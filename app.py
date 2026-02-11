"""
SmartLamppost - Sistema de Gestão de Infraestruturas RFID
API Backend Flask
Versão: 4.0.0 - Multi-Tenant com Permissões Granulares

Autor: SmartLamppost Team
Descrição: API REST para gestão de ativos de infraestrutura com tags RFID.
           
ALTERAÇÃO v4.0: Sistema Multi-Tenant
- Isolamento total entre empresas (BD separada por tenant)
- 3 níveis de utilizador: superadmin, admin, user
- Permissões granulares por secção e campo
- Autenticação 2FA (email/SMS)
- Login por email
- Catálogo partilhado entre tenants

ALTERAÇÃO v3.0: O Número de Série é agora a referência principal.
O RFID Tag é um campo opcional que pode ser alterado se danificado.
"""

import sys
import os
import sqlite3
import json
import zipfile
import random
import string
import re
import logging
from io import BytesIO
from datetime import datetime, timedelta
from functools import wraps

from flask import Flask, request, jsonify, g, send_file
from flask_cors import CORS

# Import security utilities (werkzeug-based password hashing, path validation)
from utils.security import (
    hash_password, verify_password, generate_token,
    validate_safe_path, validate_permission_action,
    VALID_PERMISSION_ACTIONS
)
from utils.database import (
    init_paths as db_init_paths,
    obter_bd, obter_bd_catalogo, obter_config,
    fechar_ligacoes, carregar_tenants, guardar_tenants,
    obter_tenant, tenant_existe, inicializar_bd_tenant,
    inicializar_catalogo, registar_auditoria,
    obter_caminho_bd_tenant, MASTER_TENANT_ID as DB_MASTER_TENANT_ID
)
from utils.email_service import (
    enviar_email_2fa, enviar_sms_2fa, enviar_email_reset_password,
    CODIGO_2FA_EXPIRACAO_MINUTOS, CODIGO_2FA_TAMANHO, MAX_TENTATIVAS_2FA
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(name)s: %(message)s'
)
logger = logging.getLogger(__name__)

# Importar openpyxl para exportação e importação Excel
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_DISPONIVEL = True
except ImportError:
    EXCEL_DISPONIVEL = False
    logger.warning("Excel features unavailable - openpyxl not installed")

# =============================================================================
# CONFIGURAÇÍO DA APLICAÇÍO
# =============================================================================

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', os.urandom(32).hex())

# Configurar CORS para permitir pedidos de qualquer origem
# CORS: restrict origins in production via CORS_ORIGINS env var (comma-separated)
_cors_origins = os.environ.get('CORS_ORIGINS', '*').split(',')
CORS(app, resources={
    r"/api/*": {
        "origins": _cors_origins,
        "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
        "allow_headers": ["Content-Type", "Authorization"]
    }
})

# Caminhos e configurações
PASTA_BASE = os.path.dirname(os.path.abspath(__file__))
db_init_paths(PASTA_BASE)
CHAVE_SECRETA = os.environ.get('SECRET_KEY', 'CHANGE-ME-IN-PRODUCTION-' + os.urandom(16).hex())
if 'CHANGE-ME-IN-PRODUCTION' in CHAVE_SECRETA:
    logger.warning("SECRET_KEY not set! Using random key - sessions will not persist across restarts")
EXPIRACAO_TOKEN_HORAS = 24

# =============================================================================
# CONFIGURAÇÕES MULTI-TENANT
# =============================================================================
PASTA_TENANTS = os.path.join(PASTA_BASE, 'tenants')
PASTA_SHARED = os.path.join(PASTA_BASE, 'shared')
PASTA_CONFIG = os.path.join(PASTA_BASE, 'config')
CATALOGO_PARTILHADO = os.path.join(PASTA_SHARED, 'catalog.db')
FICHEIRO_TENANTS = os.path.join(PASTA_CONFIG, 'tenants.json')

# Tenant master (superadmin)
MASTER_TENANT_ID = DB_MASTER_TENANT_ID

# Configurações de backup
PASTA_BACKUPS = os.path.join(PASTA_BASE, 'backups')
MAX_BACKUPS = 30

# Configurações de uploads para intervenções
PASTA_UPLOADS = os.path.join(PASTA_BASE, 'uploads')
PASTA_INTERVENCOES = os.path.join(PASTA_UPLOADS, 'interventions')
PASTA_ASSETS = os.path.join(PASTA_BASE, 'assets')
MAX_FILE_SIZE = 3 * 1024 * 1024  # 3MB
ALLOWED_EXTENSIONS = {'pdf', 'jpg', 'jpeg', 'png'}

# Configurações 2FA
CODIGO_2FA_EXPIRACAO_MINUTOS = 10
CODIGO_2FA_TAMANHO = 6
MAX_TENTATIVAS_2FA = 3

# Configurações de Email (configurar com variáveis de ambiente)
SMTP_SERVER = os.environ.get('SMTP_SERVER', 'smtp.gmail.com')
SMTP_PORT = int(os.environ.get('SMTP_PORT', '587'))
SMTP_USER = os.environ.get('SMTP_USER', '')
SMTP_PASSWORD = os.environ.get('SMTP_PASSWORD', '')
EMAIL_FROM = os.environ.get('EMAIL_FROM', 'noreply@smartlamppost.com')

# Configurações de SMS (placeholder - integrar com provider)
SMS_API_URL = os.environ.get('SMS_API_URL', '')
SMS_API_KEY = os.environ.get('SMS_API_KEY', '')

# Criar diretórios necessários
os.makedirs(PASTA_BACKUPS, exist_ok=True)
os.makedirs(PASTA_INTERVENCOES, exist_ok=True)
os.makedirs(PASTA_ASSETS, exist_ok=True)
os.makedirs(PASTA_TENANTS, exist_ok=True)
os.makedirs(PASTA_SHARED, exist_ok=True)
os.makedirs(PASTA_CONFIG, exist_ok=True)
os.makedirs(os.path.join(PASTA_TENANTS, MASTER_TENANT_ID), exist_ok=True)

# =============================================================================
# GESTÍO DA BASE DE DADOS - MULTI-TENANT
# =============================================================================

# obter_caminho_bd_tenant imported from utils.database
# obter_bd imported from utils.database

# obter_bd_catalogo imported from utils.database

# obter_config imported from utils.database

# DB connection cleanup imported from utils.database
@app.teardown_appcontext
def fechar_ligacao(excecao):
    fechar_ligacoes(excecao)

# =============================================================================
# GESTÍO DE TENANTS
# =============================================================================

# carregar_tenants imported from utils.database

# guardar_tenants imported from utils.database

# obter_tenant imported from utils.database

# tenant_existe imported from utils.database

# criar_tenant - uses utils.database
def criar_tenant(tenant_id, nome, short_name=None):
    from utils.database import carregar_tenants as _carregar, guardar_tenants as _guardar, tenant_existe as _existe
    dados = _carregar()
    if _existe(tenant_id):
        return False, 'Tenant ja existe'
    pasta_tenant = os.path.join(PASTA_TENANTS, tenant_id)
    os.makedirs(pasta_tenant, exist_ok=True)
    novo_tenant = {
        'id': tenant_id, 'name': nome,
        'short_name': short_name or tenant_id.upper()[:3],
        'is_master': tenant_id == MASTER_TENANT_ID,
        'active': True, 'created_at': datetime.now().isoformat()
    }
    dados['tenants'].append(novo_tenant)
    _guardar(dados)
    inicializar_bd_tenant(tenant_id)
    return True, novo_tenant

# =============================================================================
# FUNÇÕES DE 2FA
# =============================================================================

def gerar_codigo_2fa():
    return ''.join(random.choices(string.digits, k=CODIGO_2FA_TAMANHO))

# enviar_email_2fa imported from utils

# enviar_sms_2fa imported from utils

# enviar_email_reset_password imported from utils

# inicializar_bd_tenant imported from utils.database


# inicializar_catalogo imported from utils.database


# inicializar_bd removed - use inicializar_bd_tenant from utils.database


# =============================================================================
# AUTENTICAÇÍO
# =============================================================================

# hash_password imported from utils.security

# generate_token imported from utils.security as gerar_token
gerar_token = generate_token

def requer_autenticacao(f):
    """Decorador que exige autenticação para aceder ao endpoint."""
    @wraps(f)
    def decorado(*args, **kwargs):
        # Tentar obter token do header Authorization ou do query parameter
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        if not token:
            token = request.args.get('token', '')
        
        if not token:
            return jsonify({'error': 'Token de autenticação necessário'}), 401
        
        # Tentar obter tenant_id do header ou query
        tenant_id = request.headers.get('X-Tenant-ID', '')
        if not tenant_id:
            tenant_id = request.args.get('tenant_id', '')
        
        # Procurar sessão em todos os tenants se não especificado
        sessao = None
        tenant_encontrado = None
        
        if tenant_id:
            # Procurar apenas no tenant especificado
            try:
                bd = obter_bd(tenant_id)
                sessao = bd.execute('''
                    SELECT s.*, u.email, u.role, u.first_name, u.last_name,
                           u.must_change_password, u.two_factor_enabled
                    FROM sessions s 
                    JOIN users u ON s.user_id = u.id 
                    WHERE s.token = ? AND s.expires_at > datetime('now') AND u.active = 1
                ''', (token,)).fetchone()
                if sessao:
                    tenant_encontrado = tenant_id
            except Exception:
                pass
        else:
            # Procurar em todos os tenants
            dados_tenants = carregar_tenants()
            for tenant_info in dados_tenants.get('tenants', []):
                tid = tenant_info['id']
                try:
                    bd = obter_bd(tid)
                    sessao = bd.execute('''
                        SELECT s.*, u.email, u.role, u.first_name, u.last_name,
                               u.must_change_password, u.two_factor_enabled
                        FROM sessions s 
                        JOIN users u ON s.user_id = u.id 
                        WHERE s.token = ? AND s.expires_at > datetime('now') AND u.active = 1
                    ''', (token,)).fetchone()
                    if sessao:
                        tenant_encontrado = tid
                        break
                except Exception:
                    continue
        
        if not sessao:
            return jsonify({'error': 'Token inválido ou expirado'}), 401
        
        # Definir contexto
        g.tenant_id = tenant_encontrado
        g.utilizador_atual = dict(sessao)
        g.utilizador_atual['tenant_id'] = tenant_encontrado
        
        return f(*args, **kwargs)
    return decorado

def requer_superadmin(f):
    """Decorador que exige permissões de superadmin."""
    @wraps(f)
    @requer_autenticacao
    def decorado(*args, **kwargs):
        if g.utilizador_atual['role'] != 'superadmin':
            return jsonify({'error': 'Acesso restrito a superadministradores'}), 403
        return f(*args, **kwargs)
    return decorado

def verificar_permissao(section, action='view'):
    """Verifica se o utilizador tem permissao para uma acao numa seccao."""
    user = g.utilizador_atual

    # Superadmin e admin tem todas as permissoes
    if user['role'] in ['superadmin', 'admin']:
        return True

    # Whitelist action to prevent SQL injection
    validated_action = validate_permission_action(action)
    if validated_action is None:
        logger.warning("Invalid permission action: %s", action)
        return False

    # Verificar permissoes especificas
    bd = obter_bd()
    campo = f'can_{validated_action}'

    perm = bd.execute(f'''
        SELECT {campo} FROM user_permissions
        WHERE user_id = ? AND section = ? AND (field_name IS NULL OR field_name = '')
    ''', (user['user_id'], section)).fetchone()

    return perm and perm[campo] == 1

def requer_admin(f):
    """Decorador que exige permissões de administrador ou superior."""
    @wraps(f)
    @requer_autenticacao
    def decorado(*args, **kwargs):
        if g.utilizador_atual['role'] not in ['admin', 'superadmin']:
            return jsonify({'error': 'Acesso restrito a administradores'}), 403
        return f(*args, **kwargs)
    return decorado

# Alias for backward compatibility
requer_admin_ou_superior = requer_admin

# registar_auditoria imported from utils.database

# =============================================================================
# ENDPOINTS DE HEALTH CHECK E DEBUG
# =============================================================================

@app.route('/api/health', methods=['GET'])
def health_check():
    """Endpoint de health check para verificar se o servidor está a funcionar."""
    return jsonify({
        'status': 'ok',
        'timestamp': datetime.now().isoformat(),
        'version': '4.0.0',
        'excel_available': EXCEL_DISPONIVEL
    })


# =============================================================================
# ROTAS FRONTEND - LANDING PAGE E APP V5
# =============================================================================

# Pasta do frontend v5 compilado
V5_FRONTEND_DIST = os.path.join(PASTA_BASE, 'v5', 'frontend', 'dist')
V5_LANDING = os.path.join(PASTA_BASE, 'v5', 'frontend', 'public', 'landing')

@app.route('/', methods=['GET'])
def index():
    """Serve a landing page como página inicial."""
    landing_path = os.path.join(V5_LANDING, 'index.html')
    if os.path.exists(landing_path):
        return send_file(landing_path)
    # Fallback para o frontend v5 compilado
    return send_file(os.path.join(V5_FRONTEND_DIST, 'index.html'))

@app.route('/assets/<path:filename>', methods=['GET'])
def serve_v5_assets(filename):
    """
    Serve ficheiros estáticos do frontend v5 compilado (JS, CSS, etc).
    Se o ficheiro não existir, serve o index.html para SPA routing.
    """
    from flask import send_from_directory
    assets_path = os.path.join(V5_FRONTEND_DIST, 'assets')
    file_path = os.path.join(assets_path, filename)

    # Se é um ficheiro estático que existe, servir
    if os.path.isfile(file_path):
        return send_from_directory(assets_path, filename)

    # Caso contrário, é uma rota SPA - servir index.html
    return send_file(os.path.join(V5_FRONTEND_DIST, 'index.html'))

@app.route('/landing/<path:filename>', methods=['GET'])
def serve_landing_assets(filename):
    """Serve ficheiros estáticos da landing page."""
    from flask import send_from_directory
    return send_from_directory(V5_LANDING, filename)

@app.route('/assets', methods=['GET'], endpoint='assets_page')
def assets_page():
    """Serve a página de ativos do frontend v5."""
    return send_file(os.path.join(V5_FRONTEND_DIST, 'index.html'))

@app.route('/app', methods=['GET'])
@app.route('/app/<path:subpath>', methods=['GET'])
@app.route('/login', methods=['GET'])
@app.route('/dashboard', methods=['GET'])
@app.route('/dashboard/<path:subpath>', methods=['GET'])
@app.route('/interventions', methods=['GET'])
@app.route('/interventions/<path:subpath>', methods=['GET'])
@app.route('/catalog', methods=['GET'])
@app.route('/catalog/<path:subpath>', methods=['GET'])
@app.route('/technicians', methods=['GET'])
@app.route('/technicians/<path:subpath>', methods=['GET'])
@app.route('/reports', methods=['GET'])
@app.route('/reports/<path:subpath>', methods=['GET'])
@app.route('/custom-reports', methods=['GET'])
@app.route('/analytics', methods=['GET'])
@app.route('/data', methods=['GET'])
@app.route('/data/<path:subpath>', methods=['GET'])
@app.route('/users', methods=['GET'])
@app.route('/users/<path:subpath>', methods=['GET'])
@app.route('/tenants', methods=['GET'])
@app.route('/tenants/<path:subpath>', methods=['GET'])
@app.route('/settings', methods=['GET'])
@app.route('/settings/<path:subpath>', methods=['GET'])
@app.route('/map', methods=['GET'])
@app.route('/map/<path:subpath>', methods=['GET'])
@app.route('/scan', methods=['GET'])
@app.route('/plans', methods=['GET'])
def app_v5(subpath=None):
    """Serve o frontend v5 compilado para todas as rotas da aplicação."""
    return send_file(os.path.join(V5_FRONTEND_DIST, 'index.html'))

@app.route('/favicon.svg', methods=['GET'])
@app.route('/logo-192.svg', methods=['GET'])
@app.route('/logo-512.svg', methods=['GET'])
@app.route('/manifest.webmanifest', methods=['GET'])
@app.route('/registerSW.js', methods=['GET'])
@app.route('/sw.js', methods=['GET'])
@app.route('/workbox-4b126c97.js', methods=['GET'])
def serve_v5_static():
    """Serve ficheiros estáticos da raiz do frontend v5."""
    from flask import send_from_directory
    filename = request.path.lstrip('/')
    return send_from_directory(V5_FRONTEND_DIST, filename)


# =============================================================================
# ENDPOINTS DE AUTENTICAÇÍO
# =============================================================================


# =============================================================================
# SISTEMA DE AUTENTICAÇÍO MULTI-TENANT COM 2FA
# =============================================================================

def identificar_tenant_por_email(email):
    """Identifica o tenant a partir do email do utilizador."""
    # Primeiro, verificar no tenant master
    dados_tenants = carregar_tenants()
    
    for tenant_info in dados_tenants.get('tenants', []):
        tenant_id = tenant_info['id']
        try:
            bd = obter_bd(tenant_id)
            user = bd.execute('SELECT id FROM users WHERE email = ?', (email,)).fetchone()
            if user:
                return tenant_id
        except Exception:
            continue
    
    return None

@app.route('/api/auth/login', methods=['POST'])
def login():
    """
    Autentica um utilizador por email.
    Se 2FA estiver ativo, retorna requires_2fa=True e aguarda verificação.
    """
    dados = request.json
    email = dados.get('email', '').strip().lower()
    password = dados.get('password', '')
    tenant_id = dados.get('tenant_id', '')
    
    if not email or not password:
        return jsonify({'error': 'Email e password obrigatórios'}), 400
    
    # Identificar tenant se não especificado
    if not tenant_id:
        tenant_id = identificar_tenant_por_email(email)
        if not tenant_id:
            return jsonify({'error': 'Credenciais inválidas'}), 401
    
    # Definir tenant no contexto
    g.tenant_id = tenant_id
    bd = obter_bd(tenant_id)
    
    # Verificar utilizador
    utilizador = bd.execute(
        'SELECT * FROM users WHERE email = ?',
        (email,)
    ).fetchone()
    
    if not utilizador:
        return jsonify({'error': 'Credenciais inválidas'}), 401
    
    # Verificar se conta está ativa
    if utilizador['active'] == 0:
        return jsonify({'error': 'Conta desativada. Contacte o administrador.'}), 401
    
    # Verificar se conta está bloqueada
    if utilizador['locked_until']:
        locked_until = datetime.fromisoformat(utilizador['locked_until'])
        if locked_until > datetime.now():
            return jsonify({'error': f'Conta bloqueada até {locked_until.strftime("%H:%M")}'}), 401
        else:
            # Desbloquear
            bd.execute('UPDATE users SET locked_until = NULL, failed_login_attempts = 0 WHERE id = ?', 
                      (utilizador['id'],))
    
    # Verificar password
    if not verify_password(password, utilizador['password_hash']):
        # Incrementar tentativas falhadas
        tentativas = (utilizador['failed_login_attempts'] or 0) + 1
        if tentativas >= 5:
            # Bloquear por 15 minutos
            bloqueio = datetime.now() + timedelta(minutes=15)
            bd.execute('UPDATE users SET failed_login_attempts = ?, locked_until = ? WHERE id = ?',
                      (tentativas, bloqueio.isoformat(), utilizador['id']))
        else:
            bd.execute('UPDATE users SET failed_login_attempts = ? WHERE id = ?',
                      (tentativas, utilizador['id']))
        bd.commit()
        return jsonify({'error': 'Credenciais inválidas'}), 401
    
    # Reset tentativas falhadas
    bd.execute('UPDATE users SET failed_login_attempts = 0 WHERE id = ?', (utilizador['id'],))
    
    # Verificar se 2FA está ativo
    if utilizador['two_factor_enabled']:
        # Gerar código 2FA
        codigo = gerar_codigo_2fa()
        expira = datetime.now() + timedelta(minutes=CODIGO_2FA_EXPIRACAO_MINUTOS)
        
        # Limpar códigos antigos
        bd.execute('DELETE FROM two_factor_codes WHERE user_id = ?', (utilizador['id'],))
        
        # Guardar novo código
        bd.execute('''
            INSERT INTO two_factor_codes (user_id, code, method, expires_at)
            VALUES (?, ?, ?, ?)
        ''', (utilizador['id'], codigo, utilizador['two_factor_method'], expira.isoformat()))
        bd.commit()
        
        # Enviar código
        if utilizador['two_factor_method'] == 'sms' and utilizador['phone']:
            enviar_sms_2fa(utilizador['phone'], codigo)
        else:
            enviar_email_2fa(email, codigo, utilizador['first_name'])
        
        return jsonify({
            'requires_2fa': True,
            'method': utilizador['two_factor_method'],
            'user_id': utilizador['id'],
            'tenant_id': tenant_id,
            'message': f'Código enviado por {utilizador["two_factor_method"]}'
        })
    
    # Login direto (sem 2FA)
    return completar_login(bd, utilizador, tenant_id)

def completar_login(bd, utilizador, tenant_id):
    """Completa o processo de login após verificação."""
    # Limpar sessões antigas
    bd.execute('DELETE FROM sessions WHERE user_id = ?', (utilizador['id'],))
    
    # Criar nova sessão
    token = gerar_token()
    expira = datetime.now() + timedelta(hours=EXPIRACAO_TOKEN_HORAS)
    
    bd.execute('''
        INSERT INTO sessions (user_id, token, expires_at, ip_address, user_agent)
        VALUES (?, ?, ?, ?, ?)
    ''', (utilizador['id'], token, expira.isoformat(), 
           request.remote_addr, request.headers.get('User-Agent', '')[:200]))
    
    bd.execute('UPDATE users SET last_login = ? WHERE id = ?', (datetime.now().isoformat(), utilizador['id']))
    bd.commit()
    
    # Verificar se precisa trocar password
    must_change = utilizador['must_change_password'] == 1
    
    return jsonify({
        'token': token,
        'tenant_id': tenant_id,
        'user': {
            'id': utilizador['id'],
            'email': utilizador['email'],
            'role': utilizador['role'],
            'first_name': utilizador['first_name'],
            'last_name': utilizador['last_name'],
            'two_factor_enabled': bool(utilizador['two_factor_enabled'])
        },
        'must_change_password': must_change,
        'expires_at': expira.isoformat()
    })

@app.route('/api/auth/verify-2fa', methods=['POST'])
def verificar_2fa():
    """Verifica o código 2FA e completa o login."""
    dados = request.json
    user_id = dados.get('user_id')
    codigo = dados.get('code', '').strip()
    tenant_id = dados.get('tenant_id', MASTER_TENANT_ID)
    
    if not user_id or not codigo:
        return jsonify({'error': 'User ID e código obrigatórios'}), 400
    
    g.tenant_id = tenant_id
    bd = obter_bd(tenant_id)
    
    # Verificar código
    codigo_db = bd.execute('''
        SELECT * FROM two_factor_codes 
        WHERE user_id = ? AND code = ? AND used = 0 AND expires_at > datetime('now')
    ''', (user_id, codigo)).fetchone()
    
    if not codigo_db:
        # Verificar tentativas
        codigo_existente = bd.execute(
            'SELECT * FROM two_factor_codes WHERE user_id = ? AND used = 0',
            (user_id,)
        ).fetchone()
        
        if codigo_existente:
            tentativas = (codigo_existente['attempts'] or 0) + 1
            if tentativas >= MAX_TENTATIVAS_2FA:
                bd.execute('DELETE FROM two_factor_codes WHERE user_id = ?', (user_id,))
                bd.commit()
                return jsonify({'error': 'Demasiadas tentativas. Faça login novamente.'}), 401
            
            bd.execute('UPDATE two_factor_codes SET attempts = ? WHERE id = ?',
                      (tentativas, codigo_existente['id']))
            bd.commit()
        
        return jsonify({'error': 'Código inválido ou expirado'}), 401
    
    # Marcar código como usado
    bd.execute('UPDATE two_factor_codes SET used = 1 WHERE id = ?', (codigo_db['id'],))
    
    # Obter utilizador
    utilizador = bd.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
    
    return completar_login(bd, utilizador, tenant_id)

@app.route('/api/auth/resend-2fa', methods=['POST'])
def reenviar_2fa():
    """Reenvia o código 2FA."""
    dados = request.json
    user_id = dados.get('user_id')
    tenant_id = dados.get('tenant_id', MASTER_TENANT_ID)
    
    if not user_id:
        return jsonify({'error': 'User ID obrigatório'}), 400
    
    g.tenant_id = tenant_id
    bd = obter_bd(tenant_id)
    
    utilizador = bd.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
    if not utilizador:
        return jsonify({'error': 'Utilizador não encontrado'}), 404
    
    # Gerar novo código
    codigo = gerar_codigo_2fa()
    expira = datetime.now() + timedelta(minutes=CODIGO_2FA_EXPIRACAO_MINUTOS)
    
    # Limpar códigos antigos
    bd.execute('DELETE FROM two_factor_codes WHERE user_id = ?', (user_id,))
    
    # Guardar novo código
    bd.execute('''
        INSERT INTO two_factor_codes (user_id, code, method, expires_at)
        VALUES (?, ?, ?, ?)
    ''', (user_id, codigo, utilizador['two_factor_method'], expira.isoformat()))
    bd.commit()
    
    # Enviar código
    if utilizador['two_factor_method'] == 'sms' and utilizador['phone']:
        enviar_sms_2fa(utilizador['phone'], codigo)
    else:
        enviar_email_2fa(utilizador['email'], codigo, utilizador['first_name'])
    
    return jsonify({'message': 'Código reenviado com sucesso'})

@app.route('/api/auth/forgot-password', methods=['POST'])
def esqueci_password():
    """Inicia processo de recuperação de password."""
    dados = request.json
    email = dados.get('email', '').strip().lower()
    
    if not email:
        return jsonify({'error': 'Email obrigatório'}), 400
    
    # Encontrar tenant do utilizador
    tenant_id = identificar_tenant_por_email(email)
    if not tenant_id:
        # Não revelar se email existe ou não
        return jsonify({'message': 'Se o email existir, receberá instruções para redefinir a password.'})
    
    bd = obter_bd(tenant_id)
    utilizador = bd.execute('SELECT * FROM users WHERE email = ?', (email,)).fetchone()
    
    if utilizador:
        # Gerar token
        token = ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
        expira = datetime.now() + timedelta(minutes=30)
        
        # Limpar tokens antigos
        bd.execute('DELETE FROM password_reset_tokens WHERE user_id = ?', (utilizador['id'],))
        
        # Guardar token
        bd.execute('''
            INSERT INTO password_reset_tokens (user_id, token, expires_at)
            VALUES (?, ?, ?)
        ''', (utilizador['id'], token, expira.isoformat()))
        bd.commit()
        
        # Enviar email
        enviar_email_reset_password(email, token, utilizador['first_name'])
    
    return jsonify({'message': 'Se o email existir, receberá instruções para redefinir a password.'})

@app.route('/api/auth/reset-password', methods=['POST'])
def reset_password():
    """Redefine a password usando o token."""
    dados = request.json
    email = dados.get('email', '').strip().lower()
    token = dados.get('token', '').strip().upper()
    nova_password = dados.get('new_password', '')
    
    if not email or not token or not nova_password:
        return jsonify({'error': 'Email, token e nova password obrigatórios'}), 400
    
    if len(nova_password) < 8:
        return jsonify({'error': 'Password deve ter pelo menos 8 caracteres'}), 400
    
    # Encontrar tenant
    tenant_id = identificar_tenant_por_email(email)
    if not tenant_id:
        return jsonify({'error': 'Token inválido ou expirado'}), 401
    
    bd = obter_bd(tenant_id)
    
    # Verificar token
    resultado = bd.execute('''
        SELECT prt.*, u.id as user_id FROM password_reset_tokens prt
        JOIN users u ON prt.user_id = u.id
        WHERE u.email = ? AND prt.token = ? AND prt.used = 0 AND prt.expires_at > datetime('now')
    ''', (email, token)).fetchone()
    
    if not resultado:
        return jsonify({'error': 'Token inválido ou expirado'}), 401
    
    # Atualizar password
    bd.execute('''
        UPDATE users SET password_hash = ?, must_change_password = 0 WHERE id = ?
    ''', (hash_password(nova_password), resultado['user_id']))
    
    # Marcar token como usado
    bd.execute('UPDATE password_reset_tokens SET used = 1 WHERE id = ?', (resultado['id'],))
    
    # Limpar todas as sessões
    bd.execute('DELETE FROM sessions WHERE user_id = ?', (resultado['user_id'],))
    
    bd.commit()
    
    return jsonify({'message': 'Password alterada com sucesso'})

@app.route('/api/auth/change-password', methods=['POST'])
@requer_autenticacao
def alterar_password():
    """Altera a password do utilizador autenticado."""
    dados = request.json
    password_atual = dados.get('current_password', '')
    nova_password = dados.get('new_password', '')
    
    if not password_atual or not nova_password:
        return jsonify({'error': 'Password atual e nova password obrigatórias'}), 400
    
    if len(nova_password) < 8:
        return jsonify({'error': 'Password deve ter pelo menos 8 caracteres'}), 400
    
    bd = obter_bd()
    utilizador = bd.execute('SELECT * FROM users WHERE id = ?', 
                           (g.utilizador_atual['user_id'],)).fetchone()
    
    if not verify_password(password_atual, utilizador['password_hash']):
        return jsonify({'error': 'Password atual incorreta'}), 401
    
    bd.execute('''
        UPDATE users SET password_hash = ?, must_change_password = 0 WHERE id = ?
    ''', (hash_password(nova_password), utilizador['id']))
    bd.commit()
    
    return jsonify({'message': 'Password alterada com sucesso'})

@app.route('/api/auth/logout', methods=['POST'])
@requer_autenticacao
def logout():
    """Termina a sessão do utilizador."""
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    bd = obter_bd()
    bd.execute('DELETE FROM sessions WHERE token = ?', (token,))
    bd.commit()
    return jsonify({'message': 'Sessão terminada'})

@app.route('/api/auth/me', methods=['GET'])
@requer_autenticacao
def utilizador_atual():
    """Devolve informação do utilizador autenticado."""
    return jsonify({
        'id': g.utilizador_atual['user_id'],
        'email': g.utilizador_atual.get('email', ''),
        'role': g.utilizador_atual['role'],
        'first_name': g.utilizador_atual.get('first_name', ''),
        'last_name': g.utilizador_atual.get('last_name', ''),
        'tenant_id': g.utilizador_atual.get('tenant_id', MASTER_TENANT_ID),
        'must_change_password': bool(g.utilizador_atual.get('must_change_password', 0)),
        'two_factor_enabled': bool(g.utilizador_atual.get('two_factor_enabled', 0))
    })



# =============================================================================
# GESTÍO DE TENANTS (APENAS SUPERADMIN)
# =============================================================================

@app.route('/api/tenants', methods=['GET'])
@requer_superadmin
def listar_tenants():
    """Lista todos os tenants."""
    dados = carregar_tenants()
    tenants = []
    
    for tenant in dados.get('tenants', []):
        # Contar utilizadores e ativos de cada tenant
        try:
            bd = obter_bd(tenant['id'])
            user_count = bd.execute('SELECT COUNT(*) FROM users').fetchone()[0]
            asset_count = bd.execute('SELECT COUNT(*) FROM assets').fetchone()[0]
        except Exception:
            user_count = 0
            asset_count = 0
        
        tenants.append({
            **tenant,
            'user_count': user_count,
            'asset_count': asset_count
        })
    
    return jsonify(tenants)

@app.route('/api/tenants/<string:tenant_id>', methods=['GET'])
@requer_superadmin
def obter_tenant_info(tenant_id):
    """Obtém informação detalhada de um tenant."""
    tenant = obter_tenant(tenant_id)
    if not tenant:
        return jsonify({'error': 'Tenant não encontrado'}), 404
    
    # Estatísticas
    try:
        bd = obter_bd(tenant_id)
        stats = {
            'users': bd.execute('SELECT COUNT(*) FROM users').fetchone()[0],
            'assets': bd.execute('SELECT COUNT(*) FROM assets').fetchone()[0],
            'interventions': bd.execute('SELECT COUNT(*) FROM interventions').fetchone()[0],
        }
    except Exception:
        stats = {'users': 0, 'assets': 0, 'interventions': 0}
    
    return jsonify({**tenant, 'stats': stats})

@app.route('/api/tenants', methods=['POST'])
@requer_superadmin
def criar_novo_tenant():
    """Cria um novo tenant."""
    dados = request.json
    tenant_id = dados.get('id', '').strip().lower()
    nome = dados.get('name', '').strip()
    short_name = dados.get('short_name', '').strip()
    
    if not tenant_id or not nome:
        return jsonify({'error': 'ID e nome obrigatórios'}), 400
    
    # Validar ID (apenas letras, números e hífens)
    if not re.match(r'^[a-z0-9-]+$', tenant_id):
        return jsonify({'error': 'ID deve conter apenas letras minúsculas, números e hífens'}), 400
    
    sucesso, resultado = criar_tenant(tenant_id, nome, short_name)
    
    if not sucesso:
        return jsonify({'error': resultado}), 400
    
    # Criar admin inicial para o tenant
    admin_email = dados.get('admin_email', f'admin@{tenant_id}.com')
    admin_password = dados.get('admin_password', 'admin123')
    
    bd = obter_bd(tenant_id)
    hash_pwd = hash_password(admin_password)
    
    bd.execute('''
        INSERT INTO users (email, password_hash, role, first_name, must_change_password, active)
        VALUES (?, ?, 'admin', 'Administrador', 1, 1)
    ''', (admin_email, hash_pwd))
    bd.commit()
    
    return jsonify({
        'message': 'Tenant criado com sucesso',
        'tenant': resultado,
        'admin_email': admin_email
    }), 201

@app.route('/api/tenants/<string:tenant_id>', methods=['PUT'])
@requer_superadmin
def atualizar_tenant(tenant_id):
    """Atualiza informação de um tenant."""
    tenant = obter_tenant(tenant_id)
    if not tenant:
        return jsonify({'error': 'Tenant não encontrado'}), 404
    
    dados = request.json
    dados_tenants = carregar_tenants()
    
    for i, t in enumerate(dados_tenants['tenants']):
        if t['id'] == tenant_id:
            dados_tenants['tenants'][i]['name'] = dados.get('name', t['name'])
            dados_tenants['tenants'][i]['short_name'] = dados.get('short_name', t['short_name'])
            dados_tenants['tenants'][i]['active'] = dados.get('active', t.get('active', True))
            break
    
    guardar_tenants(dados_tenants)
    return jsonify({'message': 'Tenant atualizado'})

@app.route('/api/tenants/<string:tenant_id>/logo', methods=['POST'])
@requer_superadmin
def upload_logo_tenant(tenant_id):
    """Upload do logo de um tenant."""
    if 'logo' not in request.files:
        return jsonify({'error': 'Ficheiro de logo não fornecido'}), 400
    
    tenant = obter_tenant(tenant_id)
    if not tenant:
        return jsonify({'error': 'Tenant não encontrado'}), 404
    
    ficheiro = request.files['logo']
    if ficheiro.filename == '':
        return jsonify({'error': 'Nenhum ficheiro selecionado'}), 400
    
    # Verificar extensão
    ext = ficheiro.filename.rsplit('.', 1)[-1].lower()
    if ext not in ['png', 'jpg', 'jpeg']:
        return jsonify({'error': 'Formato inválido. Use PNG ou JPG'}), 400
    
    # Guardar
    pasta_tenant = os.path.join(PASTA_TENANTS, tenant_id)
    os.makedirs(pasta_tenant, exist_ok=True)
    caminho = os.path.join(pasta_tenant, f'logo.{ext}')
    ficheiro.save(caminho)
    
    return jsonify({'message': 'Logo atualizado', 'path': f'/api/tenants/{tenant_id}/logo'})

@app.route('/api/tenants/<string:tenant_id>/logo', methods=['GET'])
def obter_logo_tenant(tenant_id):
    """Devolve o logo de um tenant."""
    from flask import make_response
    pasta_tenant = os.path.join(PASTA_TENANTS, tenant_id)

    for ext in ['png', 'jpg', 'jpeg']:
        caminho = os.path.join(pasta_tenant, f'logo.{ext}')
        if os.path.exists(caminho):
            resp = make_response(send_file(caminho))
            resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            return resp

    # Logo padrão (master tenant)
    for ext in ['png', 'jpg', 'jpeg']:
        logo_padrao = os.path.join(PASTA_TENANTS, MASTER_TENANT_ID, f'logo.{ext}')
        if os.path.exists(logo_padrao):
            resp = make_response(send_file(logo_padrao))
            resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            return resp

    return jsonify({'error': 'Logo não encontrado'}), 404

# =============================================================================
# GESTÃO DE UTILIZADORES POR TENANT (SUPERADMIN)
# =============================================================================

@app.route('/api/tenants/<string:tenant_id>/users', methods=['GET'])
@requer_superadmin
def listar_utilizadores_tenant(tenant_id):
    """Lista utilizadores de um tenant específico (apenas superadmin)."""
    tenant = obter_tenant(tenant_id)
    if not tenant:
        return jsonify({'error': 'Tenant não encontrado'}), 404
    
    bd = obter_bd(tenant_id)
    utilizadores = bd.execute('''
        SELECT id, email, role, first_name, last_name, phone, active, created_at, last_login 
        FROM users ORDER BY created_at DESC
    ''').fetchall()
    
    return jsonify({
        'tenant': tenant,
        'users': [dict(u) for u in utilizadores]
    })

@app.route('/api/tenants/<string:tenant_id>/users', methods=['POST'])
@requer_superadmin
def criar_utilizador_tenant(tenant_id):
    """Cria utilizador num tenant específico (apenas superadmin)."""
    tenant = obter_tenant(tenant_id)
    if not tenant:
        return jsonify({'error': 'Tenant não encontrado'}), 404
    
    dados = request.json
    email = dados.get('email', '').strip().lower()
    password = dados.get('password')
    role = dados.get('role', 'user')
    first_name = dados.get('first_name', '').strip()
    last_name = dados.get('last_name', '').strip()
    phone = dados.get('phone', '').strip()
    
    if not email or not password:
        return jsonify({'error': 'Email e password obrigatórios'}), 400
    
    if not first_name:
        return jsonify({'error': 'Nome obrigatório'}), 400
    
    # Roles válidos (superadmin só pode ser criado no tenant master)
    valid_roles = ['admin', 'user', 'operator', 'visitor']
    if tenant_id == MASTER_TENANT_ID:
        valid_roles.append('superadmin')
    
    if role not in valid_roles:
        return jsonify({'error': f'Role inválido para este tenant. Opções: {", ".join(valid_roles)}'}), 400
    
    bd = obter_bd(tenant_id)
    try:
        bd.execute('''
            INSERT INTO users (email, password_hash, role, first_name, last_name, phone, active, must_change_password)
            VALUES (?, ?, ?, ?, ?, ?, 1, 1)
        ''', (email, hash_password(password), role, first_name, last_name, phone))
        bd.commit()
        return jsonify({'message': 'Utilizador criado com sucesso'}), 201
    except sqlite3.IntegrityError:
        return jsonify({'error': 'Email já existe neste tenant'}), 400

@app.route('/api/tenants/<string:tenant_id>/users/<int:user_id>', methods=['PUT'])
@requer_superadmin
def atualizar_utilizador_tenant(tenant_id, user_id):
    """Atualiza utilizador de um tenant específico (apenas superadmin)."""
    tenant = obter_tenant(tenant_id)
    if not tenant:
        return jsonify({'error': 'Tenant não encontrado'}), 404
    
    bd = obter_bd(tenant_id)
    user = bd.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
    if not user:
        return jsonify({'error': 'Utilizador não encontrado'}), 404
    
    dados = request.json
    email = dados.get('email', user['email']).strip().lower()
    role = dados.get('role', user['role'])
    first_name = dados.get('first_name', user['first_name'])
    last_name = dados.get('last_name', user['last_name'])
    phone = dados.get('phone', user['phone'])
    active = dados.get('active', user['active'])
    
    # Roles válidos
    valid_roles = ['admin', 'user', 'operator', 'visitor']
    if tenant_id == MASTER_TENANT_ID:
        valid_roles.append('superadmin')
    
    if role not in valid_roles:
        return jsonify({'error': f'Role inválido para este tenant'}), 400
    
    try:
        bd.execute('''
            UPDATE users SET email = ?, role = ?, first_name = ?, last_name = ?, phone = ?, active = ?
            WHERE id = ?
        ''', (email, role, first_name, last_name, phone, active, user_id))
        
        # Se foi fornecida nova password
        if dados.get('password'):
            bd.execute('UPDATE users SET password_hash = ? WHERE id = ?', 
                      (hash_password(dados['password']), user_id))
        
        bd.commit()
        return jsonify({'message': 'Utilizador atualizado'})
    except sqlite3.IntegrityError:
        return jsonify({'error': 'Email já existe neste tenant'}), 400

@app.route('/api/tenants/<string:tenant_id>/users/<int:user_id>', methods=['DELETE'])
@requer_superadmin
def eliminar_utilizador_tenant(tenant_id, user_id):
    """Elimina utilizador de um tenant específico (apenas superadmin)."""
    tenant = obter_tenant(tenant_id)
    if not tenant:
        return jsonify({'error': 'Tenant não encontrado'}), 404
    
    bd = obter_bd(tenant_id)
    user = bd.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
    if not user:
        return jsonify({'error': 'Utilizador não encontrado'}), 404
    
    # Não permitir eliminar o último admin/superadmin do tenant
    admin_count = bd.execute('''
        SELECT COUNT(*) FROM users WHERE role IN ('admin', 'superadmin') AND id != ?
    ''', (user_id,)).fetchone()[0]
    
    if user['role'] in ['admin', 'superadmin'] and admin_count == 0:
        return jsonify({'error': 'Não pode eliminar o último administrador do tenant'}), 400
    
    bd.execute('DELETE FROM sessions WHERE user_id = ?', (user_id,))
    bd.execute('DELETE FROM users WHERE id = ?', (user_id,))
    bd.commit()
    
    return jsonify({'message': 'Utilizador eliminado'})

# =============================================================================
# GESTÃO DE PERMISSÕES DE UTILIZADORES
# =============================================================================

@app.route('/api/users/<int:user_id>/permissions', methods=['GET'])
@requer_admin
def obter_permissoes_user(user_id):
    """Obtém permissões de um utilizador."""
    bd = obter_bd()
    
    # Verificar se utilizador existe
    user = bd.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
    if not user:
        return jsonify({'error': 'Utilizador não encontrado'}), 404
    
    # Se admin ou superadmin, tem todas as permissões
    if user['role'] in ['admin', 'superadmin']:
        return jsonify({
            'user_id': user_id,
            'role': user['role'],
            'full_access': True,
            'permissions': []
        })
    
    # Obter permissões específicas
    perms = bd.execute('''
        SELECT * FROM user_permissions WHERE user_id = ? ORDER BY section
    ''', (user_id,)).fetchall()
    
    return jsonify({
        'user_id': user_id,
        'role': user['role'],
        'full_access': False,
        'permissions': [dict(p) for p in perms]
    })

@app.route('/api/users/<int:user_id>/permissions', methods=['PUT'])
@requer_admin
def definir_permissoes_user(user_id):
    """Define permissões de um utilizador."""
    bd = obter_bd()
    
    # Verificar se utilizador existe e não é admin/superadmin
    user = bd.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
    if not user:
        return jsonify({'error': 'Utilizador não encontrado'}), 404
    
    if user['role'] in ['admin', 'superadmin']:
        return jsonify({'error': 'Não é possível definir permissões para admins'}), 400
    
    dados = request.json
    permissions = dados.get('permissions', [])
    
    # Limpar permissões existentes
    bd.execute('DELETE FROM user_permissions WHERE user_id = ?', (user_id,))
    
    # Inserir novas permissões
    for perm in permissions:
        bd.execute('''
            INSERT INTO user_permissions (user_id, section, field_name, can_view, can_create, can_edit, can_delete)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (
            user_id,
            perm.get('section'),
            perm.get('field_name'),
            1 if perm.get('can_view', True) else 0,
            1 if perm.get('can_create', False) else 0,
            1 if perm.get('can_edit', False) else 0,
            1 if perm.get('can_delete', False) else 0
        ))
    
    bd.commit()
    return jsonify({'message': 'Permissões atualizadas'})

@app.route('/api/permission-sections', methods=['GET'])
@requer_autenticacao
def listar_seccoes_permissao():
    """Lista todas as secções disponíveis para permissões."""
    seccoes = [
        {'id': 'dashboard', 'label': 'Dashboard', 'description': 'Visualização de estatísticas'},
        {'id': 'assets', 'label': 'Ativos', 'description': 'Gestão de ativos/infraestruturas'},
        {'id': 'interventions', 'label': 'Intervenções', 'description': 'Gestão de intervenções'},
        {'id': 'catalog', 'label': 'Catálogo', 'description': 'Consulta do catálogo de produtos'},
        {'id': 'reports', 'label': 'Relatórios', 'description': 'Geração de relatórios'},
        {'id': 'users', 'label': 'Utilizadores', 'description': 'Gestão de utilizadores (admin)'},
        {'id': 'settings', 'label': 'Configurações', 'description': 'Configurações do sistema (admin)'},
    ]
    return jsonify(seccoes)

# =============================================================================
# GESTÍO DE UTILIZADORES (ATUALIZADO PARA EMAIL)
# =============================================================================

@app.route('/api/users/<int:user_id>/force-password-reset', methods=['POST'])
@requer_admin
def forcar_reset_password(user_id):
    """Força um utilizador a redefinir a password no próximo login."""
    bd = obter_bd()
    
    user = bd.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
    if not user:
        return jsonify({'error': 'Utilizador não encontrado'}), 404
    
    # Não permitir em superadmins se não for superadmin
    if user['role'] == 'superadmin' and g.utilizador_atual['role'] != 'superadmin':
        return jsonify({'error': 'Sem permissão'}), 403
    
    # Gerar token de reset
    token = ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
    expira = datetime.now() + timedelta(hours=24)
    
    bd.execute('DELETE FROM password_reset_tokens WHERE user_id = ?', (user_id,))
    bd.execute('''
        INSERT INTO password_reset_tokens (user_id, token, expires_at)
        VALUES (?, ?, ?)
    ''', (user_id, token, expira.isoformat()))
    
    bd.execute('UPDATE users SET must_change_password = 1 WHERE id = ?', (user_id,))
    bd.commit()
    
    # Enviar email
    enviar_email_reset_password(user['email'], token, user['first_name'])
    
    return jsonify({'message': 'Email de reset enviado ao utilizador'})

@app.route('/api/users/<int:user_id>/toggle-2fa', methods=['POST'])
@requer_admin
def toggle_2fa_user(user_id):
    """Ativa/desativa 2FA para um utilizador."""
    bd = obter_bd()
    dados = request.json
    
    user = bd.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
    if not user:
        return jsonify({'error': 'Utilizador não encontrado'}), 404
    
    enabled = 1 if dados.get('enabled', False) else 0
    method = dados.get('method', 'email')
    
    if method not in ['email', 'sms']:
        return jsonify({'error': 'Método inválido. Use email ou sms'}), 400
    
    bd.execute('''
        UPDATE users SET two_factor_enabled = ?, two_factor_method = ? WHERE id = ?
    ''', (enabled, method, user_id))
    bd.commit()
    
    return jsonify({
        'message': f'2FA {"ativado" if enabled else "desativado"}',
        'two_factor_enabled': bool(enabled),
        'two_factor_method': method
    })


# =============================================================================
# GESTÍO DE UTILIZADORES
# =============================================================================

@app.route('/api/users', methods=['GET'])
@requer_admin
def listar_utilizadores():
    """Lista todos os utilizadores do sistema."""
    bd = obter_bd()
    utilizadores = bd.execute('SELECT id, email, role, first_name, last_name, phone, active, created_at, last_login FROM users').fetchall()
    return jsonify([dict(u) for u in utilizadores])

@app.route('/api/users', methods=['POST'])
@requer_admin
def criar_utilizador():
    """Cria um novo utilizador."""
    dados = request.json
    password = dados.get('password')
    role = dados.get('role', 'user')
    first_name = dados.get('first_name')
    last_name = dados.get('last_name')
    email = dados.get('email', '').strip().lower()
    
    if not email or not password:
        return jsonify({'error': 'Email e password obrigatórios'}), 400
    
    if not first_name or not last_name:
        return jsonify({'error': 'Nome e apelido são obrigatórios'}), 400
    
    # Roles válidos no sistema multi-tenant
    valid_roles = ['superadmin', 'admin', 'user', 'operator', 'visitor']
    if role not in valid_roles:
        return jsonify({'error': f'Role inválido. Deve ser: {", ".join(valid_roles)}'}), 400
    
    # Apenas superadmin pode criar outros superadmins
    if role == 'superadmin':
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        user_info = obter_utilizador_por_token(token)
        if not user_info or user_info.get('role') != 'superadmin':
            return jsonify({'error': 'Apenas superadmin pode criar outros superadmins'}), 403
    
    bd = obter_bd()
    try:
        bd.execute('''
            INSERT INTO users (email, password_hash, role, first_name, last_name, must_change_password, active) 
            VALUES (?, ?, ?, ?, ?, 1, 1)
        ''', (email, hash_password(password), role, first_name, last_name))
        bd.commit()
        return jsonify({'message': 'Utilizador criado com sucesso'}), 201
    except sqlite3.IntegrityError:
        return jsonify({'error': 'Email já existe'}), 400

@app.route('/api/users/<int:user_id>', methods=['PUT'])
@requer_admin
def actualizar_utilizador(user_id):
    """Actualiza um utilizador existente."""
    dados = request.json
    bd = obter_bd()
    
    user = bd.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
    if not user:
        return jsonify({'error': 'Utilizador não encontrado'}), 404
    
    # Campos actualizáveis
    first_name = dados.get('first_name', user['first_name'])
    last_name = dados.get('last_name', user['last_name'])
    email = dados.get('email', user['email'])
    role = dados.get('role', user['role'])
    
    # Roles válidos no sistema multi-tenant
    valid_roles = ['superadmin', 'admin', 'user', 'operator', 'visitor']
    if role not in valid_roles:
        return jsonify({'error': f'Role inválido. Deve ser: {", ".join(valid_roles)}'}), 400
    
    # Apenas superadmin pode promover a superadmin
    if role == 'superadmin' and user['role'] != 'superadmin':
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        user_info = obter_utilizador_por_token(token)
        if not user_info or user_info.get('role') != 'superadmin':
            return jsonify({'error': 'Apenas superadmin pode promover a superadmin'}), 403
    
    bd.execute('''
        UPDATE users SET first_name = ?, last_name = ?, email = ?, role = ? WHERE id = ?
    ''', (first_name, last_name, email, role, user_id))
    
    # Se foi fornecida nova password
    if dados.get('password'):
        bd.execute('UPDATE users SET password_hash = ? WHERE id = ?', (hash_password(dados['password']), user_id))
    
    bd.commit()
    return jsonify({'message': 'Utilizador actualizado'})

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@requer_admin
def eliminar_utilizador(user_id):
    """Elimina um utilizador."""
    if user_id == g.utilizador_atual['user_id']:
        return jsonify({'error': 'Não pode eliminar a própria conta'}), 400
    
    bd = obter_bd()
    bd.execute('DELETE FROM sessions WHERE user_id = ?', (user_id,))
    bd.execute('DELETE FROM users WHERE id = ?', (user_id,))
    bd.commit()
    return jsonify({'message': 'Utilizador eliminado'})

# =============================================================================
# GESTÍO DO ESQUEMA (CAMPOS DINÍ‚MICOS)
# =============================================================================

@app.route('/api/schema', methods=['GET'])
@requer_autenticacao
def obter_esquema():
    """Devolve a lista de campos do esquema."""
    bd = obter_bd()
    campos = bd.execute('SELECT * FROM schema_fields ORDER BY field_order').fetchall()
    
    resultado = []
    for c in campos:
        campo_dict = dict(c)
        if campo_dict['field_options']:
            try:
                # Tentar parse JSON primeiro
                campo_dict['field_options'] = json.loads(campo_dict['field_options'])
            except (json.JSONDecodeError, TypeError):
                # Se não for JSON, assumir string separada por vírgulas
                campo_dict['field_options'] = [o.strip() for o in campo_dict['field_options'].split(',')]
        resultado.append(campo_dict)
    
    return jsonify(resultado)

@app.route('/api/schema', methods=['POST'])
@requer_admin
def adicionar_campo():
    """Adiciona um novo campo ao esquema."""
    dados = request.json
    campos_obrigatorios = ['field_name', 'field_type', 'field_label', 'required']
    
    if not all(c in dados for c in campos_obrigatorios):
        return jsonify({'error': 'Campos obrigatórios: field_name, field_type, field_label, required'}), 400
    
    if not dados['field_name'].replace('_', '').isalnum():
        return jsonify({'error': 'field_name deve conter apenas letras, números e underscores'}), 400
    
    bd = obter_bd()
    max_ordem = bd.execute('SELECT MAX(field_order) FROM schema_fields').fetchone()[0] or 0
    opcoes = json.dumps(dados.get('field_options')) if dados.get('field_options') else None
    
    try:
        cursor = bd.execute('''
            INSERT INTO schema_fields (field_name, field_type, field_label, required, field_order, field_category, field_options)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (
            dados['field_name'], dados['field_type'], dados['field_label'],
            dados.get('required', 0), dados.get('field_order', max_ordem + 1),
            dados.get('field_category', 'custom'), opcoes
        ))
        bd.commit()
        registar_auditoria(bd, g.utilizador_atual['user_id'], 'CREATE', 'schema_fields', cursor.lastrowid, None, dados)
        return jsonify({'message': 'Campo adicionado com sucesso'}), 201
    except sqlite3.IntegrityError:
        return jsonify({'error': 'Campo já existe'}), 400

@app.route('/api/schema/<int:field_id>', methods=['PUT'])
@requer_admin
def atualizar_campo(field_id):
    """Atualiza um campo do esquema."""
    dados = request.json
    bd = obter_bd()
    
    campo_antigo = bd.execute('SELECT * FROM schema_fields WHERE id = ?', (field_id,)).fetchone()
    if not campo_antigo:
        return jsonify({'error': 'Campo não encontrado'}), 404
    
    opcoes = json.dumps(dados.get('field_options')) if dados.get('field_options') else None
    
    bd.execute('''
        UPDATE schema_fields 
        SET field_label = ?, required = ?, field_order = ?, field_category = ?, field_options = ?
        WHERE id = ?
    ''', (
        dados.get('field_label', campo_antigo['field_label']),
        dados.get('required', campo_antigo['required']),
        dados.get('field_order', campo_antigo['field_order']),
        dados.get('field_category', campo_antigo['field_category']),
        opcoes, field_id
    ))
    bd.commit()
    registar_auditoria(bd, g.utilizador_atual['user_id'], 'UPDATE', 'schema_fields', field_id, dict(campo_antigo), dados)
    return jsonify({'message': 'Campo atualizado'})

@app.route('/api/schema/<int:field_id>', methods=['DELETE'])
@requer_admin
def eliminar_campo(field_id):
    """Elimina um campo do esquema e os dados associados."""
    bd = obter_bd()
    
    campo = bd.execute('SELECT * FROM schema_fields WHERE id = ?', (field_id,)).fetchone()
    if not campo:
        return jsonify({'error': 'Campo não encontrado'}), 404
    
    bd.execute('DELETE FROM asset_data WHERE field_name = ?', (campo['field_name'],))
    bd.execute('DELETE FROM schema_fields WHERE id = ?', (field_id,))
    bd.commit()
    registar_auditoria(bd, g.utilizador_atual['user_id'], 'DELETE', 'schema_fields', field_id, dict(campo), None)
    return jsonify({'message': 'Campo eliminado'})

# =============================================================================
# GESTÍO DE ATIVOS RFID
# =============================================================================

@app.route('/api/assets', methods=['GET'])
@requer_autenticacao
def listar_ativos():
    """Lista todos os ativos com paginação e pesquisa."""
    bd = obter_bd()
    
    pagina = request.args.get('page', 1, type=int)
    por_pagina = request.args.get('per_page', 50, type=int)
    offset = (pagina - 1) * por_pagina
    pesquisa = request.args.get('search', '')
    
    if pesquisa:
        total = bd.execute('''
            SELECT COUNT(DISTINCT a.id) FROM assets a
            LEFT JOIN asset_data ad ON a.id = ad.asset_id
            WHERE a.serial_number LIKE ? OR ad.field_value LIKE ?
        ''', (f'%{pesquisa}%', f'%{pesquisa}%')).fetchone()[0]
        
        ativos = bd.execute('''
            SELECT DISTINCT a.* FROM assets a
            LEFT JOIN asset_data ad ON a.id = ad.asset_id
            WHERE a.serial_number LIKE ? OR ad.field_value LIKE ?
            ORDER BY a.updated_at DESC LIMIT ? OFFSET ?
        ''', (f'%{pesquisa}%', f'%{pesquisa}%', por_pagina, offset)).fetchall()
    else:
        total = bd.execute('SELECT COUNT(*) FROM assets').fetchone()[0]
        ativos = bd.execute('''
            SELECT * FROM assets ORDER BY updated_at DESC LIMIT ? OFFSET ?
        ''', (por_pagina, offset)).fetchall()
    
    resultado = []
    for ativo in ativos:
        ativo_dict = dict(ativo)
        dados = bd.execute('SELECT field_name, field_value FROM asset_data WHERE asset_id = ?', (ativo['id'],)).fetchall()
        ativo_dict['data'] = {d['field_name']: d['field_value'] for d in dados}
        resultado.append(ativo_dict)
    
    return jsonify({
        'assets': resultado,
        'total': total,
        'page': pagina,
        'per_page': por_pagina,
        'pages': (total + por_pagina - 1) // por_pagina
    })

@app.route('/api/assets/<string:serial_number>', methods=['GET'])
@requer_autenticacao
def obter_ativo(serial_number):
    """Devolve os detalhes de um ativo específico pelo número de série."""
    bd = obter_bd()
    ativo = bd.execute('SELECT * FROM assets WHERE serial_number = ?', (serial_number,)).fetchone()
    
    if not ativo:
        return jsonify({'error': 'Ativo não encontrado'}), 404
    
    ativo_dict = dict(ativo)
    dados = bd.execute('SELECT field_name, field_value FROM asset_data WHERE asset_id = ?', (ativo['id'],)).fetchall()
    ativo_dict['data'] = {d['field_name']: d['field_value'] for d in dados}
    
    # Histórico de manutenção
    try:
        manutencoes = bd.execute('''
            SELECT m.*, u.first_name || ' ' || COALESCE(u.last_name, '') as performed_by_name
            FROM maintenance_log m LEFT JOIN users u ON m.performed_by = u.id
            WHERE m.asset_id = ? ORDER BY m.performed_at DESC
        ''', (ativo['id'],)).fetchall()
        ativo_dict['maintenance_history'] = [dict(m) for m in manutencoes]
    except Exception:
        ativo_dict['maintenance_history'] = []
    
    # Histórico de alterações de estado
    try:
        alteracoes_estado = bd.execute('''
            SELECT s.*, u.first_name || ' ' || COALESCE(u.last_name, '') as changed_by_name
            FROM status_history s LEFT JOIN users u ON s.changed_by = u.id
            WHERE s.asset_id = ? ORDER BY s.changed_at DESC
        ''', (ativo['id'],)).fetchall()
        ativo_dict['status_history'] = [dict(s) for s in alteracoes_estado]
    except Exception:
        ativo_dict['status_history'] = []
    
    # Intervenções
    try:
        intervencoes = bd.execute('''
            SELECT i.*, u.first_name || ' ' || COALESCE(u.last_name, '') as created_by_name
            FROM interventions i 
            LEFT JOIN users u ON i.created_by = u.id
            WHERE i.asset_serial_number = ? 
            ORDER BY i.created_at DESC
        ''', (serial_number,)).fetchall()
        ativo_dict['interventions'] = [dict(i) for i in intervencoes]
    except Exception:
        ativo_dict['interventions'] = []
    
    return jsonify(ativo_dict)

@app.route('/api/assets', methods=['POST'])
@requer_autenticacao
def criar_ativo():
    """Cria um novo ativo. O serial_number é obrigatório e único."""
    dados = request.json
    serial_number = dados.get('serial_number')
    
    if not serial_number:
        return jsonify({'error': 'Número de Série obrigatório'}), 400
    
    bd = obter_bd()
    
    # Validar campos obrigatórios
    campos_obrigatorios = bd.execute('SELECT field_name, field_label FROM schema_fields WHERE required = 1').fetchall()
    dados_ativo = dados.get('data', {})
    
    # Definir valor padrão para status se não fornecido
    if 'status' not in dados_ativo or not dados_ativo['status']:
        dados_ativo['status'] = 'Operacional'
    
    campos_em_falta = []
    for campo in campos_obrigatorios:
        if campo['field_name'] not in dados_ativo or not dados_ativo[campo['field_name']]:
            campos_em_falta.append(campo['field_label'])
    
    if campos_em_falta:
        return jsonify({'error': f'Campos obrigatórios em falta: {", ".join(campos_em_falta)}'}), 400
    
    try:
        cursor = bd.execute('''
            INSERT INTO assets (serial_number, created_by, updated_by) VALUES (?, ?, ?)
        ''', (serial_number, g.utilizador_atual['user_id'], g.utilizador_atual['user_id']))
        
        asset_id = cursor.lastrowid
        
        for nome_campo, valor in dados_ativo.items():
            if valor is not None:
                bd.execute('''
                    INSERT INTO asset_data (asset_id, field_name, field_value) VALUES (?, ?, ?)
                ''', (asset_id, nome_campo, str(valor)))
        
        # Incrementar contador de assets se o número de série usa o prefixo configurado
        prefixo = obter_config('prefix_assets', 'SLP')
        if serial_number.startswith(prefixo):
            bd.execute('''
                UPDATE sequence_counters 
                SET current_value = current_value + 1, updated_at = CURRENT_TIMESTAMP
                WHERE counter_type = 'assets'
            ''')
            # Criar contador se não existir
            if bd.execute('SELECT changes()').fetchone()[0] == 0:
                bd.execute('''
                    INSERT OR IGNORE INTO sequence_counters (counter_type, current_value) VALUES ('assets', 1)
                ''')
        
        bd.commit()
        registar_auditoria(bd, g.utilizador_atual['user_id'], 'CREATE', 'assets', asset_id, None, dados)
        return jsonify({'message': 'Ativo criado com sucesso', 'serial_number': serial_number}), 201
    except sqlite3.IntegrityError:
        return jsonify({'error': 'Número de Série já existe'}), 400

@app.route('/api/assets/<string:serial_number>', methods=['PUT'])
@requer_autenticacao
def atualizar_ativo(serial_number):
    """Atualiza um ativo existente. Permite alterar RFID Tag se danificado."""
    bd = obter_bd()
    ativo = bd.execute('SELECT * FROM assets WHERE serial_number = ?', (serial_number,)).fetchone()
    
    if not ativo:
        return jsonify({'error': 'Ativo não encontrado'}), 404
    
    dados = request.json
    dados_ativo = dados.get('data', {})
    
    # Obter dados antigos para auditoria
    dados_antigos = bd.execute('SELECT field_name, field_value FROM asset_data WHERE asset_id = ?', (ativo['id'],)).fetchall()
    dados_antigos_dict = {d['field_name']: d['field_value'] for d in dados_antigos}
    
    for nome_campo, valor in dados_ativo.items():
        bd.execute('''
            INSERT OR REPLACE INTO asset_data (asset_id, field_name, field_value) VALUES (?, ?, ?)
        ''', (ativo['id'], nome_campo, str(valor) if valor is not None else None))
    
    bd.execute('UPDATE assets SET updated_at = ?, updated_by = ? WHERE id = ?',
               (datetime.now().isoformat(), g.utilizador_atual['user_id'], ativo['id']))
    bd.commit()
    registar_auditoria(bd, g.utilizador_atual['user_id'], 'UPDATE', 'assets', ativo['id'], dados_antigos_dict, dados_ativo)
    return jsonify({'message': 'Ativo atualizado'})

@app.route('/api/assets/<string:serial_number>', methods=['DELETE'])
@requer_autenticacao
def eliminar_ativo(serial_number):
    """Elimina um ativo pelo número de série."""
    bd = obter_bd()
    ativo = bd.execute('SELECT * FROM assets WHERE serial_number = ?', (serial_number,)).fetchone()
    
    if not ativo:
        return jsonify({'error': 'Ativo não encontrado'}), 404
    
    dados_antigos = bd.execute('SELECT field_name, field_value FROM asset_data WHERE asset_id = ?', (ativo['id'],)).fetchall()
    dados_antigos_dict = {d['field_name']: d['field_value'] for d in dados_antigos}
    
    bd.execute('DELETE FROM asset_data WHERE asset_id = ?', (ativo['id'],))
    bd.execute('DELETE FROM maintenance_log WHERE asset_id = ?', (ativo['id'],))
    bd.execute('DELETE FROM assets WHERE id = ?', (ativo['id'],))
    bd.commit()
    registar_auditoria(bd, g.utilizador_atual['user_id'], 'DELETE', 'assets', ativo['id'], 
                       {'serial_number': serial_number, **dados_antigos_dict}, None)
    return jsonify({'message': 'Ativo eliminado'})

@app.route('/api/assets/<string:serial_number>/maintenance', methods=['POST'])
@requer_autenticacao
def adicionar_manutencao(serial_number):
    """Adiciona um registo de manutenção a um ativo."""
    bd = obter_bd()
    ativo = bd.execute('SELECT * FROM assets WHERE serial_number = ?', (serial_number,)).fetchone()
    
    if not ativo:
        return jsonify({'error': 'Ativo não encontrado'}), 404
    
    dados = request.json
    bd.execute('''
        INSERT INTO maintenance_log (asset_id, action_type, description, performed_by)
        VALUES (?, ?, ?, ?)
    ''', (ativo['id'], dados.get('action_type'), dados.get('description'), g.utilizador_atual['user_id']))
    bd.commit()
    return jsonify({'message': 'Registo de manutenção adicionado'}), 201

# =============================================================================
# ALTERAÇÃO DE ESTADO EM MASSA
# =============================================================================

@app.route('/api/assets/change-status', methods=['POST'])
@requer_autenticacao
def alterar_estado_ativos():
    """
    Altera o estado de um ou mais ativos.
    Espera: { serial_numbers: [...], new_status: '...', description: '...' }
    """
    dados = request.json
    serial_numbers = dados.get('serial_numbers', [])
    novo_estado = dados.get('new_status')
    descricao = dados.get('description', '').strip()
    
    # Validações
    if not serial_numbers:
        return jsonify({'error': 'Nenhum ativo selecionado'}), 400
    
    if not novo_estado:
        return jsonify({'error': 'Estado não especificado'}), 400
    
    if not descricao:
        return jsonify({'error': 'Descrição obrigatória'}), 400
    
    estados_validos = ['Operacional', 'Manutenção Necessária', 'Em Reparação', 'Desativado']
    if novo_estado not in estados_validos:
        return jsonify({'error': f'Estado inválido. Estados válidos: {", ".join(estados_validos)}'}), 400
    
    bd = obter_bd()
    user_id = g.utilizador_atual['user_id']
    
    resultados = {'sucesso': [], 'erros': []}
    
    for sn in serial_numbers:
        try:
            # Obter ativo
            ativo = bd.execute('SELECT id FROM assets WHERE serial_number = ?', (sn,)).fetchone()
            if not ativo:
                resultados['erros'].append({'serial_number': sn, 'erro': 'Ativo não encontrado'})
                continue
            
            asset_id = ativo['id']
            
            # Obter estado anterior
            estado_anterior_row = bd.execute('''
                SELECT field_value FROM asset_data 
                WHERE asset_id = ? AND field_name = 'condition_status'
            ''', (asset_id,)).fetchone()
            estado_anterior = estado_anterior_row['field_value'] if estado_anterior_row else None
            
            # Actualizar estado em asset_data
            bd.execute('''
                INSERT OR REPLACE INTO asset_data (asset_id, field_name, field_value)
                VALUES (?, 'condition_status', ?)
            ''', (asset_id, novo_estado))
            
            # Actualizar timestamp do ativo
            bd.execute('''
                UPDATE assets SET updated_at = CURRENT_TIMESTAMP, updated_by = ?
                WHERE id = ?
            ''', (user_id, asset_id))
            
            # Registar no histórico de alterações de estado
            bd.execute('''
                INSERT INTO status_change_log (asset_id, previous_status, new_status, description, changed_by)
                VALUES (?, ?, ?, ?, ?)
            ''', (asset_id, estado_anterior, novo_estado, descricao, user_id))
            
            # Registar na auditoria
            bd.execute('''
                INSERT INTO audit_log (user_id, action, table_name, record_id, old_values, new_values)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (user_id, 'STATUS_CHANGE', 'assets', asset_id, 
                  json.dumps({'condition_status': estado_anterior}),
                  json.dumps({'condition_status': novo_estado, 'description': descricao})))
            
            resultados['sucesso'].append({
                'serial_number': sn,
                'estado_anterior': estado_anterior,
                'novo_estado': novo_estado
            })
            
        except Exception as e:
            resultados['erros'].append({'serial_number': sn, 'erro': str(e)})
    
    bd.commit()
    
    return jsonify({
        'message': f'{len(resultados["sucesso"])} ativo(s) atualizado(s)',
        'resultados': resultados
    })

@app.route('/api/assets/<string:serial_number>/status-history', methods=['GET'])
@requer_autenticacao
def obter_historico_estado(serial_number):
    """Devolve o histórico de alterações de estado de um ativo."""
    bd = obter_bd()
    ativo = bd.execute('SELECT id FROM assets WHERE serial_number = ?', (serial_number,)).fetchone()
    
    if not ativo:
        return jsonify({'error': 'Ativo não encontrado'}), 404
    
    historico = bd.execute('''
        SELECT s.*, COALESCE(u.first_name || ' ' || u.last_name, u.email) as changed_by_name
        FROM status_change_log s 
        LEFT JOIN users u ON s.changed_by = u.id
        WHERE s.asset_id = ? 
        ORDER BY s.changed_at DESC
    ''', (ativo['id'],)).fetchall()
    
    return jsonify({
        'serial_number': serial_number,
        'history': [dict(h) for h in historico]
    })

# =============================================================================
# SISTEMA DE INTERVENÇÍ•ES
# =============================================================================

def allowed_file(filename):
    """Verifica se a extensão do ficheiro é permitida."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ----- TÉCNICOS EXTERNOS -----

@app.route('/api/external-technicians', methods=['GET'])
@requer_autenticacao
def listar_tecnicos_externos():
    """Lista todos os técnicos externos activos."""
    bd = obter_bd()
    tecnicos = bd.execute('''
        SELECT et.*, (u.first_name || ' ' || COALESCE(u.last_name, '')) as created_by_name
        FROM external_technicians et
        LEFT JOIN users u ON et.created_by = u.id
        WHERE et.active = 1
        ORDER BY et.company, et.name
    ''').fetchall()
    return jsonify([dict(t) for t in tecnicos])

@app.route('/api/external-technicians', methods=['POST'])
@requer_autenticacao
def criar_tecnico_externo():
    """Cria um novo técnico externo."""
    dados = request.json
    
    if not dados.get('name') or not dados.get('company'):
        return jsonify({'error': 'Nome e empresa são obrigatórios'}), 400
    
    bd = obter_bd()
    try:
        cursor = bd.execute('''
            INSERT INTO external_technicians (name, company, phone, email, notes, created_by)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (dados['name'], dados['company'], dados.get('phone'), dados.get('email'),
              dados.get('notes'), g.utilizador_atual['user_id']))
        bd.commit()
        return jsonify({'id': cursor.lastrowid, 'message': 'Técnico externo criado'}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/external-technicians/<int:id>', methods=['DELETE'])
@requer_autenticacao
@requer_admin
def desativar_tecnico_externo(id):
    """Desactiva um técnico externo (soft delete)."""
    bd = obter_bd()
    bd.execute('UPDATE external_technicians SET active = 0 WHERE id = ?', (id,))
    bd.commit()
    return jsonify({'message': 'Técnico externo desativado'})

# ----- INTERVENÇÍ•ES -----

@app.route('/api/interventions', methods=['GET'])
@requer_autenticacao
def listar_intervencoes():
    """Lista todas as intervenções com filtros opcionais."""
    bd = obter_bd()
    
    # Parâmetros de filtro
    status = request.args.get('status')
    tipo = request.args.get('type')
    asset_id = request.args.get('asset_id')
    serial_number = request.args.get('serial_number')
    pagina = int(request.args.get('page', 1))
    por_pagina = int(request.args.get('per_page', 20))
    
    # Query base
    query = '''
        SELECT i.*, a.serial_number, COALESCE(u.first_name || ' ' || u.last_name, u.email) as created_by_name,
               (SELECT GROUP_CONCAT(COALESCE(COALESCE(u2.first_name || ' ' || u2.last_name, u2.email), et.name || ' (' || et.company || ')'), ', ')
                FROM intervention_technicians it
                LEFT JOIN users u2 ON it.user_id = u2.id
                LEFT JOIN external_technicians et ON it.external_technician_id = et.id
                WHERE it.intervention_id = i.id) as technicians,
               (SELECT COUNT(*) FROM intervention_files WHERE intervention_id = i.id) as file_count,
               (SELECT COALESCE(SUM(cost_value), 0) FROM intervention_files WHERE intervention_id = i.id AND cost_value IS NOT NULL) as total_file_costs
        FROM interventions i
        JOIN assets a ON i.asset_id = a.id
        LEFT JOIN users u ON i.created_by = u.id
        WHERE 1=1
    '''
    params = []
    
    if status:
        query += ' AND i.status = ?'
        params.append(status)
    
    if tipo:
        query += ' AND i.intervention_type = ?'
        params.append(tipo)
    
    if asset_id:
        query += ' AND i.asset_id = ?'
        params.append(asset_id)
    
    if serial_number:
        query += ' AND a.serial_number = ?'
        params.append(serial_number)
    
    # Contagem total
    count_query = query.replace(
        'SELECT i.*, a.serial_number, COALESCE(u.first_name || ' ' || u.last_name, u.email) as created_by_name,',
        'SELECT COUNT(*) as total FROM (SELECT i.id'
    ).split('FROM interventions')[0] + 'FROM interventions' + query.split('FROM interventions')[1]
    # Simplificar para contagem
    total_query = f'''
        SELECT COUNT(*) as total FROM interventions i
        JOIN assets a ON i.asset_id = a.id
        WHERE 1=1 {' AND i.status = ?' if status else ''} {' AND i.intervention_type = ?' if tipo else ''} 
        {' AND i.asset_id = ?' if asset_id else ''} {' AND a.serial_number = ?' if serial_number else ''}
    '''
    count_params = [p for p in params]
    total = bd.execute(total_query, count_params).fetchone()['total']
    
    # Paginação e ordenação
    query += ' ORDER BY i.created_at DESC LIMIT ? OFFSET ?'
    params.extend([por_pagina, (pagina - 1) * por_pagina])
    
    intervencoes = bd.execute(query, params).fetchall()
    
    return jsonify({
        'interventions': [dict(i) for i in intervencoes],
        'total': total,
        'page': pagina,
        'per_page': por_pagina,
        'pages': (total + por_pagina - 1) // por_pagina
    })

@app.route('/api/interventions/<int:id>', methods=['GET'])
@requer_autenticacao
def obter_intervencao(id):
    """Obtém detalhes completos de uma intervenção."""
    bd = obter_bd()
    
    intervencao = bd.execute('''
        SELECT i.*, a.serial_number, COALESCE(u.first_name || ' ' || u.last_name, u.email) as created_by_name,
               COALESCE(u2.first_name || ' ' || u2.last_name, u2.email) as updated_by_name
        FROM interventions i
        JOIN assets a ON i.asset_id = a.id
        LEFT JOIN users u ON i.created_by = u.id
        LEFT JOIN users u2 ON i.updated_by = u2.id
        WHERE i.id = ?
    ''', (id,)).fetchone()
    
    if not intervencao:
        return jsonify({'error': 'Intervenção não encontrada'}), 404
    
    resultado = dict(intervencao)
    
    # Técnicos participantes
    tecnicos = bd.execute('''
        SELECT it.*, COALESCE(u.first_name || ' ' || u.last_name, u.email), et.name as external_name, et.company as external_company
        FROM intervention_technicians it
        LEFT JOIN users u ON it.user_id = u.id
        LEFT JOIN external_technicians et ON it.external_technician_id = et.id
        WHERE it.intervention_id = ?
    ''', (id,)).fetchall()
    resultado['technicians'] = [dict(t) for t in tecnicos]
    
    # Ficheiros
    ficheiros = bd.execute('''
        SELECT f.*, COALESCE(u.first_name || ' ' || u.last_name, u.email) as uploaded_by_name
        FROM intervention_files f
        LEFT JOIN users u ON f.uploaded_by = u.id
        WHERE f.intervention_id = ?
        ORDER BY f.file_category, f.uploaded_at
    ''', (id,)).fetchall()
    resultado['files'] = [dict(f) for f in ficheiros]
    
    # Log de edições
    edicoes = bd.execute('''
        SELECT e.*, COALESCE(u.first_name || ' ' || u.last_name, u.email) as edited_by_name,
               COALESCE(u.first_name || ' ' || u.last_name, COALESCE(u.first_name || ' ' || u.last_name, u.email)) as edited_by_display
        FROM intervention_edit_log e
        LEFT JOIN users u ON e.edited_by = u.id
        WHERE e.intervention_id = ?
        ORDER BY e.edited_at DESC
    ''', (id,)).fetchall()
    resultado['edit_history'] = [dict(e) for e in edicoes]
    
    # Registos de tempo
    tempo_logs = bd.execute('''
        SELECT t.*, COALESCE(u.first_name || ' ' || u.last_name, u.email) as logged_by_name,
               COALESCE(u.first_name || ' ' || u.last_name, COALESCE(u.first_name || ' ' || u.last_name, u.email)) as logged_by_display
        FROM intervention_time_logs t
        LEFT JOIN users u ON t.logged_by = u.id
        WHERE t.intervention_id = ?
        ORDER BY t.logged_at DESC
    ''', (id,)).fetchall()
    resultado['time_logs'] = [dict(t) for t in tempo_logs]
    resultado['total_time_spent'] = sum(t['time_spent'] for t in tempo_logs)
    
    # Formatar tempo total em HH:MM
    total_hours = int(resultado['total_time_spent'])
    total_minutes = int((resultado['total_time_spent'] - total_hours) * 60)
    resultado['total_time_formatted'] = f"{total_hours}:{total_minutes:02d}"
    
    # Contar número de actualizações
    resultado['update_count'] = len(resultado['edit_history'])
    
    return jsonify(resultado)

@app.route('/api/interventions', methods=['POST'])
@requer_autenticacao
def criar_intervencao():
    """Cria uma nova intervenção."""
    dados = request.json
    
    # Validações obrigatórias
    if not dados.get('asset_id') and not dados.get('serial_number'):
        return jsonify({'error': 'Ativo (asset_id ou serial_number) obrigatório'}), 400
    
    if not dados.get('intervention_type'):
        return jsonify({'error': 'Tipo de intervenção obrigatório'}), 400
    
    tipos_validos = ['Manutenção Preventiva', 'Manutenção Corretiva', 'Inspeção', 'Substituição de Componente']
    if dados['intervention_type'] not in tipos_validos:
        return jsonify({'error': f'Tipo inválido. Tipos válidos: {", ".join(tipos_validos)}'}), 400
    
    if not dados.get('duration_hours'):
        return jsonify({'error': 'Duração da intervenção obrigatória'}), 400
    
    if not dados.get('parts_used'):
        return jsonify({'error': 'Peças/Materiais utilizados obrigatório'}), 400
    
    # Validações condicionais baseadas no tipo
    tipo = dados['intervention_type']
    if tipo != 'Manutenção Preventiva' and not dados.get('problem_description'):
        return jsonify({'error': 'Descrição do problema obrigatória para este tipo de intervenção'}), 400
    
    if tipo == 'Manutenção Corretiva' and not dados.get('solution_description'):
        return jsonify({'error': 'Descrição da solução obrigatória para Manutenção Corretiva'}), 400
    
    bd = obter_bd()
    
    # Obter asset_id
    asset_id = dados.get('asset_id')
    if not asset_id and dados.get('serial_number'):
        ativo = bd.execute('SELECT id FROM assets WHERE serial_number = ?', (dados['serial_number'],)).fetchone()
        if not ativo:
            return jsonify({'error': 'Ativo não encontrado'}), 404
        asset_id = ativo['id']
    
    # Obter estado actual do ativo
    estado_actual = bd.execute('''
        SELECT field_value FROM asset_data 
        WHERE asset_id = ? AND field_name = 'condition_status'
    ''', (asset_id,)).fetchone()
    previous_status = estado_actual['field_value'] if estado_actual else 'Não definido'
    
    user_id = g.utilizador_atual['user_id']
    
    try:
        # Criar intervenção
        cursor = bd.execute('''
            INSERT INTO interventions (
                asset_id, intervention_type, problem_description, solution_description,
                parts_used, total_cost, duration_hours, status, previous_asset_status,
                notes, created_by
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            asset_id, dados['intervention_type'], dados.get('problem_description'),
            dados.get('solution_description'), dados['parts_used'],
            dados.get('total_cost', 0), dados['duration_hours'],
            'em_curso', previous_status, dados.get('notes'), user_id
        ))
        
        intervention_id = cursor.lastrowid
        
        # Adicionar técnico responsável (o utilizador que criou)
        bd.execute('''
            INSERT INTO intervention_technicians (intervention_id, user_id, role)
            VALUES (?, ?, 'responsavel')
        ''', (intervention_id, user_id))
        
        # Adicionar técnicos acompanhantes (internos)
        for tech_id in dados.get('internal_technicians', []):
            if tech_id != user_id:
                bd.execute('''
                    INSERT INTO intervention_technicians (intervention_id, user_id, role)
                    VALUES (?, ?, 'participante')
                ''', (intervention_id, tech_id))
        
        # Adicionar técnicos externos
        for ext_id in dados.get('external_technicians', []):
            bd.execute('''
                INSERT INTO intervention_technicians (intervention_id, external_technician_id, role)
                VALUES (?, ?, 'participante')
            ''', (intervention_id, ext_id))
        
        # Mudar estado para "Em Reparação" se não for Inspeção
        # (Inspeções não alteram o estado do ativo)
        if tipo != 'Inspeção':
            bd.execute('''
                INSERT OR REPLACE INTO asset_data (asset_id, field_name, field_value)
                VALUES (?, 'condition_status', 'Em Reparação')
            ''', (asset_id,))
            
            bd.execute('''
                INSERT INTO status_change_log (asset_id, previous_status, new_status, description, changed_by, intervention_id)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (asset_id, previous_status, 'Em Reparação', 
                  f'Estado alterado automaticamente ao iniciar intervenção #{intervention_id} ({tipo})',
                  user_id, intervention_id))
        
        bd.commit()
        
        return jsonify({
            'id': intervention_id,
            'message': 'Intervenção criada com sucesso',
            'status_changed': tipo != 'Inspeção'
        }), 201
        
    except Exception as e:
        bd.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/interventions/<int:id>/complete', methods=['POST'])
@requer_autenticacao
def concluir_intervencao(id):
    """Conclui ou actualiza estado de uma intervenção."""
    dados = request.json
    
    bd = obter_bd()
    intervencao = bd.execute('SELECT * FROM interventions WHERE id = ?', (id,)).fetchone()
    
    if not intervencao:
        return jsonify({'error': 'Intervenção não encontrada'}), 404
    
    if intervencao['status'] == 'concluida':
        return jsonify({'error': 'Intervenção já está concluída'}), 400
    
    novo_estado = dados.get('final_status')
    descricao_estado = dados.get('status_description', '')
    
    user_id = g.utilizador_atual['user_id']
    
    # Obter estado actual do ativo
    estado_actual = bd.execute('''
        SELECT field_value FROM asset_data 
        WHERE asset_id = ? AND field_name = 'condition_status'
    ''', (intervencao['asset_id'],)).fetchone()
    current_status = estado_actual['field_value'] if estado_actual else 'Em Reparação'
    
    # Contar número de actualizações existentes
    update_count = bd.execute('''
        SELECT COUNT(*) as count FROM intervention_edit_log WHERE intervention_id = ?
    ''', (id,)).fetchone()['count']
    
    try:
        # Só concluir intervenção se estado for Operacional
        if novo_estado == 'Operacional':
            bd.execute('''
                UPDATE interventions 
                SET status = 'concluida', final_asset_status = ?, completed_at = CURRENT_TIMESTAMP,
                    updated_at = CURRENT_TIMESTAMP, updated_by = ?,
                    solution_description = COALESCE(?, solution_description)
                WHERE id = ?
            ''', (novo_estado, user_id, dados.get('solution_description'), id))
            
            # Registar no log
            bd.execute('''
                INSERT INTO intervention_edit_log (intervention_id, edited_by, field_name, old_value, new_value)
                VALUES (?, ?, ?, ?, ?)
            ''', (id, user_id, 'status', 'em_curso', f'concluida (Actualização #{update_count + 1})'))
            
            message = 'Intervenção concluída com sucesso'
        else:
            # Apenas actualizar estado do ativo, intervenção continua aberta
            bd.execute('''
                UPDATE interventions 
                SET updated_at = CURRENT_TIMESTAMP, updated_by = ?
                WHERE id = ?
            ''', (user_id, id))
            
            # Registar no log
            if novo_estado:
                bd.execute('''
                    INSERT INTO intervention_edit_log (intervention_id, edited_by, field_name, old_value, new_value)
                    VALUES (?, ?, ?, ?, ?)
                ''', (id, user_id, 'asset_status', current_status, f'{novo_estado} (Actualização #{update_count + 1})'))
            
            message = f'Estado actualizado. Intervenção continua em curso (Actualização #{update_count + 1})'
        
        # Se foi especificado um novo estado para o ativo
        if novo_estado and novo_estado != current_status:
            bd.execute('''
                INSERT OR REPLACE INTO asset_data (asset_id, field_name, field_value)
                VALUES (?, 'condition_status', ?)
            ''', (intervencao['asset_id'], novo_estado))
            
            bd.execute('''
                INSERT INTO status_change_log (asset_id, previous_status, new_status, description, changed_by, intervention_id)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (intervencao['asset_id'], current_status, 
                  novo_estado, descricao_estado or f'Alteração via intervenção #{id}', user_id, id))
        
        bd.commit()
        
        return jsonify({
            'message': message,
            'concluded': novo_estado == 'Operacional',
            'update_count': update_count + 1
        })
        
    except Exception as e:
        bd.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/interventions/<int:id>', methods=['PUT'])
@requer_autenticacao
def editar_intervencao(id):
    """Edita/actualiza uma intervenção. Regista log de alterações."""
    dados = request.json
    
    bd = obter_bd()
    intervencao = bd.execute('SELECT * FROM interventions WHERE id = ?', (id,)).fetchone()
    
    if not intervencao:
        return jsonify({'error': 'Intervenção não encontrada'}), 404
    
    # Só permitir edição se intervenção não estiver concluída OU se for admin
    if intervencao['status'] == 'concluida' and g.utilizador_atual['role'] != 'admin':
        return jsonify({'error': 'Intervenção já concluída. Apenas administradores podem editar.'}), 403
    
    user_id = g.utilizador_atual['user_id']
    campos_editaveis = ['problem_description', 'solution_description', 'parts_used', 
                        'total_cost', 'duration_hours', 'notes']
    
    try:
        for campo in campos_editaveis:
            if campo in dados and dados[campo] != intervencao[campo]:
                # Registar alteração no log
                bd.execute('''
                    INSERT INTO intervention_edit_log (intervention_id, edited_by, field_name, old_value, new_value)
                    VALUES (?, ?, ?, ?, ?)
                ''', (id, user_id, campo, str(intervencao[campo] or ''), str(dados[campo] or '')))
                
                # Actualizar campo
                # campo is from campos_editaveis whitelist - safe for interpolation
                bd.execute(f'UPDATE interventions SET {campo} = ?, updated_at = CURRENT_TIMESTAMP, updated_by = ? WHERE id = ?',
                          (dados[campo], user_id, id))
        
        # Registar tempo de trabalho se fornecido
        if dados.get('time_spent') and float(dados.get('time_spent', 0)) > 0:
            time_formatted = dados.get('time_formatted', f"{float(dados['time_spent']):.2f}h")
            
            bd.execute('''
                INSERT INTO intervention_time_logs (intervention_id, logged_by, time_spent, description)
                VALUES (?, ?, ?, ?)
            ''', (id, user_id, float(dados['time_spent']), dados.get('solution_description', 'Actualização de progresso')))
            
            # Actualizar duração total da intervenção
            total_time = bd.execute('''
                SELECT COALESCE(SUM(time_spent), 0) as total FROM intervention_time_logs WHERE intervention_id = ?
            ''', (id,)).fetchone()['total']
            bd.execute('UPDATE interventions SET duration_hours = ? WHERE id = ?', (total_time, id))
            
            # Registar tempo no histórico com formato legível
            bd.execute('''
                INSERT INTO intervention_edit_log (intervention_id, edited_by, field_name, old_value, new_value)
                VALUES (?, ?, ?, ?, ?)
            ''', (id, user_id, 'time_spent', '', f'+{time_formatted}'))
        
        # Contar número de actualizações
        update_count = bd.execute('''
            SELECT COUNT(DISTINCT edited_at) as count FROM intervention_edit_log WHERE intervention_id = ?
        ''', (id,)).fetchone()['count']
        
        bd.commit()
        return jsonify({'message': 'Intervenção actualizada com sucesso', 'update_count': update_count})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ----- FICHEIROS DE INTERVENÇÍ•ES -----

@app.route('/api/interventions/<int:id>/files', methods=['POST'])
@requer_autenticacao
def upload_ficheiro_intervencao(id):
    """Upload de ficheiro para uma intervenção."""
    bd = obter_bd()
    
    intervencao = bd.execute('SELECT * FROM interventions WHERE id = ?', (id,)).fetchone()
    if not intervencao:
        return jsonify({'error': 'Intervenção não encontrada'}), 404
    
    if 'file' not in request.files:
        return jsonify({'error': 'Nenhum ficheiro enviado'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Nenhum ficheiro selecionado'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'error': f'Tipo de ficheiro não permitido. Permitidos: {", ".join(ALLOWED_EXTENSIONS)}'}), 400
    
    # Verificar tamanho
    file.seek(0, 2)
    size = file.tell()
    file.seek(0)
    
    if size > MAX_FILE_SIZE:
        return jsonify({'error': f'Ficheiro muito grande. Máximo: {MAX_FILE_SIZE // (1024*1024)}MB'}), 400
    
    # Obter metadados
    category = request.form.get('category', 'outros')  # antes, durante, depois, fatura, outros
    description = request.form.get('description', '')
    cost_value = request.form.get('cost_value')
    
    if category not in ['antes', 'durante', 'depois', 'fatura', 'outros']:
        return jsonify({'error': 'Categoria inválida'}), 400
    
    # Gerar nome único
    ext = file.filename.rsplit('.', 1)[1].lower()
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    unique_name = f'INT{id}_{category}_{timestamp}.{ext}'
    
    # Criar pasta para a intervenção se não existir
    pasta_intervencao = os.path.join(PASTA_INTERVENCOES, str(id))
    os.makedirs(pasta_intervencao, exist_ok=True)
    
    # Guardar ficheiro
    file_path = os.path.join(pasta_intervencao, unique_name)
    file.save(file_path)
    
    # Registar na BD
    try:
        cursor = bd.execute('''
            INSERT INTO intervention_files (
                intervention_id, file_category, file_name, original_name, 
                file_path, file_type, file_size, description, cost_value, uploaded_by
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (id, category, unique_name, file.filename, file_path, ext, size,
              description, float(cost_value) if cost_value else None,
              g.utilizador_atual['user_id']))
        
        # Actualizar custo total da intervenção se for fatura com valor
        if cost_value and category == 'fatura':
            bd.execute('''
                UPDATE interventions SET total_cost = (
                    SELECT COALESCE(SUM(cost_value), 0) FROM intervention_files 
                    WHERE intervention_id = ? AND cost_value IS NOT NULL
                ) WHERE id = ?
            ''', (id, id))
        
        bd.commit()
        
        return jsonify({
            'id': cursor.lastrowid,
            'filename': unique_name,
            'message': 'Ficheiro carregado com sucesso'
        }), 201
        
    except Exception as e:
        # Remover ficheiro se falhou o registo na BD
        if os.path.exists(file_path):
            os.remove(file_path)
        return jsonify({'error': str(e)}), 500

@app.route('/api/interventions/<int:int_id>/files/<int:file_id>', methods=['GET'])
@requer_autenticacao
def download_ficheiro_intervencao(int_id, file_id):
    """Download de um ficheiro de intervenção."""
    bd = obter_bd()
    
    ficheiro = bd.execute('''
        SELECT * FROM intervention_files WHERE id = ? AND intervention_id = ?
    ''', (file_id, int_id)).fetchone()
    
    if not ficheiro:
        return jsonify({'error': 'Ficheiro não encontrado'}), 404
    
    file_path = ficheiro['file_path']
    safe_path = validate_safe_path(file_path, PASTA_UPLOADS)
    if not safe_path or not os.path.exists(safe_path):
        return jsonify({'error': 'Ficheiro nao existe no servidor'}), 404

    return send_file(safe_path, download_name=ficheiro['original_name'])

@app.route('/api/interventions/<int:int_id>/files/<int:file_id>', methods=['DELETE'])
@requer_autenticacao
@requer_admin
def eliminar_ficheiro_intervencao(int_id, file_id):
    """Elimina um ficheiro de intervenção (apenas admin)."""
    bd = obter_bd()
    
    ficheiro = bd.execute('''
        SELECT * FROM intervention_files WHERE id = ? AND intervention_id = ?
    ''', (file_id, int_id)).fetchone()
    
    if not ficheiro:
        return jsonify({'error': 'Ficheiro não encontrado'}), 404
    
    try:
        # Eliminar ficheiro físico
        if os.path.exists(ficheiro['file_path']):
            os.remove(ficheiro['file_path'])
        
        # Eliminar registo na BD
        bd.execute('DELETE FROM intervention_files WHERE id = ?', (file_id,))
        
        # Actualizar custo total
        bd.execute('''
            UPDATE interventions SET total_cost = (
                SELECT COALESCE(SUM(cost_value), 0) FROM intervention_files 
                WHERE intervention_id = ? AND cost_value IS NOT NULL
            ) WHERE id = ?
        ''', (int_id, int_id))
        
        bd.commit()
        return jsonify({'message': 'Ficheiro eliminado com sucesso'})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ----- LISTA DE UTILIZADORES PARA SELEÇÍO -----

@app.route('/api/users/technicians', methods=['GET'])
@requer_autenticacao
def listar_tecnicos():
    """Lista utilizadores que podem ser técnicos (operators e admins)."""
    bd = obter_bd()
    users = bd.execute('''
        SELECT id, email, role, first_name, last_name FROM users 
        WHERE role IN ('admin', 'operator') AND active = 1
        ORDER BY first_name, last_name, email
    ''').fetchall()
    result = []
    for u in users:
        user_dict = dict(u)
        # Criar nome completo para exibição
        if u['first_name'] and u['last_name']:
            user_dict['display_name'] = f"{u['first_name']} {u['last_name']}"
        else:
            user_dict['display_name'] = u['email']
        result.append(user_dict)
    return jsonify(result)
    return jsonify(result)

# =============================================================================
# CONSULTA PÍšBLICA RFID
# =============================================================================

@app.route('/api/public/lookup/<string:identifier>', methods=['GET'])
def consulta_publica(identifier):
    """
    Endpoint público para leitores RFID ou pesquisa por serial number.
    Procura primeiro por RFID Tag, depois por Número de Série.
    """
    bd = obter_bd()
    
    # Procurar por RFID Tag em asset_data
    resultado_rfid = bd.execute('''
        SELECT a.* FROM assets a
        JOIN asset_data ad ON a.id = ad.asset_id
        WHERE ad.field_name = 'rfid_tag' AND ad.field_value = ?
    ''', (identifier,)).fetchone()
    
    # Se não encontrou por RFID, procurar por serial_number
    ativo = resultado_rfid if resultado_rfid else bd.execute('SELECT * FROM assets WHERE serial_number = ?', (identifier,)).fetchone()
    
    if not ativo:
        return jsonify({'found': False, 'message': 'Ativo não encontrado'}), 404
    
    dados = bd.execute('''
        SELECT field_name, field_value FROM asset_data 
        WHERE asset_id = ? AND field_name IN (
            'rfid_tag', 'product_reference', 'model', 'manufacturer',
            'installation_location', 'condition_status', 'warranty_end_date',
            'power_watts', 'connection_type'
        )
    ''', (ativo['id'],)).fetchall()
    
    return jsonify({
        'found': True,
        'serial_number': ativo['serial_number'],
        'data': {d['field_name']: d['field_value'] for d in dados}
    })

# =============================================================================
# GESTÍO DE BACKUPS
# =============================================================================

def limpar_backups_antigos():
    """Remove backups antigos, mantendo apenas os últimos MAX_BACKUPS."""
    try:
        backups = []
        for ficheiro in os.listdir(PASTA_BACKUPS):
            if ficheiro.startswith('backup_') and ficheiro.endswith('.zip'):
                caminho = os.path.join(PASTA_BACKUPS, ficheiro)
                backups.append({'filename': ficheiro, 'path': caminho, 'mtime': os.path.getmtime(caminho)})
        
        backups.sort(key=lambda x: x['mtime'], reverse=True)
        
        for backup in backups[MAX_BACKUPS:]:
            try:
                os.remove(backup['path'])
                print(f"âœ“ Backup antigo removido: {backup['filename']}")
            except Exception as e:
                print(f"âœ— Erro ao remover backup: {e}")
    except Exception as e:
        print(f"âœ— Erro na limpeza de backups: {e}")

def criar_backup_interno(tenant_id=None):
    """Cria um backup da base de dados do tenant em formato ZIP."""
    try:
        if tenant_id is None:
            tenant_id = getattr(g, 'tenant_id', MASTER_TENANT_ID)
        
        # Caminho da BD do tenant
        db_path = obter_caminho_bd_tenant(tenant_id)
        if not os.path.exists(db_path):
            return False, 'Base de dados não encontrada', None
        
        # Criar pasta de backup do tenant
        pasta_backup_tenant = os.path.join(PASTA_BACKUPS, tenant_id)
        os.makedirs(pasta_backup_tenant, exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        nome_ficheiro = f'backup_{tenant_id}_{timestamp}.zip'
        caminho_backup = os.path.join(pasta_backup_tenant, nome_ficheiro)
        
        with zipfile.ZipFile(caminho_backup, 'w', zipfile.ZIP_DEFLATED) as zipf:
            zipf.write(db_path, 'database.db')
            metadata = {
                'backup_date': timestamp,
                'tenant_id': tenant_id,
                'database_size': os.path.getsize(db_path),
                'created_by': 'SmartLamppost System'
            }
            zipf.writestr('backup_info.json', json.dumps(metadata, indent=2))
        
        limpar_backups_antigos()
        return True, f'Backup criado: {nome_ficheiro}', nome_ficheiro
    except Exception as e:
        return False, f'Erro ao criar backup: {str(e)}', None

@app.route('/api/backup/create', methods=['POST'])
@requer_autenticacao
def criar_backup():
    """Cria um backup manual da base de dados."""
    sucesso, mensagem, ficheiro = criar_backup_interno()
    
    if sucesso:
        bd = obter_bd()
        registar_auditoria(bd, g.utilizador_atual['user_id'], 'BACKUP_CREATE', 'database', None, None, {'filename': ficheiro})
        bd.commit()
        return jsonify({'success': True, 'message': mensagem, 'filename': ficheiro}), 201
    return jsonify({'success': False, 'error': mensagem}), 500

@app.route('/api/backup/list', methods=['GET'])
@requer_autenticacao
def listar_backups():
    """Lista todos os backups disponíveis."""
    try:
        backups = []
        for ficheiro in os.listdir(PASTA_BACKUPS):
            if ficheiro.startswith('backup_') and ficheiro.endswith('.zip'):
                caminho = os.path.join(PASTA_BACKUPS, ficheiro)
                stats = os.stat(caminho)
                backups.append({
                    'filename': ficheiro,
                    'size': stats.st_size,
                    'size_mb': round(stats.st_size / (1024 * 1024), 2),
                    'created_at': datetime.fromtimestamp(stats.st_mtime).isoformat(),
                    'created_at_formatted': datetime.fromtimestamp(stats.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                })
        
        backups.sort(key=lambda x: x['created_at'], reverse=True)
        return jsonify({'backups': backups, 'total': len(backups), 'max_backups': MAX_BACKUPS})
    except Exception as e:
        return jsonify({'error': f'Erro ao listar backups: {str(e)}'}), 500

@app.route('/api/backup/download/<filename>', methods=['GET'])
@requer_autenticacao
def descarregar_backup(filename):
    """Permite o download de um backup específico."""
    if not filename.startswith('backup_') or not filename.endswith('.zip'):
        return jsonify({'error': 'Nome de ficheiro invalido'}), 400

    caminho = os.path.join(PASTA_BACKUPS, filename)
    safe_path = validate_safe_path(caminho, PASTA_BACKUPS)
    if not safe_path or not os.path.exists(safe_path):
        return jsonify({'error': 'Backup nao encontrado'}), 404

    return send_file(safe_path, mimetype='application/zip', as_attachment=True, download_name=filename)

@app.route('/api/backup/delete/<filename>', methods=['DELETE'])
@requer_admin
def eliminar_backup(filename):
    """Elimina um backup específico (apenas admin)."""
    if not filename.startswith('backup_') or not filename.endswith('.zip'):
        return jsonify({'error': 'Nome de ficheiro invalido'}), 400

    caminho = os.path.join(PASTA_BACKUPS, filename)
    safe_path = validate_safe_path(caminho, PASTA_BACKUPS)
    if not safe_path or not os.path.exists(safe_path):
        return jsonify({'error': 'Backup não encontrado'}), 404
    
    try:
        os.remove(caminho)
        bd = obter_bd()
        registar_auditoria(bd, g.utilizador_atual['user_id'], 'BACKUP_DELETE', 'database', None, {'filename': filename}, None)
        bd.commit()
        return jsonify({'message': f'Backup {filename} eliminado'})
    except Exception as e:
        return jsonify({'error': f'Erro ao eliminar: {str(e)}'}), 500

# =============================================================================
# EXPORTAÇÍO
# =============================================================================

def obter_ativos_para_exportacao(campos_selecionados=None):
    """Obtém todos os ativos formatados para exportação."""
    bd = obter_bd()
    
    esquema = bd.execute('SELECT * FROM schema_fields ORDER BY field_order').fetchall()
    esquema_dict = {f['field_name']: f for f in esquema}
    
    if campos_selecionados is None:
        campos_selecionados = [f['field_name'] for f in esquema]
    
    ativos = bd.execute('SELECT * FROM assets ORDER BY serial_number').fetchall()
    dados_exportacao = []
    
    for ativo in ativos:
        dados_ativo = bd.execute('SELECT field_name, field_value FROM asset_data WHERE asset_id = ?', (ativo['id'],)).fetchall()
        dados_ativo_dict = {d['field_name']: d['field_value'] for d in dados_ativo}
        
        manutencoes = bd.execute('''
            SELECT m.performed_at, m.action_type, m.description, COALESCE(u.first_name || ' ' || u.last_name, u.email) as performed_by
            FROM maintenance_log m LEFT JOIN users u ON m.performed_by = u.id
            WHERE m.asset_id = ? ORDER BY m.performed_at
        ''', (ativo['id'],)).fetchall()
        
        if not manutencoes:
            linha = {'serial_number': ativo['serial_number'], 'created_at': ativo['created_at'], 'updated_at': ativo['updated_at']}
            for campo in campos_selecionados:
                linha[campo] = dados_ativo_dict.get(campo, '')
            linha.update({'maintenance_date': '', 'maintenance_type': '', 'maintenance_description': '', 'maintenance_performed_by': ''})
            dados_exportacao.append(linha)
        else:
            for m in manutencoes:
                linha = {'serial_number': ativo['serial_number'], 'created_at': ativo['created_at'], 'updated_at': ativo['updated_at']}
                for campo in campos_selecionados:
                    linha[campo] = dados_ativo_dict.get(campo, '')
                linha.update({
                    'maintenance_date': m['performed_at'] or '',
                    'maintenance_type': m['action_type'] or '',
                    'maintenance_description': m['description'] or '',
                    'maintenance_performed_by': m['performed_by'] or ''
                })
                dados_exportacao.append(linha)
    
    return dados_exportacao, esquema_dict, campos_selecionados

def criar_ficheiro_excel(campos_selecionados=None):
    """Cria um ficheiro Excel com os dados dos ativos."""
    if not EXCEL_DISPONIVEL:
        raise Exception("Biblioteca openpyxl não disponível")
    
    dados_exportacao, esquema_dict, campos_incluidos = obter_ativos_para_exportacao(campos_selecionados)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Ativos"
    
    # Estilos do cabeçalho
    estilo_cabecalho = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    fonte_cabecalho = Font(bold=True, color="FFFFFF", size=11)
    alinhamento_cabecalho = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Criar cabeçalhos - Nº Série é o primeiro
    cabecalhos = ['Nº Série', 'Data Criação', 'Íšltima Atualização']
    for campo in campos_incluidos:
        info = esquema_dict.get(campo)
        cabecalhos.append(info['field_label'] if info else campo)
    cabecalhos.extend(['Data Manutenção', 'Tipo Manutenção', 'Descrição Manutenção', 'Executado Por'])
    
    for col, cabecalho in enumerate(cabecalhos, 1):
        celula = ws.cell(row=1, column=col, value=cabecalho)
        celula.fill = estilo_cabecalho
        celula.font = fonte_cabecalho
        celula.alignment = alinhamento_cabecalho
    
    # Escrever dados
    for linha_num, dados_linha in enumerate(dados_exportacao, 2):
        ws.cell(row=linha_num, column=1, value=dados_linha['serial_number'])
        ws.cell(row=linha_num, column=2, value=dados_linha['created_at'])
        ws.cell(row=linha_num, column=3, value=dados_linha['updated_at'])
        
        col = 4
        for campo in campos_incluidos:
            ws.cell(row=linha_num, column=col, value=dados_linha.get(campo, ''))
            col += 1
        
        ws.cell(row=linha_num, column=col, value=dados_linha.get('maintenance_date', ''))
        ws.cell(row=linha_num, column=col + 1, value=dados_linha.get('maintenance_type', ''))
        ws.cell(row=linha_num, column=col + 2, value=dados_linha.get('maintenance_description', ''))
        ws.cell(row=linha_num, column=col + 3, value=dados_linha.get('maintenance_performed_by', ''))
    
    # Ajustar largura das colunas
    for coluna in ws.columns:
        largura_max = max(len(str(celula.value or '')) for celula in coluna)
        ws.column_dimensions[coluna[0].column_letter].width = min(largura_max + 2, 50)
    
    ws.freeze_panes = 'A2'
    
    bd = obter_bd()
    
    # FOLHA 2: Histórico de Alterações de Estado (com referência a intervenção)
    historico_estados = bd.execute('''
        SELECT a.serial_number, s.previous_status, s.new_status, s.description, 
               s.changed_at, COALESCE(u.first_name || ' ' || u.last_name, u.email) as changed_by, s.intervention_id
        FROM status_change_log s
        JOIN assets a ON s.asset_id = a.id
        LEFT JOIN users u ON s.changed_by = u.id
        ORDER BY s.changed_at DESC
    ''').fetchall()
    
    ws2 = wb.create_sheet(title="Histórico Estados")
    
    # Cabeçalhos
    cabecalhos_hist = ['Nº Série', 'Estado Anterior', 'Novo Estado', 'Descrição', 'Data/Hora', 'Alterado Por', 'ID Intervenção']
    for col, cabecalho in enumerate(cabecalhos_hist, 1):
        celula = ws2.cell(row=1, column=col, value=cabecalho)
        celula.fill = estilo_cabecalho
        celula.font = fonte_cabecalho
        celula.alignment = alinhamento_cabecalho
    
    # Dados
    for linha_num, hist in enumerate(historico_estados, 2):
        ws2.cell(row=linha_num, column=1, value=hist['serial_number'])
        ws2.cell(row=linha_num, column=2, value=hist['previous_status'] or 'N/D')
        ws2.cell(row=linha_num, column=3, value=hist['new_status'])
        ws2.cell(row=linha_num, column=4, value=hist['description'])
        ws2.cell(row=linha_num, column=5, value=hist['changed_at'])
        ws2.cell(row=linha_num, column=6, value=hist['changed_by'])
        ws2.cell(row=linha_num, column=7, value=hist['intervention_id'] if hist['intervention_id'] else '')
    
    # Ajustar largura
    for coluna in ws2.columns:
        largura_max = max(len(str(celula.value or '')) for celula in coluna)
        ws2.column_dimensions[coluna[0].column_letter].width = min(largura_max + 2, 50)
    
    ws2.freeze_panes = 'A2'
    
    # FOLHA 3: Intervenções
    intervencoes = bd.execute('''
        SELECT i.id, a.serial_number, i.intervention_type, i.problem_description,
               i.solution_description, i.parts_used, i.total_cost, i.duration_hours,
               i.status, i.previous_asset_status, i.final_asset_status,
               i.created_at, i.completed_at, COALESCE(u.first_name || ' ' || u.last_name, u.email) as created_by,
               i.notes,
               (SELECT GROUP_CONCAT(COALESCE(COALESCE(u2.first_name || ' ' || u2.last_name, u2.email), et.name || ' (' || et.company || ')'), '; ')
                FROM intervention_technicians it
                LEFT JOIN users u2 ON it.user_id = u2.id
                LEFT JOIN external_technicians et ON it.external_technician_id = et.id
                WHERE it.intervention_id = i.id) as technicians
        FROM interventions i
        JOIN assets a ON i.asset_id = a.id
        LEFT JOIN users u ON i.created_by = u.id
        ORDER BY i.created_at DESC
    ''').fetchall()
    
    ws3 = wb.create_sheet(title="Intervenções")
    
    # Cabeçalhos de intervenções
    cabecalhos_int = ['ID', 'Nº Série', 'Tipo', 'Descrição Problema', 'Descrição Solução', 
                      'Peças/Materiais', 'Custo Total (â‚¬)', 'Duração (h)', 'Estado Intervenção',
                      'Estado Ativo Antes', 'Estado Ativo Depois', 'Data Criação', 
                      'Data Conclusão', 'Criado Por', 'Técnicos', 'Notas']
    
    for col, cabecalho in enumerate(cabecalhos_int, 1):
        celula = ws3.cell(row=1, column=col, value=cabecalho)
        celula.fill = estilo_cabecalho
        celula.font = fonte_cabecalho
        celula.alignment = alinhamento_cabecalho
    
    # Dados de intervenções
    for linha_num, inter in enumerate(intervencoes, 2):
        ws3.cell(row=linha_num, column=1, value=inter['id'])
        ws3.cell(row=linha_num, column=2, value=inter['serial_number'])
        ws3.cell(row=linha_num, column=3, value=inter['intervention_type'])
        ws3.cell(row=linha_num, column=4, value=inter['problem_description'] or '')
        ws3.cell(row=linha_num, column=5, value=inter['solution_description'] or '')
        ws3.cell(row=linha_num, column=6, value=inter['parts_used'] or '')
        ws3.cell(row=linha_num, column=7, value=inter['total_cost'] or 0)
        ws3.cell(row=linha_num, column=8, value=inter['duration_hours'])
        ws3.cell(row=linha_num, column=9, value='Concluída' if inter['status'] == 'concluida' else 'Em Curso')
        ws3.cell(row=linha_num, column=10, value=inter['previous_asset_status'] or '')
        ws3.cell(row=linha_num, column=11, value=inter['final_asset_status'] or '')
        ws3.cell(row=linha_num, column=12, value=inter['created_at'])
        ws3.cell(row=linha_num, column=13, value=inter['completed_at'] or '')
        ws3.cell(row=linha_num, column=14, value=inter['created_by'])
        ws3.cell(row=linha_num, column=15, value=inter['technicians'] or '')
        ws3.cell(row=linha_num, column=16, value=inter['notes'] or '')
    
    # Ajustar largura
    for coluna in ws3.columns:
        largura_max = max(len(str(celula.value or '')) for celula in coluna)
        ws3.column_dimensions[coluna[0].column_letter].width = min(largura_max + 2, 50)
    
    ws3.freeze_panes = 'A2'
    
    ficheiro_excel = BytesIO()
    wb.save(ficheiro_excel)
    ficheiro_excel.seek(0)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return ficheiro_excel, f'ativos_export_{timestamp}.xlsx'

@app.route('/api/export/excel', methods=['POST'])
@requer_autenticacao
def exportar_excel():
    """Exporta ativos para ficheiro Excel."""
    if not EXCEL_DISPONIVEL:
        return jsonify({'error': 'Exportação Excel não disponível'}), 500
    
    try:
        dados = request.json or {}
        campos_selecionados = dados.get('fields', None)
        
        ficheiro_excel, nome_ficheiro = criar_ficheiro_excel(campos_selecionados)
        
        bd = obter_bd()
        registar_auditoria(bd, g.utilizador_atual['user_id'], 'EXPORT_EXCEL', 'assets', None, None,
                          {'filename': nome_ficheiro, 'fields': campos_selecionados or 'all'})
        bd.commit()
        
        return send_file(ficheiro_excel, 
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True, download_name=nome_ficheiro)
    except Exception as e:
        return jsonify({'error': f'Erro ao exportar: {str(e)}'}), 500

@app.route('/api/export/excel/fields', methods=['GET'])
@requer_autenticacao
def campos_disponiveis_export():
    """Lista campos disponíveis para exportação, agrupados por categoria."""
    bd = obter_bd()
    esquema = bd.execute('SELECT * FROM schema_fields ORDER BY field_category, field_order').fetchall()
    
    categorias = {}
    for campo in esquema:
        cat = campo['field_category']
        if cat not in categorias:
            categorias[cat] = []
        categorias[cat].append({
            'field_name': campo['field_name'],
            'field_label': campo['field_label'],
            'field_type': campo['field_type']
        })
    
    nomes_categorias = {
        'identification': 'Identificação', 'specifications': 'Especificações',
        'installation': 'Instalação', 'warranty': 'Garantia',
        'maintenance': 'Manutenção', 'equipment': 'Equipamento',
        'other': 'Outros', 'custom': 'Personalizados'
    }
    
    return jsonify([
        {'category': cat, 'category_label': nomes_categorias.get(cat, cat.title()), 'fields': campos}
        for cat, campos in categorias.items()
    ])

@app.route('/api/export', methods=['GET'])
@requer_autenticacao
def exportar_json():
    """Exporta todos os ativos em formato JSON."""
    bd = obter_bd()
    ativos = bd.execute('SELECT * FROM assets').fetchall()
    
    resultado = []
    for ativo in ativos:
        ativo_dict = dict(ativo)
        dados = bd.execute('SELECT field_name, field_value FROM asset_data WHERE asset_id = ?', (ativo['id'],)).fetchall()
        ativo_dict['data'] = {d['field_name']: d['field_value'] for d in dados}
        resultado.append(ativo_dict)
    
    return jsonify({'exported_at': datetime.now().isoformat(), 'total_assets': len(resultado), 'assets': resultado})

@app.route('/api/import', methods=['POST'])
@requer_admin
def importar_json():
    """Importa ativos a partir de JSON."""
    dados = request.json
    ativos = dados.get('assets', [])
    
    bd = obter_bd()
    importados = 0
    erros = []
    
    for ativo in ativos:
        try:
            cursor = bd.execute('''
                INSERT INTO assets (rfid_tag, created_by, updated_by) VALUES (?, ?, ?)
            ''', (ativo['rfid_tag'], g.utilizador_atual['user_id'], g.utilizador_atual['user_id']))
            
            asset_id = cursor.lastrowid
            for nome_campo, valor in ativo.get('data', {}).items():
                if valor is not None:
                    bd.execute('INSERT INTO asset_data (asset_id, field_name, field_value) VALUES (?, ?, ?)',
                              (asset_id, nome_campo, str(valor)))
            importados += 1
        except sqlite3.IntegrityError:
            erros.append(f"RFID duplicado: {ativo['rfid_tag']}")
    
    bd.commit()
    return jsonify({'imported': importados, 'errors': erros})

# =============================================================================
# IMPORTAÇÍO EXCEL
# =============================================================================

@app.route('/api/import/excel/template', methods=['GET'])
@requer_autenticacao
def descarregar_template_import():
    """Gera um template Excel para importação de ativos."""
    if not EXCEL_DISPONIVEL:
        return jsonify({'error': 'Funcionalidade Excel não disponível'}), 500
    
    bd = obter_bd()
    esquema = bd.execute('SELECT * FROM schema_fields ORDER BY field_order').fetchall()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Template Importação"
    
    # Estilos
    estilo_cabecalho = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    fonte_cabecalho = Font(bold=True, color="FFFFFF", size=11)
    estilo_obrigatorio = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    
    # Cabeçalhos - Nº Série é sempre a primeira coluna e obrigatória
    cabecalhos = [('serial_number', 'Nº Série', True)]
    for campo in esquema:
        cabecalhos.append((campo['field_name'], campo['field_label'], campo['required'] == 1))
    
    for col, (nome, label, obrigatorio) in enumerate(cabecalhos, 1):
        celula = ws.cell(row=1, column=col, value=label)
        celula.fill = estilo_obrigatorio if obrigatorio else estilo_cabecalho
        celula.font = fonte_cabecalho
        ws.cell(row=2, column=col, value=f"({nome})")
    
    for col in range(1, len(cabecalhos) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20
    
    ficheiro = BytesIO()
    wb.save(ficheiro)
    ficheiro.seek(0)
    
    return send_file(ficheiro,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name='template_importacao_ativos.xlsx')

@app.route('/api/import/excel', methods=['POST'])
@requer_admin
def importar_excel():
    """
    Importa ativos a partir de ficheiro Excel.
    O Número de Série é a referência principal obrigatória.
    """
    if not EXCEL_DISPONIVEL:
        return jsonify({'error': 'Funcionalidade Excel não disponível'}), 500
    
    if 'file' not in request.files:
        return jsonify({'error': 'Nenhum ficheiro enviado'}), 400
    
    ficheiro = request.files['file']
    if ficheiro.filename == '':
        return jsonify({'error': 'Nenhum ficheiro selecionado'}), 400
    
    if not ficheiro.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Formato inválido. Use ficheiros .xlsx ou .xls'}), 400
    
    modo = request.form.get('mode', 'upsert')
    campos_selecionados = request.form.get('selected_fields', '[]')
    try:
        campos_selecionados = json.loads(campos_selecionados)
    except Exception:
        campos_selecionados = []
    
    try:
        wb = load_workbook(ficheiro, data_only=True)
        ws = wb.active
        
        bd = obter_bd()
        esquema = bd.execute('SELECT * FROM schema_fields').fetchall()
        
        # Mapeamento de labels/nomes para field_names
        mapeamento = {
            'nº série': 'serial_number', 'n série': 'serial_number', 'numero de serie': 'serial_number',
            'número de série': 'serial_number', 'serial_number': 'serial_number', 'serial': 'serial_number',
            'rfid tag': 'rfid_tag', 'rfid_tag': 'rfid_tag', 'rfid': 'rfid_tag'
        }
        for campo in esquema:
            mapeamento[campo['field_name'].lower()] = campo['field_name']
            mapeamento[campo['field_label'].lower()] = campo['field_name']
        
        # Ler cabeçalhos
        cabecalhos = []
        serial_col = None
        for col in range(1, ws.max_column + 1):
            valor = ws.cell(row=1, column=col).value
            if valor:
                valor_lower = str(valor).lower().strip()
                if valor_lower in mapeamento:
                    nome_campo = mapeamento[valor_lower]
                    if nome_campo == 'serial_number':
                        serial_col = col
                    cabecalhos.append(nome_campo)
                else:
                    cabecalhos.append(None)
            else:
                cabecalhos.append(None)
        
        if serial_col is None:
            return jsonify({'error': 'Coluna "Nº Série" não encontrada. Esta coluna é obrigatória.'}), 400
        
        linha_inicio = 2
        primeira_celula = ws.cell(row=2, column=1).value
        if primeira_celula and str(primeira_celula).startswith('('):
            linha_inicio = 3
        
        criados = 0
        atualizados = 0
        ignorados = 0
        erros = []
        
        for linha in range(linha_inicio, ws.max_row + 1):
            serial_number = ws.cell(row=linha, column=serial_col).value
            
            if not serial_number:
                continue
            
            serial_number = str(serial_number).strip()
            
            ativo_existente = bd.execute('SELECT id FROM assets WHERE serial_number = ?', (serial_number,)).fetchone()
            
            if modo == 'create' and ativo_existente:
                ignorados += 1
                continue
            elif modo == 'update' and not ativo_existente:
                ignorados += 1
                continue
            
            dados_ativo = {}
            for col, campo in enumerate(cabecalhos, 1):
                if campo and campo != 'serial_number':
                    if campos_selecionados and campo not in campos_selecionados:
                        continue
                    valor = ws.cell(row=linha, column=col).value
                    if valor is not None:
                        dados_ativo[campo] = str(valor).strip()
            
            try:
                if ativo_existente:
                    for nome_campo, valor in dados_ativo.items():
                        bd.execute('''
                            INSERT OR REPLACE INTO asset_data (asset_id, field_name, field_value)
                            VALUES (?, ?, ?)
                        ''', (ativo_existente['id'], nome_campo, valor))
                    
                    bd.execute('UPDATE assets SET updated_at = ?, updated_by = ? WHERE id = ?',
                              (datetime.now(), g.utilizador_atual['user_id'], ativo_existente['id']))
                    atualizados += 1
                else:
                    cursor = bd.execute('''
                        INSERT INTO assets (serial_number, created_by, updated_by)
                        VALUES (?, ?, ?)
                    ''', (serial_number, g.utilizador_atual['user_id'], g.utilizador_atual['user_id']))
                    
                    asset_id = cursor.lastrowid
                    for nome_campo, valor in dados_ativo.items():
                        bd.execute('''
                            INSERT INTO asset_data (asset_id, field_name, field_value)
                            VALUES (?, ?, ?)
                        ''', (asset_id, nome_campo, valor))
                    criados += 1
                    
            except Exception as e:
                erros.append(f"Linha {linha} (Nº Série: {serial_number}): {str(e)}")
        
        bd.commit()
        
        # Registar auditoria
        registar_auditoria(bd, g.utilizador_atual['user_id'], 'IMPORT_EXCEL', 'assets', None, None, {
            'filename': ficheiro.filename,
            'mode': modo,
            'created': criados,
            'updated': atualizados,
            'ignored': ignorados,
            'errors': len(erros)
        })
        bd.commit()
        
        return jsonify({
            'success': True,
            'created': criados,
            'updated': atualizados,
            'ignored': ignorados,
            'errors': erros,
            'total_processed': criados + atualizados + ignorados
        })
        
    except Exception as e:
        return jsonify({'error': f'Erro ao processar ficheiro: {str(e)}'}), 500

# =============================================================================
# ESTATÍSTICAS
# =============================================================================

@app.route('/api/stats', methods=['GET'])
@requer_autenticacao
def obter_estatisticas():
    """Devolve estatísticas gerais do sistema."""
    bd = obter_bd()
    
    total_ativos = bd.execute('SELECT COUNT(*) FROM assets').fetchone()[0]
    
    condicoes = bd.execute('''
        SELECT field_value, COUNT(*) as count FROM asset_data 
        WHERE field_name = 'condition_status' GROUP BY field_value
    ''').fetchall()
    
    municipios = bd.execute('''
        SELECT field_value, COUNT(*) as count FROM asset_data 
        WHERE field_name = 'municipality' AND field_value IS NOT NULL
        GROUP BY field_value ORDER BY count DESC LIMIT 10
    ''').fetchall()
    
    # Manutenções realizadas nos últimos 30 dias
    manutencoes_recentes = bd.execute('''
        SELECT COUNT(*) FROM maintenance_log WHERE performed_at > datetime('now', '-30 days')
    ''').fetchone()[0]
    
    # Manutenções programadas para os próximos 30 dias
    manutencoes_programadas = bd.execute('''
        SELECT COUNT(*) FROM asset_data 
        WHERE field_name = 'next_maintenance_date' 
        AND field_value IS NOT NULL 
        AND field_value != ''
        AND date(field_value) BETWEEN date('now') AND date('now', '+30 days')
    ''').fetchone()[0]
    
    # Lista de ativos com manutenção programada nos próximos 30 dias
    ativos_manutencao = bd.execute('''
        SELECT a.serial_number, ad.field_value as maintenance_date,
               (SELECT field_value FROM asset_data WHERE asset_id = a.id AND field_name = 'municipality') as municipality,
               (SELECT field_value FROM asset_data WHERE asset_id = a.id AND field_name = 'maintenance_notes') as notes
        FROM assets a
        JOIN asset_data ad ON a.id = ad.asset_id
        WHERE ad.field_name = 'next_maintenance_date' 
        AND ad.field_value IS NOT NULL 
        AND ad.field_value != ''
        AND date(ad.field_value) BETWEEN date('now') AND date('now', '+30 days')
        ORDER BY ad.field_value ASC
        LIMIT 10
    ''').fetchall()
    
    return jsonify({
        'total_assets': total_ativos,
        'by_condition': [dict(c) for c in condicoes],
        'by_municipality': [dict(m) for m in municipios],
        'recent_maintenance_30d': manutencoes_recentes,
        'scheduled_maintenance_30d': manutencoes_programadas,
        'upcoming_maintenance': [dict(a) for a in ativos_manutencao]
    })

@app.route('/api/stats/summary', methods=['GET'])
@requer_autenticacao
def obter_resumo_estatisticas():
    """Devolve resumo estatístico com filtro por município."""
    bd = obter_bd()
    municipio = request.args.get('municipality', '')
    
    todos_municipios = bd.execute('''
        SELECT DISTINCT field_value FROM asset_data 
        WHERE field_name = 'municipality' AND field_value IS NOT NULL AND field_value != ''
        ORDER BY field_value
    ''').fetchall()
    
    if municipio:
        ids_ativos = bd.execute('''
            SELECT DISTINCT asset_id FROM asset_data WHERE field_name = 'municipality' AND field_value = ?
        ''', (municipio,)).fetchall()
        lista_ids = [a['asset_id'] for a in ids_ativos]
        
        if not lista_ids:
            return jsonify({
                'municipalities': [m['field_value'] for m in todos_municipios],
                'selected_municipality': municipio,
                'total_assets': 0, 'by_condition': [], 'warranty_valid': 0,
                'warranty_expiring_30d': 0, 'warranty_expiring_assets': [], 'assets_by_condition': []
            })
        
        placeholders = ','.join(['?' for _ in lista_ids])
        total_ativos = len(lista_ids)
        
        condicoes = bd.execute(f'''
            SELECT field_value, COUNT(*) as count FROM asset_data 
            WHERE field_name = 'condition_status' AND asset_id IN ({placeholders})
            GROUP BY field_value
        ''', lista_ids).fetchall()
        
        garantia_valida = bd.execute(f'''
            SELECT COUNT(*) FROM asset_data 
            WHERE field_name = 'warranty_end_date' AND field_value >= date('now') AND asset_id IN ({placeholders})
        ''', lista_ids).fetchone()[0]
        
        garantia_expirar = bd.execute(f'''
            SELECT ad.asset_id, ad.field_value as warranty_end, a.serial_number
            FROM asset_data ad JOIN assets a ON ad.asset_id = a.id
            WHERE ad.field_name = 'warranty_end_date' 
            AND ad.field_value >= date('now') AND ad.field_value <= date('now', '+30 days')
            AND ad.asset_id IN ({placeholders}) ORDER BY ad.field_value
        ''', lista_ids).fetchall()
        
        ativos_por_condicao = []
        for cond in condicoes:
            ativos_cond = bd.execute(f'''
                SELECT a.serial_number, a.id FROM assets a
                JOIN asset_data ad ON a.id = ad.asset_id
                WHERE ad.field_name = 'condition_status' AND ad.field_value = ? AND a.id IN ({placeholders})
                LIMIT 10
            ''', [cond['field_value']] + lista_ids).fetchall()
            ativos_por_condicao.append({
                'condition': cond['field_value'], 'count': cond['count'],
                'sample_assets': [{'serial_number': a['serial_number'], 'id': a['id']} for a in ativos_cond]
            })
    else:
        total_ativos = bd.execute('SELECT COUNT(*) FROM assets').fetchone()[0]
        
        condicoes = bd.execute('''
            SELECT field_value, COUNT(*) as count FROM asset_data 
            WHERE field_name = 'condition_status' GROUP BY field_value
        ''').fetchall()
        
        garantia_valida = bd.execute('''
            SELECT COUNT(*) FROM asset_data WHERE field_name = 'warranty_end_date' AND field_value >= date('now')
        ''').fetchone()[0]
        
        garantia_expirar = bd.execute('''
            SELECT ad.asset_id, ad.field_value as warranty_end, a.serial_number
            FROM asset_data ad JOIN assets a ON ad.asset_id = a.id
            WHERE ad.field_name = 'warranty_end_date' 
            AND ad.field_value >= date('now') AND ad.field_value <= date('now', '+30 days')
            ORDER BY ad.field_value LIMIT 20
        ''').fetchall()
        
        ativos_por_condicao = []
        for cond in condicoes:
            ativos_cond = bd.execute('''
                SELECT a.serial_number, a.id FROM assets a
                JOIN asset_data ad ON a.id = ad.asset_id
                WHERE ad.field_name = 'condition_status' AND ad.field_value = ?
                LIMIT 10
            ''', (cond['field_value'],)).fetchall()
            ativos_por_condicao.append({
                'condition': cond['field_value'], 'count': cond['count'],
                'sample_assets': [{'serial_number': a['serial_number'], 'id': a['id']} for a in ativos_cond]
            })
    
    return jsonify({
        'municipalities': [m['field_value'] for m in todos_municipios],
        'selected_municipality': municipio or 'Todos',
        'total_assets': total_ativos,
        'by_condition': [dict(c) for c in condicoes],
        'warranty_valid': garantia_valida,
        'warranty_expiring_30d': len(garantia_expirar),
        'warranty_expiring_assets': [{'serial_number': w['serial_number'], 'warranty_end': w['warranty_end']} for w in garantia_expirar],
        'assets_by_condition': ativos_por_condicao
    })

# =============================================================================
# CONFIGURAÇÍ•ES DO SISTEMA - Prefixos, Cores e Contadores
# =============================================================================

def obter_proximo_numero(tipo_contador):
    """
    Gera o próximo número sequencial para um tipo de contador.
    Retorna o número formatado com prefixo e zeros Í  esquerda.
    """
    bd = obter_bd()
    
    # Mapear tipo de contador para configurações de prefixo
    config_map = {
        'assets': ('prefix_assets', 'prefix_assets_digits'),
        'int_preventiva': ('prefix_int_preventiva', 'prefix_int_digits'),
        'int_corretiva': ('prefix_int_corretiva', 'prefix_int_digits'),
        'int_substituicao': ('prefix_int_substituicao', 'prefix_int_digits'),
        'int_inspecao': ('prefix_int_inspecao', 'prefix_int_digits'),
    }
    
    if tipo_contador not in config_map:
        raise ValueError(f"Tipo de contador inválido: {tipo_contador}")
    
    prefixo_key, digitos_key = config_map[tipo_contador]
    
    # Obter prefixo e número de dígitos
    prefixo = obter_config(prefixo_key, 'SLP')
    digitos = int(obter_config(digitos_key, '9'))
    
    # Incrementar contador
    bd.execute('''
        UPDATE sequence_counters 
        SET current_value = current_value + 1, updated_at = CURRENT_TIMESTAMP
        WHERE counter_type = ?
    ''', (tipo_contador,))
    
    # Obter novo valor
    resultado = bd.execute(
        'SELECT current_value FROM sequence_counters WHERE counter_type = ?',
        (tipo_contador,)
    ).fetchone()
    
    if not resultado:
        # Criar contador se não existir
        bd.execute('''
            INSERT INTO sequence_counters (counter_type, current_value) VALUES (?, 1)
        ''', (tipo_contador,))
        bd.commit()
        numero = 1
    else:
        numero = resultado['current_value']
    
    bd.commit()
    
    # Formatar número com zeros Í  esquerda
    return f"{prefixo}{str(numero).zfill(digitos)}"

def obter_proximo_numero_preview(tipo_contador):
    """
    Obtém preview do próximo número SEM incrementar o contador.
    """
    bd = obter_bd()
    
    config_map = {
        'assets': ('prefix_assets', 'prefix_assets_digits'),
        'int_preventiva': ('prefix_int_preventiva', 'prefix_int_digits'),
        'int_corretiva': ('prefix_int_corretiva', 'prefix_int_digits'),
        'int_substituicao': ('prefix_int_substituicao', 'prefix_int_digits'),
        'int_inspecao': ('prefix_int_inspecao', 'prefix_int_digits'),
    }
    
    if tipo_contador not in config_map:
        return None
    
    prefixo_key, digitos_key = config_map[tipo_contador]
    prefixo = obter_config(prefixo_key, 'SLP')
    digitos = int(obter_config(digitos_key, '9'))
    
    resultado = bd.execute(
        'SELECT current_value FROM sequence_counters WHERE counter_type = ?',
        (tipo_contador,)
    ).fetchone()
    
    proximo = (resultado['current_value'] if resultado else 0) + 1
    return f"{prefixo}{str(proximo).zfill(digitos)}"


@app.route('/api/config/prefixes', methods=['GET'])
@requer_autenticacao
def obter_prefixos():
    """Devolve todas as configurações de prefixos."""
    bd = obter_bd()
    
    configs = bd.execute('''
        SELECT config_key, config_value, description, updated_at
        FROM system_config
        WHERE config_key LIKE 'prefix_%'
        ORDER BY config_key
    ''').fetchall()
    
    prefixes = {}
    for c in configs:
        prefixes[c['config_key']] = c['config_value']
    
    return jsonify({'prefixes': prefixes})


@app.route('/api/config/prefixes', methods=['PUT'])
@requer_admin
def atualizar_prefixos():
    """Atualiza configurações de prefixos."""
    dados = request.json
    bd = obter_bd()
    user_id = request.user_id
    
    campos_validos = [
        'prefix_assets', 'prefix_assets_digits',
        'prefix_int_preventiva', 'prefix_int_corretiva',
        'prefix_int_substituicao', 'prefix_int_inspecao', 'prefix_int_digits'
    ]
    
    for campo in campos_validos:
        if campo in dados:
            bd.execute('''
                UPDATE system_config 
                SET config_value = ?, updated_at = CURRENT_TIMESTAMP, updated_by = ?
                WHERE config_key = ?
            ''', (dados[campo], user_id, campo))
    
    bd.commit()
    return jsonify({'message': 'Prefixos atualizados com sucesso'})


@app.route('/api/config/colors', methods=['GET'])
@requer_autenticacao
def obter_cores():
    """Devolve lista de cores configuradas."""
    resultado = obter_config('colors_list', '[]')
    try:
        cores = json.loads(resultado)
    except Exception:
        cores = []
    return jsonify({'colors': cores})


@app.route('/api/config/colors', methods=['PUT'])
@requer_admin
def atualizar_cores():
    """Atualiza lista de cores."""
    dados = request.json
    cores = dados.get('colors', [])
    bd = obter_bd()
    user_id = request.user_id
    
    bd.execute('''
        UPDATE system_config 
        SET config_value = ?, updated_at = CURRENT_TIMESTAMP, updated_by = ?
        WHERE config_key = 'colors_list'
    ''', (json.dumps(cores), user_id))
    
    bd.commit()
    return jsonify({'message': 'Cores atualizadas com sucesso'})


@app.route('/api/config/counters', methods=['GET'])
@requer_autenticacao
def obter_contadores():
    """Devolve estado atual dos contadores."""
    bd = obter_bd()
    contadores = bd.execute('''
        SELECT counter_type, current_value, updated_at
        FROM sequence_counters
        ORDER BY counter_type
    ''').fetchall()
    
    return jsonify({
        'counters': [dict(c) for c in contadores]
    })


@app.route('/api/config/counters/<string:counter_type>', methods=['PUT'])
@requer_admin
def atualizar_contador(counter_type):
    """Atualiza manualmente o valor de um contador."""
    dados = request.json
    novo_valor = dados.get('value')
    
    if novo_valor is None or not isinstance(novo_valor, int) or novo_valor < 0:
        return jsonify({'error': 'Valor inválido'}), 400
    
    bd = obter_bd()
    user_id = request.user_id
    
    # Verificar se contador existe
    resultado = bd.execute(
        'SELECT current_value FROM sequence_counters WHERE counter_type = ?',
        (counter_type,)
    ).fetchone()
    
    if not resultado:
        return jsonify({'error': f'Contador {counter_type} não encontrado'}), 404
    
    valor_anterior = resultado['current_value']
    
    bd.execute('''
        UPDATE sequence_counters 
        SET current_value = ?, updated_at = CURRENT_TIMESTAMP
        WHERE counter_type = ?
    ''', (novo_valor, counter_type))
    
    registar_auditoria(bd, user_id, 'COUNTER_UPDATE', 'sequence_counters', None, 
                      {'counter_type': counter_type, 'old_value': valor_anterior},
                      {'counter_type': counter_type, 'new_value': novo_valor})
    
    bd.commit()
    return jsonify({
        'message': f'Contador {counter_type} atualizado',
        'old_value': valor_anterior,
        'new_value': novo_valor
    })


@app.route('/api/config/generate-number/<string:tipo>', methods=['POST'])
@requer_autenticacao
def gerar_numero(tipo):
    """
    Gera o próximo número sequencial para um tipo específico.
    Tipos: assets, int_preventiva, int_corretiva, int_substituicao, int_inspecao
    """
    tipos_validos = ['assets', 'int_preventiva', 'int_corretiva', 'int_substituicao', 'int_inspecao']
    
    if tipo not in tipos_validos:
        return jsonify({'error': f'Tipo inválido. Tipos válidos: {", ".join(tipos_validos)}'}), 400
    
    try:
        numero = obter_proximo_numero(tipo)
        return jsonify({'number': numero, 'type': tipo})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/config/next-number/<string:tipo>', methods=['GET'])
@requer_autenticacao
def preview_proximo_numero(tipo):
    """
    Preview do próximo número SEM incrementar o contador.
    Usado para mostrar sugestão antes de criar ativo/intervenção.
    """
    tipos_validos = ['assets', 'int_preventiva', 'int_corretiva', 'int_substituicao', 'int_inspecao']
    
    if tipo not in tipos_validos:
        return jsonify({'error': f'Tipo inválido'}), 400
    
    numero = obter_proximo_numero_preview(tipo)
    return jsonify({'number': numero, 'type': tipo})

# =============================================================================
# MAPA GPS
# =============================================================================

@app.route('/api/assets/map', methods=['GET'])
@requer_autenticacao
def obter_ativos_mapa():
    """Devolve ativos com coordenadas GPS para visualização no mapa."""
    bd = obter_bd()
    municipio = request.args.get('municipality', '')
    estado = request.args.get('status', '')
    
    # Query base - usar 'status' em vez de 'condition_status'
    query = '''
        SELECT a.id, a.serial_number,
               lat.field_value as latitude,
               lon.field_value as longitude,
               loc.field_value as location,
               stat.field_value as status,
               ref.field_value as product_reference,
               mun.field_value as municipality,
               pw.field_value as power_watts,
               rfid.field_value as rfid_tag
        FROM assets a
        LEFT JOIN asset_data lat ON a.id = lat.asset_id AND lat.field_name = 'gps_latitude'
        LEFT JOIN asset_data lon ON a.id = lon.asset_id AND lon.field_name = 'gps_longitude'
        LEFT JOIN asset_data loc ON a.id = loc.asset_id AND loc.field_name = 'address'
        LEFT JOIN asset_data stat ON a.id = stat.asset_id AND stat.field_name = 'status'
        LEFT JOIN asset_data ref ON a.id = ref.asset_id AND ref.field_name = 'product_reference'
        LEFT JOIN asset_data mun ON a.id = mun.asset_id AND mun.field_name = 'municipality'
        LEFT JOIN asset_data pw ON a.id = pw.asset_id AND pw.field_name = 'power_watts'
        LEFT JOIN asset_data rfid ON a.id = rfid.asset_id AND rfid.field_name = 'rfid_tag'
        WHERE lat.field_value IS NOT NULL 
          AND lon.field_value IS NOT NULL
          AND lat.field_value != ''
          AND lon.field_value != ''
    '''
    
    params = []
    
    if municipio:
        query += ' AND mun.field_value = ?'
        params.append(municipio)
    
    if estado:
        query += ' AND stat.field_value = ?'
        params.append(estado)
    
    query += ' ORDER BY a.serial_number'
    
    ativos = bd.execute(query, params).fetchall()
    
    # Converter para lista de dicionários
    resultado = []
    for a in ativos:
        try:
            lat = float(a['latitude']) if a['latitude'] else None
            lon = float(a['longitude']) if a['longitude'] else None
            if lat is not None and lon is not None:
                resultado.append({
                    'id': a['id'],
                    'serial_number': a['serial_number'],
                    'latitude': lat,
                    'longitude': lon,
                    'location': a['location'] or '',
                    'status': a['status'] or 'N/D',
                    'product_reference': a['product_reference'] or '',
                    'municipality': a['municipality'] or '',
                    'power_watts': a['power_watts'] or '',
                    'rfid_tag': a['rfid_tag'] or ''
                })
        except (ValueError, TypeError):
            continue
    
    # Obter lista de municípios para filtro
    municipios = bd.execute('''
        SELECT DISTINCT field_value FROM asset_data 
        WHERE field_name = 'municipality' AND field_value IS NOT NULL AND field_value != ''
        ORDER BY field_value
    ''').fetchall()
    
    # Obter lista de estados para filtro
    estados = bd.execute('''
        SELECT DISTINCT field_value FROM asset_data 
        WHERE field_name = 'condition_status' AND field_value IS NOT NULL AND field_value != ''
        ORDER BY field_value
    ''').fetchall()
    
    return jsonify({
        'assets': resultado,
        'total': len(resultado),
        'filters': {
            'municipalities': [m['field_value'] for m in municipios],
            'statuses': [s['field_value'] for s in estados]
        }
    })

# =============================================================================
# AUDITORIA E SAÍšDE DO SISTEMA
# =============================================================================

@app.route('/api/audit', methods=['GET'])
@requer_admin
def obter_auditoria():
    """Devolve o log de auditoria."""
    bd = obter_bd()
    pagina = request.args.get('page', 1, type=int)
    por_pagina = request.args.get('per_page', 100, type=int)
    offset = (pagina - 1) * por_pagina
    
    logs = bd.execute('''
        SELECT a.*, COALESCE(u.first_name || ' ' || u.last_name, u.email) FROM audit_log a
        LEFT JOIN users u ON a.user_id = u.id
        ORDER BY a.timestamp DESC LIMIT ? OFFSET ?
    ''', (por_pagina, offset)).fetchall()
    
    return jsonify([dict(l) for l in logs])

@app.route('/api/health', methods=['GET'])
def verificar_saude():
    """Verifica o estado do sistema."""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'features': {'excel_export': EXCEL_DISPONIVEL, 'backup': True, 'import': True},
        'backup_info': {'max_backups': MAX_BACKUPS, 'backup_dir': PASTA_BACKUPS}
    })

# =============================================================================
# ASSETS ESTÍTICOS (LOGO)
# =============================================================================

@app.route('/api/logo')
def servir_logo():
    """Serve a logo da empresa baseada no tenant do token."""
    # Tentar obter tenant_id do header ou do token
    tenant_id = request.headers.get('X-Tenant-ID', '')
    
    if not tenant_id:
        # Tentar obter do token
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        if token:
            dados_tenants = carregar_tenants()
            for tenant_info in dados_tenants.get('tenants', []):
                tid = tenant_info['id']
                try:
                    bd = obter_bd(tid)
                    sessao = bd.execute('SELECT user_id FROM sessions WHERE token = ?', (token,)).fetchone()
                    if sessao:
                        tenant_id = tid
                        break
                except Exception:
                    continue
    
    if not tenant_id:
        tenant_id = MASTER_TENANT_ID
    
    # Procurar logo na pasta do tenant
    logo_path = os.path.join(PASTA_TENANTS, tenant_id, 'logo.png')
    if os.path.exists(logo_path):
        return send_file(logo_path, mimetype='image/png')
    
    # Fallback para pasta assets (compatibilidade)
    logo_path_old = os.path.join(PASTA_ASSETS, 'logo.png')
    if os.path.exists(logo_path_old):
        return send_file(logo_path_old, mimetype='image/png')
    
    return '', 404


# =============================================================================
# CATÍLOGO DE REFERÊNCIAS SmartLamppost - REESTRUTURADO
# =============================================================================

@app.route('/api/catalog/columns', methods=['GET'])
@requer_autenticacao
def obter_colunas_catalogo():
    """Devolve lista de colunas do catálogo, opcionalmente filtrada por pack."""
    bd = obter_bd_catalogo()
    pack = request.args.get('pack', '')
    height = request.args.get('height', type=int)
    
    query = 'SELECT * FROM catalog_columns WHERE 1=1'
    params = []
    
    if pack:
        query += ' AND pack = ?'
        params.append(pack)
    if height:
        query += ' AND height_m = ?'
        params.append(height)
    
    query += ' ORDER BY pack, height_m, reference'
    colunas = bd.execute(query, params).fetchall()
    
    return jsonify([dict(c) for c in colunas])


@app.route('/api/catalog/packs', methods=['GET'])
@requer_autenticacao
def obter_packs_catalogo():
    """Devolve lista de packs disponíveis."""
    bd = obter_bd_catalogo()
    packs = bd.execute('SELECT DISTINCT pack FROM catalog_columns ORDER BY pack').fetchall()
    return jsonify([p['pack'] for p in packs])


@app.route('/api/catalog/luminaires', methods=['GET'])
@requer_autenticacao
def obter_luminarias():
    """Devolve lista de luminárias, filtrada por altura e/ou tipo."""
    bd = obter_bd_catalogo()
    height = request.args.get('height', type=int)
    lum_type = request.args.get('type', '')  # 'Tipo 1' ou 'Tipo 2'
    
    query = 'SELECT * FROM catalog_luminaires WHERE 1=1'
    params = []
    
    if height:
        query += ' AND column_height_m = ?'
        params.append(height)
    if lum_type == 'Tipo 1':
        query += ' AND type_1 = 1'
    elif lum_type == 'Tipo 2':
        query += ' AND type_2 = 1'
    
    query += ' ORDER BY column_height_m, reference'
    luminarias = bd.execute(query, params).fetchall()
    
    return jsonify([dict(l) for l in luminarias])


@app.route('/api/catalog/electrical-panels', methods=['GET'])
@requer_autenticacao
def obter_quadros_eletricos():
    """Devolve lista de quadros elétricos."""
    bd = obter_bd_catalogo()
    panels = bd.execute('SELECT * FROM catalog_electrical_panels ORDER BY panel_type, reference').fetchall()
    return jsonify([dict(p) for p in panels])


@app.route('/api/catalog/fuse-boxes', methods=['GET'])
@requer_autenticacao
def obter_cofretes_fusivel():
    """Devolve lista de cofretes fusível, filtrada por tipo se especificado."""
    bd = obter_bd_catalogo()
    fuse_type = request.args.get('type', '')  # 'Tipo S' ou 'Tipo D'
    
    query = 'SELECT * FROM catalog_fuse_boxes WHERE 1=1'
    params = []
    
    if fuse_type == 'Tipo S':
        query += ' AND type_s = 1'
    elif fuse_type == 'Tipo D':
        query += ' AND type_d = 1'
    
    query += ' ORDER BY reference'
    cofretes = bd.execute(query, params).fetchall()
    
    return jsonify([dict(c) for c in cofretes])


@app.route('/api/catalog/telemetry-panels', methods=['GET'])
@requer_autenticacao
def obter_quadros_telemetria():
    """Devolve lista de quadros de telemetria."""
    bd = obter_bd_catalogo()
    panels = bd.execute('SELECT * FROM catalog_telemetry_panels ORDER BY reference').fetchall()
    return jsonify([dict(p) for p in panels])


@app.route('/api/catalog/modules/ev', methods=['GET'])
@requer_autenticacao
def obter_modulos_ev():
    """Devolve lista de módulos EV Charger."""
    bd = obter_bd_catalogo()
    modules = bd.execute('SELECT * FROM catalog_module_ev ORDER BY reference').fetchall()
    return jsonify([dict(m) for m in modules])


@app.route('/api/catalog/modules/mupi', methods=['GET'])
@requer_autenticacao
def obter_modulos_mupi():
    """Devolve lista de módulos MUPI."""
    bd = obter_bd_catalogo()
    modules = bd.execute('SELECT * FROM catalog_module_mupi ORDER BY reference').fetchall()
    return jsonify([dict(m) for m in modules])


@app.route('/api/catalog/modules/lateral', methods=['GET'])
@requer_autenticacao
def obter_modulos_laterais():
    """Devolve lista de módulos laterais."""
    bd = obter_bd_catalogo()
    modules = bd.execute('SELECT * FROM catalog_module_lateral ORDER BY reference').fetchall()
    return jsonify([dict(m) for m in modules])


@app.route('/api/catalog/modules/antenna', methods=['GET'])
@requer_autenticacao
def obter_modulos_antena():
    """Devolve lista de módulos de antena, opcionalmente filtrada por altura."""
    bd = obter_bd_catalogo()
    height = request.args.get('height', type=int)
    
    if height:
        modules = bd.execute('''
            SELECT * FROM catalog_module_antenna WHERE column_height_m = ? ORDER BY reference
        ''', (height,)).fetchall()
    else:
        modules = bd.execute('SELECT * FROM catalog_module_antenna ORDER BY column_height_m, reference').fetchall()
    
    return jsonify([dict(m) for m in modules])


@app.route('/api/catalog/import', methods=['POST'])
@requer_autenticacao
def importar_catalogo():
    """Importa catálogo a partir do ficheiro Excel REFEMOD."""
    if 'file' not in request.files:
        return jsonify({'error': 'Ficheiro não fornecido'}), 400
    
    ficheiro = request.files['file']
    if not ficheiro.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Formato inválido. Use .xlsx ou .xls'}), 400
    
    if not EXCEL_DISPONIVEL:
        return jsonify({'error': 'openpyxl não disponível no servidor'}), 500
    
    # Verificar se deve substituir tudo
    replace_all = request.args.get('replace', 'true').lower() == 'true'
    
    try:
        import openpyxl
        
        # Guardar temporariamente
        caminho_temp = os.path.join(PASTA_BACKUPS, 'temp_catalog.xlsx')
        ficheiro.save(caminho_temp)
        
        wb = openpyxl.load_workbook(caminho_temp)
        bd = obter_bd_catalogo()
        stats = {'columns': 0, 'luminaires': 0, 'electrical': 0, 'fuse_boxes': 0, 
                 'telemetry': 0, 'ev': 0, 'mupi': 0, 'lateral': 0, 'antenna': 0}
        errors = []
        
        print(f"[CATALOG IMPORT] Folhas encontradas: {wb.sheetnames}")
        
        # Se replace_all, limpar todas as tabelas primeiro
        if replace_all:
            bd.execute('DELETE FROM catalog_columns')
            bd.execute('DELETE FROM catalog_luminaires')
            bd.execute('DELETE FROM catalog_electrical_panels')
            bd.execute('DELETE FROM catalog_fuse_boxes')
            bd.execute('DELETE FROM catalog_telemetry_panels')
            bd.execute('DELETE FROM catalog_module_ev')
            bd.execute('DELETE FROM catalog_module_mupi')
            bd.execute('DELETE FROM catalog_module_lateral')
            bd.execute('DELETE FROM catalog_module_antenna')
            print("[CATALOG IMPORT] Tabelas limpas - modo substituição total")
        
        # Helper para verificar Sim/Não
        def is_yes(val):
            if val is None:
                return 0
            return 1 if str(val).strip().lower() in ['sim', 'yes', '1', 'true', 's'] else 0
        
        def get_val(val, default=''):
            return str(val).strip() if val is not None else default
        
        # 1. Importar Colunas (folha "Coluna")
        if 'Coluna' in wb.sheetnames:
            ws = wb['Coluna']
            rows = list(ws.iter_rows(min_row=3, values_only=True))  # Começa na linha 3 (header está na linha 2)
            for row in rows:
                if not row or len(row) < 18 or not row[1]:  # Referência está na coluna B (index 1)
                    continue
                try:
                    # Mapear colunas conforme o Excel REFEMOD ATUALIZADO
                    # A=Descrição, B=Referência, C=Pack, D=Tipo, E=Fixação, F=Altura, G=Braço Luminária,
                    # H=Braço Rua, I=Braço Passeio, J=Luminária Incluída (NOVO), K=Mod1, L=Mod2, M=Mod3, 
                    # N=Mod4, O=Mod5, P=Mod6, Q=Mod7, R=Mod8
                    bd.execute('''
                        INSERT OR REPLACE INTO catalog_columns 
                        (description, reference, pack, column_type, fixing, height_m,
                         arm_count, arm_street, arm_sidewalk, luminaire_included,
                         mod1_luminaire, mod2_electrical, mod3_fuse_box, mod4_telemetry,
                         mod5_ev, mod6_mupi, mod7_lateral, mod8_antenna)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        get_val(row[0]),                    # Descrição
                        get_val(row[1]),                    # Referência
                        get_val(row[2], 'BAREBONE'),        # Pack
                        get_val(row[3], 'Standard'),        # Tipo
                        get_val(row[4], 'Flange'),          # Fixação
                        int(row[5]) if row[5] else 0,       # Altura (m)
                        int(row[6]) if row[6] else 0,       # Braço Luminária (total)
                        int(row[7]) if row[7] else 0,       # Braço Rua
                        int(row[8]) if row[8] else 0,       # Braço Passeio
                        get_val(row[9], 'Não'),             # Luminária Incluída (Sim/Não)
                        get_val(row[10], 'Não'),            # Mod 1 Luminária (Tipo 1, Tipo 2, Não)
                        get_val(row[11], 'Não'),            # Mod 2 Q. Elétrico (Sim/Não)
                        get_val(row[12], 'Não'),            # Mod 3 Cofrete (Tipo S, Tipo D, Não)
                        get_val(row[13], 'Não'),            # Mod 4 Telemetria (Sim/Não)
                        get_val(row[14], 'Não'),            # Mod 5 CVE (Sim/Não)
                        get_val(row[15], 'Não'),            # Mod 6 MUPI (Sim/Não)
                        get_val(row[16], 'Sim'),            # Mod 7 Lateral (Sim)
                        get_val(row[17], 'Sim')             # Mod 8 Antena (Sim)
                    ))
                    stats['columns'] += 1
                except Exception as e:
                    errors.append(f"Coluna {row[1]}: {str(e)}")
            print(f"[CATALOG IMPORT] Colunas importadas: {stats['columns']}")
        
        # 2. Importar Luminárias (folha "Mod. 1 Luminária")
        sheet_lum = None
        for name in wb.sheetnames:
            if 'luminária' in name.lower() or 'luminaria' in name.lower():
                sheet_lum = name
                break
        
        if sheet_lum:
            ws = wb[sheet_lum]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 5 or not row[2]:
                    continue
                try:
                    # A=Tipo, B=Descrição, C=Referência, D=Ref.Fabricante, E=Altura, F=Tipo1, G=Tipo2
                    bd.execute('''
                        INSERT OR REPLACE INTO catalog_luminaires 
                        (luminaire_type, description, reference, manufacturer_ref, column_height_m, type_1, type_2)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        get_val(row[0]),
                        get_val(row[1]),
                        get_val(row[2]),
                        get_val(row[3]) if len(row) > 3 else '',
                        int(row[4]) if len(row) > 4 and row[4] else None,
                        is_yes(row[5]) if len(row) > 5 else 0,
                        is_yes(row[6]) if len(row) > 6 else 0
                    ))
                    stats['luminaires'] += 1
                except Exception as e:
                    errors.append(f"Luminária {row[2]}: {str(e)}")
            print(f"[CATALOG IMPORT] Luminárias importadas: {stats['luminaires']}")
        
        # 3. Importar Quadros Elétricos (folha "Mod. 2 Quando Elétrico")
        sheet_elec = None
        for name in wb.sheetnames:
            if 'létrico' in name.lower() or 'eletrico' in name.lower() or 'elétrico' in name.lower():
                sheet_elec = name
                break
        
        if sheet_elec:
            ws = wb[sheet_elec]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 3 or not row[2]:
                    continue
                try:
                    ref = get_val(row[2])
                    # Criar short_reference removendo SLP- do início
                    short_ref = ref.replace('SLP-', '') if ref.startswith('SLP-') else ref
                    bd.execute('''
                        INSERT OR REPLACE INTO catalog_electrical_panels 
                        (panel_type, description, reference, short_reference)
                        VALUES (?, ?, ?, ?)
                    ''', (get_val(row[0]), get_val(row[1]), ref, short_ref))
                    stats['electrical'] += 1
                except Exception as e:
                    errors.append(f"Q.Elétrico {row[2]}: {str(e)}")
            print(f"[CATALOG IMPORT] Quadros elétricos importados: {stats['electrical']}")
        
        # 4. Importar Cofretes Fusível (folha "Mod. 3 Cofrete Fusível")
        sheet_fuse = None
        for name in wb.sheetnames:
            if 'cofrete' in name.lower() or 'fusível' in name.lower() or 'fusivel' in name.lower():
                sheet_fuse = name
                break
        
        if sheet_fuse:
            ws = wb[sheet_fuse]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 3 or not row[2]:
                    continue
                try:
                    ref = get_val(row[2])
                    short_ref = ref.replace('SLP-', '') if ref.startswith('SLP-') else ref
                    # A=Tipo, B=Descrição, C=Referência, D=TipoS, E=TipoD
                    bd.execute('''
                        INSERT OR REPLACE INTO catalog_fuse_boxes 
                        (fuse_type, description, reference, short_reference, type_s, type_d)
                        VALUES (?, ?, ?, ?, ?, ?)
                    ''', (
                        get_val(row[0]),
                        get_val(row[1]),
                        ref,
                        short_ref,
                        is_yes(row[3]) if len(row) > 3 else 0,
                        is_yes(row[4]) if len(row) > 4 else 0
                    ))
                    stats['fuse_boxes'] += 1
                except Exception as e:
                    errors.append(f"Cofrete {row[2]}: {str(e)}")
            print(f"[CATALOG IMPORT] Cofretes importados: {stats['fuse_boxes']}")
        
        # 5. Importar Quadros Telemetria (folha "Mod. 4 Quadro Telemetria")
        sheet_tel = None
        for name in wb.sheetnames:
            if 'telemetria' in name.lower():
                sheet_tel = name
                break
        
        if sheet_tel:
            ws = wb[sheet_tel]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 3 or not row[2]:
                    continue
                try:
                    ref = get_val(row[2])
                    short_ref = ref.replace('SLP-', '') if ref.startswith('SLP-') else ref
                    bd.execute('''
                        INSERT OR REPLACE INTO catalog_telemetry_panels 
                        (panel_type, description, reference, short_reference)
                        VALUES (?, ?, ?, ?)
                    ''', (get_val(row[0]), get_val(row[1]), ref, short_ref))
                    stats['telemetry'] += 1
                except Exception as e:
                    errors.append(f"Telemetria {row[2]}: {str(e)}")
            print(f"[CATALOG IMPORT] Quadros telemetria importados: {stats['telemetry']}")
        
        # 6. Importar Módulos EV (folha "Mod. 5 CVE")
        sheet_ev = None
        for name in wb.sheetnames:
            if 'cve' in name.lower() or 'mod. 5' in name.lower():
                sheet_ev = name
                break
        
        if sheet_ev:
            ws = wb[sheet_ev]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 3 or not row[2]:
                    continue
                try:
                    ref = get_val(row[2])
                    short_ref = ref.replace('SLPS-', '').replace('SLP-', '') if ref.startswith('SLP') else ref
                    bd.execute('''
                        INSERT OR REPLACE INTO catalog_module_ev 
                        (module_type, description, reference, short_reference)
                        VALUES (?, ?, ?, ?)
                    ''', (get_val(row[0]), get_val(row[1]), ref, short_ref))
                    stats['ev'] += 1
                except Exception as e:
                    errors.append(f"Módulo EV {row[2]}: {str(e)}")
            print(f"[CATALOG IMPORT] Módulos EV importados: {stats['ev']}")
        
        # 7. Importar Módulos MUPI (folha "Mod. 6 MUPI")
        sheet_mupi = None
        for name in wb.sheetnames:
            if 'mupi' in name.lower() or 'mod. 6' in name.lower():
                sheet_mupi = name
                break
        
        if sheet_mupi:
            ws = wb[sheet_mupi]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 3 or not row[2]:
                    continue
                try:
                    ref = get_val(row[2])
                    short_ref = ref.replace('SLPS-', '').replace('SLP-', '') if ref.startswith('SLP') else ref
                    bd.execute('''
                        INSERT OR REPLACE INTO catalog_module_mupi 
                        (module_type, description, reference, short_reference)
                        VALUES (?, ?, ?, ?)
                    ''', (get_val(row[0]), get_val(row[1]), ref, short_ref))
                    stats['mupi'] += 1
                except Exception as e:
                    errors.append(f"Módulo MUPI {row[2]}: {str(e)}")
            print(f"[CATALOG IMPORT] Módulos MUPI importados: {stats['mupi']}")
        
        # 8. Importar Módulos Laterais (folha "Mod. 7 Lateral")
        sheet_lat = None
        for name in wb.sheetnames:
            if 'lateral' in name.lower() or 'mod. 7' in name.lower():
                sheet_lat = name
                break
        
        if sheet_lat:
            ws = wb[sheet_lat]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 3 or not row[2]:
                    continue
                try:
                    ref = get_val(row[2])
                    short_ref = ref.replace('SLP-', '') if ref.startswith('SLP-') else ref
                    bd.execute('''
                        INSERT OR REPLACE INTO catalog_module_lateral 
                        (module_type, description, reference, short_reference)
                        VALUES (?, ?, ?, ?)
                    ''', (get_val(row[0]), get_val(row[1]), ref, short_ref))
                    stats['lateral'] += 1
                except Exception as e:
                    errors.append(f"Módulo Lateral {row[2]}: {str(e)}")
            print(f"[CATALOG IMPORT] Módulos Laterais importados: {stats['lateral']}")
        
        # 9. Importar Módulos Antena (folha "Mod. 8 Cápsula Antena")
        sheet_ant = None
        for name in wb.sheetnames:
            if 'antena' in name.lower() or 'mod. 8' in name.lower() or 'cápsula' in name.lower():
                sheet_ant = name
                break
        
        if sheet_ant:
            ws = wb[sheet_ant]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 3 or not row[2]:
                    continue
                try:
                    ref = get_val(row[2])
                    short_ref = ref.replace('SLPC-', '').replace('SLP-', '') if ref.startswith('SLP') else ref
                    height = int(row[3]) if len(row) > 3 and row[3] else None
                    bd.execute('''
                        INSERT OR REPLACE INTO catalog_module_antenna 
                        (module_type, description, reference, short_reference, column_height_m)
                        VALUES (?, ?, ?, ?, ?)
                    ''', (get_val(row[0]), get_val(row[1]), ref, short_ref, height))
                    stats['antenna'] += 1
                except Exception as e:
                    errors.append(f"Módulo Antena {row[2]}: {str(e)}")
            print(f"[CATALOG IMPORT] Módulos Antena importados: {stats['antenna']}")
        
        bd.commit()
        
        # Limpar ficheiro temporário
        if os.path.exists(caminho_temp):
            os.remove(caminho_temp)
        
        result = {
            'success': True,
            'message': 'Catálogo importado com sucesso',
            'stats': stats,
            'replace_mode': replace_all
        }
        
        if errors:
            result['warnings'] = errors[:20]
            result['total_errors'] = len(errors)
        
        return jsonify(result)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Erro ao importar: {str(e)}'}), 500


@app.route('/api/catalog/export', methods=['GET'])
@requer_autenticacao
def exportar_catalogo():
    """Exporta catálogo para Excel."""
    if not EXCEL_DISPONIVEL:
        return jsonify({'error': 'openpyxl não disponível'}), 500
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        bd = obter_bd_catalogo()
        wb = openpyxl.Workbook()
        
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        def add_headers(ws, headers):
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
        
        # 1. Folha de Colunas
        ws = wb.active
        ws.title = 'Coluna'
        headers = ['Descrição', 'Referência Coluna', 'Pack', 'Tipo', 'Fixação', 'Altura (m)', 
                   'Braço Luminária', 'Braço Rua', 'Braço Passeio', 'Luminária Incluída',
                   'Mod. 1 Luminária', 'Mod. 2 Q. Elétrico', 'Mod. 3 Cofrete', 
                   'Mod. 4 Telemetria', 'Mod. 5 CVE', 'Mod. 6 MUPI', 'Mod. 7 Lateral', 'Mod. 8 Antena']
        add_headers(ws, headers)
        
        colunas = bd.execute('SELECT * FROM catalog_columns ORDER BY pack, height_m, reference').fetchall()
        for row_num, col_data in enumerate(colunas, 2):
            ws.cell(row=row_num, column=1, value=col_data['description'])
            ws.cell(row=row_num, column=2, value=col_data['reference'])
            ws.cell(row=row_num, column=3, value=col_data['pack'])
            ws.cell(row=row_num, column=4, value=col_data['column_type'])
            ws.cell(row=row_num, column=5, value=col_data['fixing'])
            ws.cell(row=row_num, column=6, value=col_data['height_m'])
            ws.cell(row=row_num, column=7, value=col_data['arm_count'])
            ws.cell(row=row_num, column=8, value=col_data['arm_street'])
            ws.cell(row=row_num, column=9, value=col_data['arm_sidewalk'])
            ws.cell(row=row_num, column=10, value=col_data['luminaire_included'])
            ws.cell(row=row_num, column=11, value=col_data['mod1_luminaire'])
            ws.cell(row=row_num, column=12, value=col_data['mod2_electrical'])
            ws.cell(row=row_num, column=13, value=col_data['mod3_fuse_box'])
            ws.cell(row=row_num, column=14, value=col_data['mod4_telemetry'])
            ws.cell(row=row_num, column=15, value=col_data['mod5_ev'])
            ws.cell(row=row_num, column=16, value=col_data['mod6_mupi'])
            ws.cell(row=row_num, column=17, value=col_data['mod7_lateral'])
            ws.cell(row=row_num, column=18, value=col_data['mod8_antenna'])
        
        # 2. Luminárias
        ws = wb.create_sheet('Mod. 1 Luminária')
        headers = ['Tipo', 'Descrição', 'Referência', 'Ref. Fabricante', 'Altura (m)', 'Tipo 1', 'Tipo 2']
        add_headers(ws, headers)
        items = bd.execute('SELECT * FROM catalog_luminaires ORDER BY column_height_m, reference').fetchall()
        for row_num, item in enumerate(items, 2):
            ws.cell(row=row_num, column=1, value=item['luminaire_type'])
            ws.cell(row=row_num, column=2, value=item['description'])
            ws.cell(row=row_num, column=3, value=item['reference'])
            ws.cell(row=row_num, column=4, value=item['manufacturer_ref'])
            ws.cell(row=row_num, column=5, value=item['column_height_m'])
            ws.cell(row=row_num, column=6, value='Sim' if item['type_1'] else 'Não')
            ws.cell(row=row_num, column=7, value='Sim' if item['type_2'] else 'Não')
        
        # 3. Quadros Elétricos
        ws = wb.create_sheet('Mod. 2 Quadro Elétrico')
        headers = ['Tipo', 'Descrição', 'Referência']
        add_headers(ws, headers)
        items = bd.execute('SELECT * FROM catalog_electrical_panels ORDER BY reference').fetchall()
        for row_num, item in enumerate(items, 2):
            ws.cell(row=row_num, column=1, value=item['panel_type'])
            ws.cell(row=row_num, column=2, value=item['description'])
            ws.cell(row=row_num, column=3, value=item['reference'])
        
        # 4. Cofretes Fusível
        ws = wb.create_sheet('Mod. 3 Cofrete Fusível')
        headers = ['Tipo', 'Descrição', 'Referência', 'Tipo S', 'Tipo D']
        add_headers(ws, headers)
        items = bd.execute('SELECT * FROM catalog_fuse_boxes ORDER BY reference').fetchall()
        for row_num, item in enumerate(items, 2):
            ws.cell(row=row_num, column=1, value=item['fuse_type'])
            ws.cell(row=row_num, column=2, value=item['description'])
            ws.cell(row=row_num, column=3, value=item['reference'])
            ws.cell(row=row_num, column=4, value='Sim' if item['type_s'] else 'Não')
            ws.cell(row=row_num, column=5, value='Sim' if item['type_d'] else 'Não')
        
        # 5. Quadros Telemetria
        ws = wb.create_sheet('Mod. 4 Quadro Telemetria')
        headers = ['Tipo', 'Descrição', 'Referência']
        add_headers(ws, headers)
        items = bd.execute('SELECT * FROM catalog_telemetry_panels ORDER BY reference').fetchall()
        for row_num, item in enumerate(items, 2):
            ws.cell(row=row_num, column=1, value=item['panel_type'])
            ws.cell(row=row_num, column=2, value=item['description'])
            ws.cell(row=row_num, column=3, value=item['reference'])
        
        # 6. Módulos EV
        ws = wb.create_sheet('Mod. 5 CVE')
        headers = ['Tipo', 'Descrição', 'Referência']
        add_headers(ws, headers)
        items = bd.execute('SELECT * FROM catalog_module_ev ORDER BY reference').fetchall()
        for row_num, item in enumerate(items, 2):
            ws.cell(row=row_num, column=1, value=item['module_type'])
            ws.cell(row=row_num, column=2, value=item['description'])
            ws.cell(row=row_num, column=3, value=item['reference'])
        
        # 7. Módulos MUPI
        ws = wb.create_sheet('Mod. 6 MUPI')
        headers = ['Tipo', 'Descrição', 'Referência']
        add_headers(ws, headers)
        items = bd.execute('SELECT * FROM catalog_module_mupi ORDER BY reference').fetchall()
        for row_num, item in enumerate(items, 2):
            ws.cell(row=row_num, column=1, value=item['module_type'])
            ws.cell(row=row_num, column=2, value=item['description'])
            ws.cell(row=row_num, column=3, value=item['reference'])
        
        # 8. Módulos Laterais
        ws = wb.create_sheet('Mod. 7 Lateral')
        headers = ['Tipo', 'Descrição', 'Referência']
        add_headers(ws, headers)
        items = bd.execute('SELECT * FROM catalog_module_lateral ORDER BY reference').fetchall()
        for row_num, item in enumerate(items, 2):
            ws.cell(row=row_num, column=1, value=item['module_type'])
            ws.cell(row=row_num, column=2, value=item['description'])
            ws.cell(row=row_num, column=3, value=item['reference'])
        
        # 9. Módulos Antena
        ws = wb.create_sheet('Mod. 8 Cápsula Antena')
        headers = ['Tipo', 'Descrição', 'Referência', 'Altura (m)']
        add_headers(ws, headers)
        items = bd.execute('SELECT * FROM catalog_module_antenna ORDER BY column_height_m, reference').fetchall()
        for row_num, item in enumerate(items, 2):
            ws.cell(row=row_num, column=1, value=item['module_type'])
            ws.cell(row=row_num, column=2, value=item['description'])
            ws.cell(row=row_num, column=3, value=item['reference'])
            ws.cell(row=row_num, column=4, value=item['column_height_m'])
        
        # Guardar
        caminho = os.path.join(PASTA_BACKUPS, f'catalogo_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        wb.save(caminho)
        
        return send_file(caminho, as_attachment=True, download_name='catalogo_smartlamppost.xlsx')
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Erro ao exportar: {str(e)}'}), 500


@app.route('/api/catalog/stats', methods=['GET'])
@requer_autenticacao
def obter_stats_catalogo():
    """Devolve estatísticas do catálogo."""
    bd = obter_bd_catalogo()
    try:
        return jsonify({
            'columns': bd.execute('SELECT COUNT(*) FROM catalog_columns').fetchone()[0],
            'luminaires': bd.execute('SELECT COUNT(*) FROM catalog_luminaires').fetchone()[0],
            'electrical': bd.execute('SELECT COUNT(*) FROM catalog_electrical_panels').fetchone()[0],
            'fuse_boxes': bd.execute('SELECT COUNT(*) FROM catalog_fuse_boxes').fetchone()[0],
            'telemetry': bd.execute('SELECT COUNT(*) FROM catalog_telemetry_panels').fetchone()[0],
            'ev': bd.execute('SELECT COUNT(*) FROM catalog_module_ev').fetchone()[0],
            'mupi': bd.execute('SELECT COUNT(*) FROM catalog_module_mupi').fetchone()[0],
            'lateral': bd.execute('SELECT COUNT(*) FROM catalog_module_lateral').fetchone()[0],
            'antenna': bd.execute('SELECT COUNT(*) FROM catalog_module_antenna').fetchone()[0]
        })
    except Exception as e:
        return jsonify({'columns': 0, 'luminaires': 0, 'electrical': 0, 'fuse_boxes': 0, 
                       'telemetry': 0, 'ev': 0, 'mupi': 0, 'lateral': 0, 'antenna': 0})


@app.route('/api/catalog/item/<tab>', methods=['POST'])
@requer_admin
def adicionar_item_catalogo(tab):
    """Adiciona um item ao catálogo."""
    dados = request.json
    bd = obter_bd_catalogo()
    
    try:
        if tab == 'columns':
            bd.execute('''INSERT INTO catalog_columns 
                (description, reference, pack, column_type, fixing, height_m, arm_count, arm_street, arm_sidewalk,
                 luminaire_included, mod1_luminaire, mod2_electrical, mod3_fuse_box, mod4_telemetry, 
                 mod5_ev, mod6_mupi, mod7_lateral, mod8_antenna)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                (dados.get('description', ''), dados.get('reference', ''), dados.get('pack', 'BAREBONE'),
                 dados.get('column_type', 'Standard'), dados.get('fixing', 'Flange'),
                 dados.get('height_m', 0), dados.get('arm_count', 0), dados.get('arm_street', 0), dados.get('arm_sidewalk', 0),
                 dados.get('luminaire_included', 'Não'), dados.get('mod1_luminaire', 'Não'), dados.get('mod2_electrical', 'Não'),
                 dados.get('mod3_fuse_box', 'Não'), dados.get('mod4_telemetry', 'Não'), dados.get('mod5_ev', 'Não'),
                 dados.get('mod6_mupi', 'Não'), dados.get('mod7_lateral', 'Sim'), dados.get('mod8_antenna', 'Sim')))
        
        elif tab == 'luminaires':
            ref = dados.get('reference', '')
            bd.execute('''INSERT INTO catalog_luminaires 
                (luminaire_type, description, reference, manufacturer_ref, column_height_m, type_1, type_2)
                VALUES (?, ?, ?, ?, ?, ?, ?)''',
                (dados.get('luminaire_type', ''), dados.get('description', ''), ref,
                 dados.get('manufacturer_ref', ''), dados.get('column_height_m'),
                 1 if dados.get('type_1') else 0, 1 if dados.get('type_2') else 0))
        
        elif tab == 'electrical':
            ref = dados.get('reference', '')
            short = dados.get('short_reference') or ref.replace('SLP-', '')
            bd.execute('''INSERT INTO catalog_electrical_panels (panel_type, description, reference, short_reference)
                VALUES (?, ?, ?, ?)''', (dados.get('panel_type', ''), dados.get('description', ''), ref, short))
        
        elif tab == 'fuse_boxes':
            ref = dados.get('reference', '')
            short = dados.get('short_reference') or ref.replace('SLP-', '')
            bd.execute('''INSERT INTO catalog_fuse_boxes (fuse_type, description, reference, short_reference, type_s, type_d)
                VALUES (?, ?, ?, ?, ?, ?)''',
                (dados.get('fuse_type', ''), dados.get('description', ''), ref, short,
                 1 if dados.get('type_s') else 0, 1 if dados.get('type_d') else 0))
        
        elif tab == 'telemetry':
            ref = dados.get('reference', '')
            short = dados.get('short_reference') or ref.replace('SLP-', '')
            bd.execute('''INSERT INTO catalog_telemetry_panels (panel_type, description, reference, short_reference)
                VALUES (?, ?, ?, ?)''', (dados.get('panel_type', ''), dados.get('description', ''), ref, short))
        
        elif tab == 'ev':
            ref = dados.get('reference', '')
            short = dados.get('short_reference') or ref.replace('SLPS-', '').replace('SLP-', '')
            bd.execute('''INSERT INTO catalog_module_ev (module_type, description, reference, short_reference)
                VALUES (?, ?, ?, ?)''', (dados.get('module_type', ''), dados.get('description', ''), ref, short))
        
        elif tab == 'mupi':
            ref = dados.get('reference', '')
            short = dados.get('short_reference') or ref.replace('SLPS-', '').replace('SLP-', '')
            bd.execute('''INSERT INTO catalog_module_mupi (module_type, description, reference, short_reference)
                VALUES (?, ?, ?, ?)''', (dados.get('module_type', ''), dados.get('description', ''), ref, short))
        
        elif tab == 'lateral':
            ref = dados.get('reference', '')
            short = dados.get('short_reference') or ref.replace('SLP-', '')
            bd.execute('''INSERT INTO catalog_module_lateral (module_type, description, reference, short_reference)
                VALUES (?, ?, ?, ?)''', (dados.get('module_type', ''), dados.get('description', ''), ref, short))
        
        elif tab == 'antenna':
            ref = dados.get('reference', '')
            short = dados.get('short_reference') or ref.replace('SLPC-', '').replace('SLP-', '')
            bd.execute('''INSERT INTO catalog_module_antenna (module_type, description, reference, short_reference, column_height_m)
                VALUES (?, ?, ?, ?, ?)''',
                (dados.get('module_type', ''), dados.get('description', ''), ref, short, dados.get('column_height_m')))
        
        else:
            return jsonify({'error': f'Tab desconhecida: {tab}'}), 400
        
        bd.commit()
        return jsonify({'success': True, 'message': 'Item adicionado com sucesso'})
    
    except sqlite3.IntegrityError as e:
        return jsonify({'error': f'Referência já existe: {str(e)}'}), 400
    except Exception as e:
        return jsonify({'error': f'Erro ao adicionar: {str(e)}'}), 500


@app.route('/api/catalog/item/<tab>/<int:item_id>', methods=['DELETE'])
@requer_admin
def remover_item_catalogo(tab, item_id):
    """Remove um item do catálogo."""
    bd = obter_bd_catalogo()
    
    tables = {
        'columns': 'catalog_columns',
        'luminaires': 'catalog_luminaires',
        'electrical': 'catalog_electrical_panels',
        'fuse_boxes': 'catalog_fuse_boxes',
        'telemetry': 'catalog_telemetry_panels',
        'ev': 'catalog_module_ev',
        'mupi': 'catalog_module_mupi',
        'lateral': 'catalog_module_lateral',
        'antenna': 'catalog_module_antenna'
    }
    
    if tab not in tables:
        return jsonify({'error': f'Tab desconhecida: {tab}'}), 400
    
    try:
        bd.execute(f'DELETE FROM {tables[tab]} WHERE id = ?', (item_id,))
        bd.commit()
        return jsonify({'success': True, 'message': 'Item removido com sucesso'})
    except Exception as e:
        return jsonify({'error': f'Erro ao remover: {str(e)}'}), 500


@app.route('/api/catalog/clear', methods=['DELETE'])
@requer_admin
def limpar_catalogo():
    """Limpa todo o catálogo (endpoint alternativo)."""
    try:
        bd = obter_bd_catalogo()
        bd.execute('DELETE FROM catalog_columns')
        bd.execute('DELETE FROM catalog_luminaires')
        bd.execute('DELETE FROM catalog_electrical_panels')
        bd.execute('DELETE FROM catalog_fuse_boxes')
        bd.execute('DELETE FROM catalog_telemetry_panels')
        bd.execute('DELETE FROM catalog_module_ev')
        bd.execute('DELETE FROM catalog_module_mupi')
        bd.execute('DELETE FROM catalog_module_lateral')
        bd.execute('DELETE FROM catalog_module_antenna')
        bd.commit()
        return jsonify({'success': True, 'message': 'Catálogo limpo com sucesso'})
    except Exception as e:
        return jsonify({'error': f'Erro ao limpar: {str(e)}'}), 500


@app.route('/api/catalog/reset', methods=['POST'])
@requer_admin
def resetar_catalogo():
    """Limpa todo o catálogo (apenas admin)."""
    try:
        bd = obter_bd_catalogo()
        
        bd.execute('DELETE FROM catalog_columns')
        bd.execute('DELETE FROM catalog_luminaires')
        bd.execute('DELETE FROM catalog_electrical_panels')
        bd.execute('DELETE FROM catalog_fuse_boxes')
        bd.execute('DELETE FROM catalog_telemetry_panels')
        bd.execute('DELETE FROM catalog_module_ev')
        bd.execute('DELETE FROM catalog_module_mupi')
        bd.execute('DELETE FROM catalog_module_lateral')
        bd.execute('DELETE FROM catalog_module_antenna')
        
        bd.commit()
        
        return jsonify({
            'success': True,
            'message': 'Catálogo limpo com sucesso'
        })
    except Exception as e:
        return jsonify({'error': f'Erro ao limpar catálogo: {str(e)}'}), 500


@app.route('/api/catalog/load-default', methods=['POST'])
@requer_admin
def carregar_catalogo_padrao():
    """Carrega dados padrão do catálogo baseado no Excel REFEMOD."""
    try:
        bd = obter_bd_catalogo()
        stats = {'columns': 0, 'luminaires': 0, 'electrical': 0, 'fuse_boxes': 0,
                 'telemetry': 0, 'ev': 0, 'mupi': 0, 'lateral': 0, 'antenna': 0}
        
        # Limpar tabelas existentes
        bd.execute('DELETE FROM catalog_columns')
        bd.execute('DELETE FROM catalog_luminaires')
        bd.execute('DELETE FROM catalog_electrical_panels')
        bd.execute('DELETE FROM catalog_fuse_boxes')
        bd.execute('DELETE FROM catalog_telemetry_panels')
        bd.execute('DELETE FROM catalog_module_ev')
        bd.execute('DELETE FROM catalog_module_mupi')
        bd.execute('DELETE FROM catalog_module_lateral')
        bd.execute('DELETE FROM catalog_module_antenna')
        
        # =================================================================
        # LUMINÍRIAS (Mod. 1) - Baseado no Excel
        # =================================================================
        luminaires_data = [
            # Tipo, Descrição, Referência, Ref.Fabricante, Altura, Tipo1, Tipo2
            ('Luminária Braço', 'Luminária para colunas de 3 metros', 'LUSA-16-700', 'L75B44261030005', 3, 1, 0),
            ('Luminária Braço', 'Luminária para colunas de 4 metros', 'LUSA-16-700', 'L75B44261030005', 4, 1, 0),
            ('Luminária Braço', 'Luminária para colunas de 6 metros', 'LUSA-24-700', 'L75B44261430005', 6, 1, 0),
            ('Luminária Braço', 'Luminária para colunas de 8 metros', 'LUSA-32-700', 'L75B44262030005', 8, 1, 0),
            ('Luminária Braço', 'Luminária para colunas de 10 metros', 'LUSA-36-700', 'L76B44262630005', 10, 1, 0),
            ('Luminária Topo', 'Luminária 360 para colunas de 4m sem braço', 'APE360', 'LNX114111230000', 4, 0, 1),
        ]
        for ltype, desc, ref, mref, height, t1, t2 in luminaires_data:
            bd.execute('''INSERT INTO catalog_luminaires 
                (luminaire_type, description, reference, manufacturer_ref, column_height_m, type_1, type_2) 
                VALUES (?, ?, ?, ?, ?, ?, ?)''', (ltype, desc, ref, mref, height, t1, t2))
        stats['luminaires'] = len(luminaires_data)
        
        # =================================================================
        # QUADROS ELÉTRICOS (Mod. 2) - Baseado no Excel
        # =================================================================
        electrical_data = [
            ('Q.E. Monofásico', 'Quadro elétrico SLP MONO PLAIN PACK 1', 'SLP-1LPL-10000-040A'),
            ('Q.E. Monofásico', 'Quadro elétrico SLP MONO BASIC PACK 1', 'SLP-1LBA-01000-040A'),
            ('Q.E. Monofásico', 'Quadro elétrico SLP MONO BASIC PACK 2', 'SLP-1LBA-02000-040A'),
            ('Q.E. Monofásico', 'Quadro elétrico SLP MONO CELL PACK 1', 'SLP-1LSC-01100-040A'),
            ('Q.E. Monofásico', 'Quadro elétrico SLP MONO EVC 1x3.7 PACK 1', 'SLP-1LVE-01011-040A'),
            ('Q.E. Monofásico', 'Quadro elétrico SLP MONO EVC 2x3.7 PACK 1', 'SLP-1LVE-01021-040A'),
            ('Q.E. Monofásico', 'Quadro elétrico SLP MONO EVC 1x7.4 PACK 1', 'SLP-1LVE-01012-040A'),
            ('Q.E. Monofásico', 'Quadro elétrico SLP MONO EVC 2x7.4 PACK 1', 'SLP-1LVE-01022-040A'),
            ('Q.E. Trifásico', 'Quadro elétrico SLP TRI PLAIN PACK 1', 'SLP-3LPL-10000-063A'),
            ('Q.E. Trifásico', 'Quadro elétrico SLP TRI BASIC PACK 1', 'SLP-3LBA-01000-063A'),
            ('Q.E. Trifásico', 'Quadro elétrico SLP TRI BASIC PACK 2', 'SLP-3LBA-02000-063A'),
            ('Q.E. Trifásico', 'Quadro elétrico SLP TRI CELL PACK 1', 'SLP-3LSC-01100-063A'),
            ('Q.E. Trifásico', 'Quadro elétrico SLP TRI EVC 1x11 PACK 1', 'SLP-3LVE-01013-063A'),
            ('Q.E. Trifásico', 'Quadro elétrico SLP TRI EVC 2x11 PACK 1', 'SLP-3LVE-01023-063A'),
            ('Q.E. Trifásico', 'Quadro elétrico SLP TRI EVC 1x22 PACK 1', 'SLP-3LVE-01014-063A'),
            ('Q.E. Trifásico', 'Quadro elétrico SLP TRI EVC 2x22 PACK 1', 'SLP-3LVE-01024-063A'),
        ]
        for ptype, desc, ref in electrical_data:
            short = ref.replace('SLP-', '')
            bd.execute('''INSERT INTO catalog_electrical_panels 
                (panel_type, description, reference, short_reference) VALUES (?, ?, ?, ?)''',
                (ptype, desc, ref, short))
        stats['electrical'] = len(electrical_data)
        
        # =================================================================
        # COFRETES FUSÍVEL (Mod. 3) - Baseado no Excel
        # =================================================================
        fuse_data = [
            ('Luminária Simples', 'Cofrete Fusível Amarelo 1', 'SLP-FUSE-1', 1, 0),
            ('Luminária Dupla', 'Cofrete Fusível Amarelo 2', 'SLP-FUSE-2', 0, 1),
        ]
        for ftype, desc, ref, ts, td in fuse_data:
            short = ref.replace('SLP-', '')
            bd.execute('''INSERT INTO catalog_fuse_boxes 
                (fuse_type, description, reference, short_reference, type_s, type_d) VALUES (?, ?, ?, ?, ?, ?)''',
                (ftype, desc, ref, short, ts, td))
        stats['fuse_boxes'] = len(fuse_data)
        
        # =================================================================
        # QUADROS TELEMETRIA (Mod. 4) - Baseado no Excel
        # =================================================================
        telemetry_data = [
            ('Q.E. Telemetria', 'Quadro elétrico SLP TELEMETRIA LITE COMPACT', 'SLP-TEL-LC'),
            ('Q.E. Telemetria', 'Quadro elétrico SLP TELEMETRIA LITE FULL', 'SLP-TEL-LF'),
            ('Q.E. Telemetria', 'Quadro elétrico SLP TELEMETRIA PERFORMANCE COMPACT', 'SLP-TEL-PC'),
            ('Q.E. Telemetria', 'Quadro elétrico SLP TELEMETRIA PERFORMANCE FULL', 'SLP-TEL-PF'),
        ]
        for ptype, desc, ref in telemetry_data:
            short = ref.replace('SLP-', '')
            bd.execute('''INSERT INTO catalog_telemetry_panels 
                (panel_type, description, reference, short_reference) VALUES (?, ?, ?, ?)''',
                (ptype, desc, ref, short))
        stats['telemetry'] = len(telemetry_data)
        
        # =================================================================
        # MÍ“DULOS EV (Mod. 5) - Baseado no Excel
        # =================================================================
        ev_data = [
            ('Módulo CVE', 'SLP Standard EV Charger Module (1x3.7kW)', 'SLPS-EVC137'),
            ('Módulo CVE', 'SLP Standard EV Charger Module (2x3.7kW)', 'SLPS-EVC237'),
            ('Módulo CVE', 'SLP Standard EV Charger Module (1x7.4kW)', 'SLPS-EVC174'),
            ('Módulo CVE', 'SLP Standard EV Charger Module (2x7.4kW)', 'SLPS-EVC274'),
            ('Módulo CVE', 'SLP Standard EV Charger Module (1x11kW)', 'SLPS-EVC111'),
            ('Módulo CVE', 'SLP Standard EV Charger Module (2x11kW)', 'SLPS-EVC211'),
            ('Módulo CVE', 'SLP Standard EV Charger Module (1x22kW)', 'SLPS-EVC122'),
            ('Módulo CVE', 'SLP Standard EV Charger Module (2x22kW)', 'SLPS-EVC222'),
        ]
        for mtype, desc, ref in ev_data:
            short = ref.replace('SLPS-', '')
            bd.execute('''INSERT INTO catalog_module_ev 
                (module_type, description, reference, short_reference) VALUES (?, ?, ?, ?)''',
                (mtype, desc, ref, short))
        stats['ev'] = len(ev_data)
        
        # =================================================================
        # MÍ“DULOS MUPI (Mod. 6) - Baseado no Excel
        # =================================================================
        mupi_data = [
            ('MUPI', 'MUPI Display 32" Inox for SLP Standard Top Hatch', 'SLPS-DISP32'),
        ]
        for mtype, desc, ref in mupi_data:
            short = ref.replace('SLPS-', '')
            bd.execute('''INSERT INTO catalog_module_mupi 
                (module_type, description, reference, short_reference) VALUES (?, ?, ?, ?)''',
                (mtype, desc, ref, short))
        stats['mupi'] = len(mupi_data)
        
        # =================================================================
        # MÍ“DULOS LATERAIS (Mod. 7) - Baseado no Excel
        # =================================================================
        lateral_data = [
            ('Módulo Lateral', 'Trash Cabinet SLP - Papeleira lateral SLP 30 litros', 'SLP-TCAB-01'),
            ('Módulo Lateral', 'Master Cabinet SLP - Armário lateral SLP', 'SLP-MCAB-02'),
        ]
        for mtype, desc, ref in lateral_data:
            short = ref.replace('SLP-', '')
            bd.execute('''INSERT INTO catalog_module_lateral 
                (module_type, description, reference, short_reference) VALUES (?, ?, ?, ?)''',
                (mtype, desc, ref, short))
        stats['lateral'] = len(lateral_data)
        
        # =================================================================
        # MÍ“DULOS ANTENA (Mod. 8) - Baseado no Excel
        # =================================================================
        antenna_data = [
            ('Módulo Antena', 'Antenna Enclosure for SLP Standard 3m (Í˜=212mm H=300mm)', 'SLPC-3MD212H300', 3),
            ('Módulo Antena', 'Antenna Enclosure for SLP Standard 4m (Í˜=196mm H=300mm)', 'SLP-4MCD196H300', 4),
            ('Módulo Antena', 'Antenna Enclosure for SLP Standard 4m (Í˜=196mm H=700mm)', 'SLPC-4MD196H700', 4),
            ('Módulo Antena', 'Antenna Enclosure for SLP Standard 4m (Í˜Base=196 Í˜=230mm H=800mm)', 'SLPC-4MD196/230H800', 4),
            ('Módulo Antena', 'Antenna Enclosure for SLP Standard 6m (Í˜=164mm H=300mm)', 'SLPC-6MD164H300', 6),
            ('Módulo Antena', 'Antenna Enclosure for SLP Standard 6m (Í˜=164mm H=700mm)', 'SLPC-6MD164H700', 6),
            ('Módulo Antena', 'Antenna Enclosure for SLP Standard 8m (Í˜=132mm H=300mm)', 'SLPC-8MD132H300', 8),
            ('Módulo Antena', 'Antenna Enclosure for SLP Standard 8m (Í˜=132mm H=700mm)', 'SLPC-8MD132H700', 8),
        ]
        for mtype, desc, ref, height in antenna_data:
            short = ref.replace('SLPC-', '').replace('SLP-', '')
            bd.execute('''INSERT INTO catalog_module_antenna 
                (module_type, description, reference, short_reference, column_height_m) VALUES (?, ?, ?, ?, ?)''',
                (mtype, desc, ref, short, height))
        stats['antenna'] = len(antenna_data)
        
        # =================================================================
        # COLUNAS - Gerar todas as combinações baseadas no Excel REFEMOD
        # =================================================================
        
        # Gerar colunas para cada pack
        packs_config = {
            'BAREBONE': {'code': 'B', 'mod2': 'Não', 'mod3': 'Não', 'mod4': 'Não', 'mod5': 'Não', 'mod6': 'Não'},
            'ESSENCIAL': {'code': 'E', 'mod2': 'Sim', 'mod3_rule': True, 'mod4': 'Não', 'mod5': 'Não', 'mod6': 'Não'},
            'CORE': {'code': 'C', 'mod2': 'Sim', 'mod3_rule': True, 'mod4': 'Não', 'mod5': 'Sim', 'mod6': 'Sim'},
            'PREMIUM': {'code': 'P', 'mod2': 'Sim', 'mod3_rule': True, 'mod4': 'Sim', 'mod5': 'Sim', 'mod6': 'Sim'},
        }
        
        heights = [3, 4, 6, 8, 10]
        fixings = [('Flange', 'F'), ('Embutido', 'E')]
        
        # Configurações de braços por altura
        arm_configs_by_height = {
            3: [(0, 0, 0, '0S'), (1, 1, 0, '1S')],
            4: [(0, 0, 0, '0S'), (1, 1, 0, '1S')],
            6: [(0, 0, 0, '0S'), (1, 1, 0, '1S'), (2, 2, 0, '2S')],
            8: [(0, 0, 0, '0S'), (1, 1, 0, '1S'), (2, 2, 0, '2S'), (2, 1, 1, '1S1P')],
            10: [(0, 0, 0, '0S'), (1, 1, 0, '1S'), (2, 2, 0, '2S'), (2, 1, 1, '1S1P')],
        }
        
        for pack_name, pack_cfg in packs_config.items():
            for height in heights:
                for fixing_name, fixing_code in fixings:
                    arm_configs = arm_configs_by_height.get(height, [(0, 0, 0, '0S')])
                    
                    for arm_count, arm_street, arm_sidewalk, arm_code in arm_configs:
                        # Determinar se tem luminária (S=Sem, L=Com)
                        # Colunas sem braço podem ter luminária tipo 2 (topo)
                        # Colunas com braço têm luminária tipo 1 (braço)
                        if arm_count == 0:
                            lum_code = 'S'  # Sem luminária por defeito, mas aceita Tipo 2
                            mod1_lum = 'Tipo 2' if height == 4 else 'Não'  # Só 4m tem luminária topo
                            lum_included = 'Não'
                        else:
                            lum_code = 'S'  # O código base é sem, o L vem quando se adiciona
                            mod1_lum = 'Tipo 1'
                            lum_included = 'Não'  # Por defeito é Não - o Excel dirá se é Sim
                        
                        # Código da referência: SLP-PXYZZ-MN
                        height_str = str(height).zfill(2)
                        ref = f"SLP-{pack_cfg['code']}S{fixing_code}{height_str}-{arm_code}{lum_code}"
                        
                        # Determinar mod3 (cofrete): só no ESSENCIAL e packs superiores, depende dos braços
                        if pack_cfg.get('mod3_rule'):
                            if arm_count == 1:
                                mod3 = 'Tipo S'
                            elif arm_count >= 2:
                                mod3 = 'Tipo D'
                            else:
                                mod3 = 'Não'
                        else:
                            mod3 = 'Não'
                        
                        # Descrição
                        desc_parts = [f"{pack_name} PACK"]
                        if arm_count == 0:
                            desc_parts.append("S/ LUMINÍRIA")
                        else:
                            desc_parts.append(f"{arm_count} Braço(s)")
                        desc_parts.append(f"- Coluna SLP {height}m")
                        desc_parts.append(f"({fixing_name})")
                        desc = " ".join(desc_parts)
                        
                        bd.execute('''INSERT OR REPLACE INTO catalog_columns 
                            (description, reference, pack, column_type, fixing, height_m,
                             arm_count, arm_street, arm_sidewalk, luminaire_included,
                             mod1_luminaire, mod2_electrical, mod3_fuse_box, mod4_telemetry,
                             mod5_ev, mod6_mupi, mod7_lateral, mod8_antenna)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                            (desc, ref, pack_name, 'Standard', fixing_name, height,
                             arm_count, arm_street, arm_sidewalk, lum_included,
                             mod1_lum, pack_cfg['mod2'], mod3, pack_cfg['mod4'],
                             pack_cfg['mod5'], pack_cfg['mod6'], 'Sim', 'Sim'))
                        stats['columns'] += 1
        
        bd.commit()
        
        return jsonify({
            'success': True,
            'message': 'Catálogo padrão carregado com sucesso',
            'stats': stats
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Erro ao carregar catálogo: {str(e)}'}), 500



# =============================================================================
# INICIALIZAÇÃO
# =============================================================================

# Inicialização automática para WSGI/Produção
# Isto garante que a BD é inicializada quando o app é importado
def _inicializar_para_producao():
    """Inicializa o sistema quando executado via WSGI."""
    tenant_db = obter_caminho_bd_tenant(MASTER_TENANT_ID)
    if os.path.exists(tenant_db):
        with app.app_context():
            inicializar_bd_tenant(MASTER_TENANT_ID)

# Executar inicialização
_inicializar_para_producao()

if __name__ == '__main__':
    print("🚀 SmartLamppost v4.0 - Sistema Multi-Tenant")
    print("=" * 55)
    
    # Verificar se o sistema foi inicializado
    tenant_db = obter_caminho_bd_tenant(MASTER_TENANT_ID)
    if not os.path.exists(tenant_db):
        print("⚠️  Sistema não inicializado!")
        print("   Execute primeiro: python init_system.py")
        print("=" * 55)
    else:
        print(f"📁 Pasta de tenants: {PASTA_TENANTS}")
        print(f"📁 Catálogo partilhado: {CATALOGO_PARTILHADO}")
        print(f"📦 Pasta de backups: {PASTA_BACKUPS}")
        print(f"📊 Exportação Excel: {'✅ Disponível' if EXCEL_DISPONIVEL else '❌ Indisponível'}")
        print("=" * 55)
        print("🌐 Servidor a iniciar em http://localhost:5000")
        print("👤 Credenciais: admin@smartlamppost.com / admin123")
        print("=" * 55)
    
    app.run(host='0.0.0.0', port=5000, debug=os.environ.get('FLASK_DEBUG', '0') == '1')
