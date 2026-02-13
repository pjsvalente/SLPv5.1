"""
SmartLamppost v5.0 - Data Import/Export Routes
Excel and JSON data management with multi-sheet support.
"""

import io
import json
import logging
from datetime import datetime
from flask import Blueprint, request, jsonify, send_file, g

from ...shared.database import obter_bd, obter_bd_catalogo
from ...shared.permissions import requer_admin, requer_autenticacao

logger = logging.getLogger(__name__)

data_bp = Blueprint('data', __name__)


# =========================================================================
# HELPER FUNCTIONS
# =========================================================================

def find_serial_column(headers):
    """
    Find the serial number column in Excel headers.
    Supports multiple naming conventions (PT, EN, variations).
    """
    serial_variants = [
        'número série', 'numero serie', 'nº série', 'nº serie',
        'serial number', 'serial_number', 'serialnumber',
        'n serie', 'nserie', 'num serie', 'numserie',
        'serial', 'série', 'serie'
    ]
    for h in headers:
        if h:
            normalized = str(h).lower().strip().replace('º', 'o').replace('_', ' ').replace('-', ' ')
            if normalized in serial_variants:
                return h
    return None


def find_status_column(headers):
    """
    Find the status column in Excel headers.
    Supports multiple naming conventions.
    """
    status_variants = ['status', 'estado', 'state', 'condition', 'situação', 'situacao']
    for h in headers:
        if h:
            normalized = str(h).lower().strip()
            if normalized in status_variants:
                return h
    return None


def style_header_row(ws, headers, row=1):
    """Apply consistent header styling."""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border


def auto_adjust_columns(ws):
    """Auto-adjust column widths based on content."""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


@data_bp.route('/export/excel', methods=['POST'])
@requer_admin
def export_excel():
    """
    Export assets to Excel format with multiple sheets.
    Sheets: Ativos, Histórico Estados, Intervenções, Atualizações
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return jsonify({'error': 'openpyxl não instalado. Execute: pip install openpyxl'}), 500

    dados = request.get_json() or {}
    selected_fields = dados.get('fields', [])  # Empty = all fields
    include_history = dados.get('include_history', True)
    include_interventions = dados.get('include_interventions', True)
    include_updates = dados.get('include_updates', True)
    asset_ids = dados.get('asset_ids', [])  # Empty = all assets

    bd = obter_bd()

    # Get schema fields
    schema_fields = bd.execute('SELECT * FROM schema_fields ORDER BY field_order').fetchall()
    field_map = {f['field_name']: f for f in schema_fields}

    # Filter fields if specified
    if selected_fields:
        schema_fields = [f for f in schema_fields if f['field_name'] in selected_fields]

    # Get assets (filtered or all)
    if asset_ids:
        placeholders = ','.join('?' * len(asset_ids))
        assets = bd.execute(
            f'SELECT * FROM assets WHERE id IN ({placeholders}) ORDER BY serial_number',
            asset_ids
        ).fetchall()
    else:
        assets = bd.execute('SELECT * FROM assets ORDER BY serial_number').fetchall()

    asset_id_list = [a['id'] for a in assets]

    # Create workbook
    wb = openpyxl.Workbook()

    # =========================================================================
    # SHEET 1: ATIVOS (Assets)
    # =========================================================================
    ws_assets = wb.active
    ws_assets.title = 'Ativos'

    # Build headers
    system_headers = ['ID', 'Número Série', 'Status', 'Data Criação', 'Última Atualização']
    dynamic_headers = [f['field_label'] for f in schema_fields]
    all_headers = system_headers + dynamic_headers

    style_header_row(ws_assets, all_headers)

    # Write data
    for row, asset in enumerate(assets, 2):
        asset_dict = dict(asset)
        dynamic_fields = {}
        if asset_dict.get('dynamic_fields'):
            try:
                dynamic_fields = json.loads(asset_dict['dynamic_fields'])
            except:
                pass

        ws_assets.cell(row=row, column=1, value=asset_dict['id'])
        ws_assets.cell(row=row, column=2, value=asset_dict['serial_number'])
        ws_assets.cell(row=row, column=3, value=asset_dict['status'])
        ws_assets.cell(row=row, column=4, value=asset_dict['created_at'])
        ws_assets.cell(row=row, column=5, value=asset_dict['updated_at'])

        for col, field in enumerate(schema_fields, 6):
            value = dynamic_fields.get(field['field_name'], '')
            ws_assets.cell(row=row, column=col, value=value)

    auto_adjust_columns(ws_assets)

    # =========================================================================
    # SHEET 2: HISTÓRICO ESTADOS (Status History)
    # =========================================================================
    if include_history and asset_id_list:
        ws_history = wb.create_sheet('Histórico Estados')

        # Check if table exists
        table_exists = bd.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='asset_status_history'"
        ).fetchone()

        if table_exists:
            placeholders = ','.join('?' * len(asset_id_list))
            history = bd.execute(f'''
                SELECT h.*, a.serial_number
                FROM asset_status_history h
                LEFT JOIN assets a ON h.asset_id = a.id
                WHERE h.asset_id IN ({placeholders})
                ORDER BY h.changed_at DESC
            ''', asset_id_list).fetchall()

            history_headers = ['ID', 'Número Série Ativo', 'Estado Anterior', 'Novo Estado',
                              'Motivo', 'Data Alteração', 'Alterado Por']
            style_header_row(ws_history, history_headers)

            for row, h in enumerate(history, 2):
                h_dict = dict(h)
                ws_history.cell(row=row, column=1, value=h_dict.get('id'))
                ws_history.cell(row=row, column=2, value=h_dict.get('serial_number', ''))
                ws_history.cell(row=row, column=3, value=h_dict.get('old_status', ''))
                ws_history.cell(row=row, column=4, value=h_dict.get('new_status', ''))
                ws_history.cell(row=row, column=5, value=h_dict.get('reason', ''))
                ws_history.cell(row=row, column=6, value=h_dict.get('changed_at', ''))
                ws_history.cell(row=row, column=7, value=h_dict.get('changed_by', ''))

            auto_adjust_columns(ws_history)

    # =========================================================================
    # SHEET 3: INTERVENÇÕES (Interventions)
    # =========================================================================
    if include_interventions and asset_id_list:
        ws_interventions = wb.create_sheet('Intervenções')

        placeholders = ','.join('?' * len(asset_id_list))
        interventions = bd.execute(f'''
            SELECT i.*, a.serial_number,
                   t.nome as technician_name
            FROM interventions i
            LEFT JOIN assets a ON i.asset_id = a.id
            LEFT JOIN technicians t ON i.technician_id = t.id
            WHERE i.asset_id IN ({placeholders})
            ORDER BY i.scheduled_date DESC
        ''', asset_id_list).fetchall()

        intervention_headers = [
            'ID', 'Número Série Ativo', 'Tipo', 'Descrição', 'Status',
            'Data Agendada', 'Data Conclusão', 'Técnico', 'Prioridade',
            'Custo Estimado', 'Custo Final', 'Notas', 'Data Criação'
        ]
        style_header_row(ws_interventions, intervention_headers)

        for row, intv in enumerate(interventions, 2):
            intv_dict = dict(intv)
            ws_interventions.cell(row=row, column=1, value=intv_dict.get('id'))
            ws_interventions.cell(row=row, column=2, value=intv_dict.get('serial_number', ''))
            ws_interventions.cell(row=row, column=3, value=intv_dict.get('intervention_type', ''))
            ws_interventions.cell(row=row, column=4, value=intv_dict.get('description', ''))
            ws_interventions.cell(row=row, column=5, value=intv_dict.get('status', ''))
            ws_interventions.cell(row=row, column=6, value=intv_dict.get('scheduled_date', ''))
            ws_interventions.cell(row=row, column=7, value=intv_dict.get('completed_date', ''))
            ws_interventions.cell(row=row, column=8, value=intv_dict.get('technician_name', ''))
            ws_interventions.cell(row=row, column=9, value=intv_dict.get('priority', ''))
            ws_interventions.cell(row=row, column=10, value=intv_dict.get('estimated_cost', ''))
            ws_interventions.cell(row=row, column=11, value=intv_dict.get('actual_cost', ''))
            ws_interventions.cell(row=row, column=12, value=intv_dict.get('notes', ''))
            ws_interventions.cell(row=row, column=13, value=intv_dict.get('created_at', ''))

        auto_adjust_columns(ws_interventions)

    # =========================================================================
    # SHEET 4: ATUALIZAÇÕES (Field Updates/Audit Log)
    # =========================================================================
    if include_updates and asset_id_list:
        ws_updates = wb.create_sheet('Atualizações')

        # Check if audit_log table exists
        table_exists = bd.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='audit_log'"
        ).fetchone()

        if table_exists:
            placeholders = ','.join('?' * len(asset_id_list))
            updates = bd.execute(f'''
                SELECT al.*, a.serial_number, u.nome as user_name
                FROM audit_log al
                LEFT JOIN assets a ON al.entity_id = a.id AND al.entity_type = 'asset'
                LEFT JOIN users u ON al.user_id = u.id
                WHERE al.entity_type = 'asset' AND al.entity_id IN ({placeholders})
                ORDER BY al.created_at DESC
            ''', asset_id_list).fetchall()

            update_headers = [
                'ID', 'Número Série Ativo', 'Ação', 'Campo Alterado',
                'Valor Anterior', 'Novo Valor', 'Utilizador', 'Data'
            ]
            style_header_row(ws_updates, update_headers)

            for row, upd in enumerate(updates, 2):
                upd_dict = dict(upd)
                ws_updates.cell(row=row, column=1, value=upd_dict.get('id'))
                ws_updates.cell(row=row, column=2, value=upd_dict.get('serial_number', ''))
                ws_updates.cell(row=row, column=3, value=upd_dict.get('action', ''))
                ws_updates.cell(row=row, column=4, value=upd_dict.get('field_name', ''))
                ws_updates.cell(row=row, column=5, value=upd_dict.get('old_value', ''))
                ws_updates.cell(row=row, column=6, value=upd_dict.get('new_value', ''))
                ws_updates.cell(row=row, column=7, value=upd_dict.get('user_name', ''))
                ws_updates.cell(row=row, column=8, value=upd_dict.get('created_at', ''))

            auto_adjust_columns(ws_updates)

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'ativos_export_{timestamp}.xlsx'

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@data_bp.route('/export/excel/simple', methods=['GET'])
@requer_admin
def export_excel_simple():
    """Simple GET endpoint for quick export of all assets."""
    try:
        import openpyxl
    except ImportError:
        return jsonify({'error': 'openpyxl não instalado'}), 500

    bd = obter_bd()
    assets = bd.execute('SELECT * FROM assets ORDER BY serial_number').fetchall()
    schema_fields = bd.execute('SELECT * FROM schema_fields ORDER BY field_order').fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ativos'

    headers = ['ID', 'Número Série', 'Status', 'Data Criação', 'Última Atualização']
    for field in schema_fields:
        headers.append(field['field_label'])

    style_header_row(ws, headers)

    for row, asset in enumerate(assets, 2):
        asset_dict = dict(asset)
        dynamic_fields = {}
        if asset_dict.get('dynamic_fields'):
            try:
                dynamic_fields = json.loads(asset_dict['dynamic_fields'])
            except:
                pass

        ws.cell(row=row, column=1, value=asset_dict['id'])
        ws.cell(row=row, column=2, value=asset_dict['serial_number'])
        ws.cell(row=row, column=3, value=asset_dict['status'])
        ws.cell(row=row, column=4, value=asset_dict['created_at'])
        ws.cell(row=row, column=5, value=asset_dict['updated_at'])

        for col, field in enumerate(schema_fields, 6):
            value = dynamic_fields.get(field['field_name'], '')
            ws.cell(row=row, column=col, value=value)

    auto_adjust_columns(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'ativos_export_{timestamp}.xlsx'

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@data_bp.route('/export/excel/fields', methods=['GET'])
@requer_admin
def get_export_fields():
    """Get available fields for export."""
    bd = obter_bd()

    schema_fields = bd.execute('SELECT * FROM schema_fields ORDER BY field_order').fetchall()

    fields = [
        {'name': 'id', 'label': 'ID', 'category': 'system'},
        {'name': 'serial_number', 'label': 'Número Série', 'category': 'system'},
        {'name': 'status', 'label': 'Status', 'category': 'system'},
        {'name': 'created_at', 'label': 'Data Criação', 'category': 'system'},
        {'name': 'updated_at', 'label': 'Última Atualização', 'category': 'system'},
    ]

    for field in schema_fields:
        fields.append({
            'name': field['field_name'],
            'label': field['field_label'],
            'category': field['field_category'] or 'other'
        })

    return jsonify({'fields': fields}), 200


@data_bp.route('/import/excel/template', methods=['GET'])
@requer_admin
def get_import_template():
    """Get Excel template for import."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({'error': 'openpyxl não instalado'}), 500

    bd = obter_bd()

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Template'

    # Get schema fields
    schema_fields = bd.execute('SELECT * FROM schema_fields ORDER BY field_order').fetchall()

    # Headers (without ID and dates)
    headers = ['Número Série', 'Status']
    for field in schema_fields:
        headers.append(field['field_label'])

    # Style for headers
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Add instruction sheet
    ws_inst = wb.create_sheet('Instruções')
    instructions = [
        ['Instruções de Importação'],
        [''],
        ['1. Preencha os dados na folha "Template"'],
        ['2. O campo "Número Série" é obrigatório e deve ser único'],
        ['3. Status válidos: ativo, inativo, manutencao, abatido'],
        ['4. Não modifique os cabeçalhos'],
        ['5. Guarde o ficheiro e importe no sistema'],
    ]
    for row, inst in enumerate(instructions, 1):
        ws_inst.cell(row=row, column=1, value=inst[0] if inst else '')

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name='template_importacao.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@data_bp.route('/import/excel/preview', methods=['POST'])
@requer_admin
def import_excel_preview():
    """
    Preview Excel import without committing changes.
    Returns parsed data with validation status.
    """
    try:
        import openpyxl
    except ImportError:
        return jsonify({'error': 'openpyxl não instalado'}), 500

    if 'file' not in request.files:
        return jsonify({'error': 'Ficheiro não fornecido'}), 400

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Formato de ficheiro inválido. Use .xlsx ou .xls'}), 400

    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
    except Exception as e:
        return jsonify({'error': f'Erro ao ler ficheiro: {str(e)}'}), 400

    # Wrap entire processing in try/except to ensure JSON response
    try:
        bd = obter_bd()

        # Get schema fields
        schema_fields = bd.execute('SELECT * FROM schema_fields ORDER BY field_order').fetchall()
        field_map = {f['field_label']: f['field_name'] for f in schema_fields}
        required_fields = [f['field_name'] for f in schema_fields if f.get('is_required')]

        # Get headers from first row and filter None values
        headers = [cell.value for cell in ws[1]]
        headers_clean = [h for h in headers if h is not None]

        if not headers_clean:
            return jsonify({'error': 'Ficheiro Excel sem cabeçalhos válidos na primeira linha'}), 400

        # Find serial number column (flexible naming)
        serial_col = find_serial_column(headers)
        if not serial_col:
            return jsonify({
                'error': 'Coluna de número de série não encontrada. Use um destes nomes: "Número Série", "Serial Number", "N Serie", "Serial"'
            }), 400

        # Find status column (flexible naming)
        status_col = find_status_column(headers)

        # Get existing serial numbers
        existing_serials = set()
        for row in bd.execute('SELECT serial_number FROM assets').fetchall():
            existing_serials.add(row['serial_number'])

        # Process rows for preview
        preview_rows = []
        stats = {'total': 0, 'new': 0, 'update': 0, 'errors': 0}

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row or all(cell is None for cell in row):
                continue

            stats['total'] += 1
            row_data = dict(zip(headers, row))

            # Use detected column names
            serial_number = str(row_data.get(serial_col, '') or '').strip()
            status = row_data.get(status_col, 'ativo') if status_col else 'ativo'
            if status is None:
                status = 'ativo'

            # Validate row
            errors = []
            warnings = []

            if not serial_number:
                errors.append('Número de série vazio')

            if status and status not in ['ativo', 'inativo', 'manutencao', 'abatido', 'suspenso']:
                warnings.append(f'Status "{status}" inválido, será convertido para "ativo"')

            # Check required fields
            for field in schema_fields:
                if field.get('is_required'):
                    label = field['field_label']
                    if label not in row_data or not row_data[label]:
                        errors.append(f'Campo obrigatório "{label}" vazio')

            # Build dynamic fields
            dynamic_fields = {}
            for label, field_name in field_map.items():
                if label in row_data and row_data[label] is not None:
                    dynamic_fields[field_name] = str(row_data[label])

            # Check if exists
            is_existing = serial_number in existing_serials
            action = 'update' if is_existing else 'create'

            if errors:
                stats['errors'] += 1
            elif is_existing:
                stats['update'] += 1
            else:
                stats['new'] += 1

            preview_rows.append({
                'row_num': row_num,
                'serial_number': serial_number,
                'status': status,
                'action': action,
                'is_valid': len(errors) == 0,
                'errors': errors,
                'warnings': warnings,
                'data': dynamic_fields
            })

            # Limit preview to first 100 rows
            if len(preview_rows) >= 100:
                break

        return jsonify({
            'headers': headers_clean,
            'rows': preview_rows,
            'stats': stats,
            'total_rows': stats['total'],
            'has_more': stats['total'] > 100,
            'detected_columns': {
                'serial': serial_col,
                'status': status_col
            }
        }), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Erro ao processar ficheiro: {str(e)}'}), 500


@data_bp.route('/import/excel', methods=['POST'])
@requer_admin
def import_excel():
    """
    Import assets from Excel file with mode selection.
    Modes:
    - 'create_only': Only insert new records, skip existing
    - 'update_only': Only update existing records, skip new
    - 'upsert': Insert new and update existing (default)
    """
    try:
        import openpyxl
    except ImportError:
        return jsonify({'error': 'openpyxl não instalado'}), 500

    if 'file' not in request.files:
        return jsonify({'error': 'Ficheiro não fornecido'}), 400

    file = request.files['file']
    mode = request.form.get('mode', 'upsert')  # create_only, update_only, upsert
    convert_suspended = request.form.get('convert_suspended', 'true') == 'true'

    if mode not in ['create_only', 'update_only', 'upsert']:
        return jsonify({'error': 'Modo inválido. Use: create_only, update_only, upsert'}), 400

    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Formato de ficheiro inválido. Use .xlsx ou .xls'}), 400

    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
    except Exception as e:
        return jsonify({'error': f'Erro ao ler ficheiro: {str(e)}'}), 400

    # Wrap entire processing in try/except to ensure JSON response
    try:
        bd = obter_bd()

        # Get schema fields
        schema_fields = bd.execute('SELECT * FROM schema_fields ORDER BY field_order').fetchall()
        field_map = {f['field_label']: f['field_name'] for f in schema_fields}

        # Get headers from first row
        headers = [cell.value for cell in ws[1]]
        headers_clean = [h for h in headers if h is not None]

        if not headers_clean:
            return jsonify({'error': 'Ficheiro Excel sem cabeçalhos válidos na primeira linha'}), 400

        # Find serial number column (flexible naming)
        serial_col = find_serial_column(headers)
        if not serial_col:
            return jsonify({
                'error': 'Coluna de número de série não encontrada. Use um destes nomes: "Número Série", "Serial Number", "N Serie", "Serial"'
            }), 400

        # Find status column (flexible naming)
        status_col = find_status_column(headers)

        # Process rows
        imported = 0
        updated = 0
        skipped = 0
        errors = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row or all(cell is None for cell in row):
                continue

            row_data = dict(zip(headers, row))

            # Use detected column names
            serial_number = str(row_data.get(serial_col, '') or '').strip()
            if not serial_number:
                errors.append(f'Linha {row_num}: Número de série vazio')
                continue

            status = row_data.get(status_col, 'ativo') if status_col else 'ativo'
            if status is None:
                status = 'ativo'

            # Convert invalid/suspended status to 'ativo'
            valid_statuses = ['ativo', 'inativo', 'manutencao', 'abatido']
            if status not in valid_statuses:
                if convert_suspended and status == 'suspenso':
                    status = 'ativo'  # Convert suspended to active as per RFID v3 spec
                else:
                    status = 'ativo'

            # Build dynamic fields
            dynamic_fields = {}
            for label, field_name in field_map.items():
                if label in row_data and row_data[label] is not None:
                    dynamic_fields[field_name] = str(row_data[label])

            # Check if exists
            existing = bd.execute(
                'SELECT id FROM assets WHERE serial_number = ?',
                (serial_number,)
            ).fetchone()

            try:
                if existing:
                    # Handle existing record
                    if mode == 'create_only':
                        skipped += 1
                        continue

                    # Update - get the asset_id first
                    asset_row = bd.execute(
                        'SELECT id FROM assets WHERE serial_number = ?',
                        (serial_number,)
                    ).fetchone()
                    asset_id = asset_row['id']

                    # Update the asset record
                    bd.execute('''
                        UPDATE assets SET updated_at = CURRENT_TIMESTAMP
                        WHERE serial_number = ?
                    ''', (serial_number,))

                    # Update dynamic fields in asset_data table
                    for field_name, field_value in dynamic_fields.items():
                        bd.execute('''
                            INSERT OR REPLACE INTO asset_data (asset_id, field_name, field_value)
                            VALUES (?, ?, ?)
                        ''', (asset_id, field_name, str(field_value) if field_value is not None else None))

                    # Store status as condition_status
                    if status:
                        bd.execute('''
                            INSERT OR REPLACE INTO asset_data (asset_id, field_name, field_value)
                            VALUES (?, 'condition_status', ?)
                        ''', (asset_id, status))

                    updated += 1
                else:
                    # Handle new record
                    if mode == 'update_only':
                        skipped += 1
                        continue

                    # Insert into assets table
                    cursor = bd.execute('''
                        INSERT INTO assets (serial_number, created_at)
                        VALUES (?, CURRENT_TIMESTAMP)
                    ''', (serial_number,))
                    asset_id = cursor.lastrowid

                    # Insert dynamic fields into asset_data table
                    for field_name, field_value in dynamic_fields.items():
                        bd.execute('''
                            INSERT OR REPLACE INTO asset_data (asset_id, field_name, field_value)
                            VALUES (?, ?, ?)
                        ''', (asset_id, field_name, str(field_value) if field_value is not None else None))

                    # Store status as condition_status
                    if status:
                        bd.execute('''
                            INSERT OR REPLACE INTO asset_data (asset_id, field_name, field_value)
                            VALUES (?, 'condition_status', ?)
                        ''', (asset_id, status))

                    imported += 1
            except Exception as e:
                errors.append(f'Linha {row_num}: {str(e)}')

        bd.commit()

        return jsonify({
            'message': 'Importação concluída',
            'mode': mode,
            'imported': imported,
            'updated': updated,
            'skipped': skipped,
            'errors': errors[:20]  # Limit errors shown
        }), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Erro ao processar importação: {str(e)}'}), 500


@data_bp.route('/export/json', methods=['POST'])
@requer_admin
def export_json():
    """Export data to JSON format."""
    dados = request.get_json() or {}
    include_assets = dados.get('assets', True)
    include_interventions = dados.get('interventions', True)
    include_technicians = dados.get('technicians', True)

    bd = obter_bd()
    export_data = {
        'version': '5.0',
        'exported_at': datetime.now().isoformat(),
        'tenant': 'smartlamppost'
    }

    if include_assets:
        assets = bd.execute('SELECT * FROM assets').fetchall()
        export_data['assets'] = [dict(a) for a in assets]

    if include_interventions:
        interventions = bd.execute('SELECT * FROM interventions').fetchall()
        export_data['interventions'] = [dict(i) for i in interventions]

    if include_technicians:
        technicians = bd.execute('SELECT * FROM technicians').fetchall()
        export_data['technicians'] = [dict(t) for t in technicians]

    buffer = io.BytesIO()
    buffer.write(json.dumps(export_data, indent=2, default=str).encode('utf-8'))
    buffer.seek(0)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'backup_data_{timestamp}.json'

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/json'
    )


@data_bp.route('/import/json', methods=['POST'])
@requer_admin
def import_json():
    """Import data from JSON file."""
    if 'file' not in request.files:
        return jsonify({'error': 'Ficheiro não fornecido'}), 400

    file = request.files['file']
    if not file.filename.endswith('.json'):
        return jsonify({'error': 'Formato de ficheiro inválido. Use .json'}), 400

    try:
        data = json.load(file)
    except Exception as e:
        return jsonify({'error': f'Erro ao ler JSON: {str(e)}'}), 400

    bd = obter_bd()
    results = {'assets': 0, 'interventions': 0, 'technicians': 0}

    # Import assets
    if 'assets' in data:
        for asset in data['assets']:
            try:
                existing = bd.execute(
                    'SELECT id FROM assets WHERE serial_number = ?',
                    (asset.get('serial_number'),)
                ).fetchone()

                if not existing:
                    bd.execute('''
                        INSERT INTO assets (serial_number, status, dynamic_fields)
                        VALUES (?, ?, ?)
                    ''', (
                        asset.get('serial_number'),
                        asset.get('status', 'ativo'),
                        asset.get('dynamic_fields', '{}')
                    ))
                    results['assets'] += 1
            except Exception as e:
                logger.error(f"Error importing asset: {e}")

    # Import technicians
    if 'technicians' in data:
        for tech in data['technicians']:
            try:
                existing = bd.execute(
                    'SELECT id FROM technicians WHERE nome = ? AND tipo = ?',
                    (tech.get('nome'), tech.get('tipo'))
                ).fetchone()

                if not existing:
                    bd.execute('''
                        INSERT INTO technicians (nome, tipo, empresa, telefone, email, especialidade, notas, ativo)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        tech.get('nome'),
                        tech.get('tipo'),
                        tech.get('empresa', ''),
                        tech.get('telefone', ''),
                        tech.get('email', ''),
                        tech.get('especialidade', ''),
                        tech.get('notas', ''),
                        tech.get('ativo', 1)
                    ))
                    results['technicians'] += 1
            except Exception as e:
                logger.error(f"Error importing technician: {e}")

    bd.commit()

    return jsonify({
        'message': 'Importação concluída',
        'imported': results
    }), 200


@data_bp.route('/stats', methods=['GET'])
@requer_admin
def get_data_stats():
    """Get data statistics."""
    bd = obter_bd()

    stats = {
        'assets': bd.execute('SELECT COUNT(*) FROM assets').fetchone()[0],
        'interventions': bd.execute('SELECT COUNT(*) FROM interventions').fetchone()[0],
        'technicians': bd.execute('SELECT COUNT(*) FROM technicians').fetchone()[0],
        'users': bd.execute('SELECT COUNT(*) FROM users').fetchone()[0],
    }

    return jsonify(stats), 200


# =========================================================================
# CATALOG EXPORT/IMPORT (catalog.db)
# =========================================================================

@data_bp.route('/export/catalog', methods=['GET'])
@requer_admin
def export_catalog():
    """
    Export entire catalog to Excel with multiple sheets.
    One sheet per module type + columns + packs.
    """
    try:
        import openpyxl
    except ImportError:
        return jsonify({'error': 'openpyxl não instalado'}), 500

    bd = obter_bd_catalogo()

    wb = openpyxl.Workbook()

    # Define catalog tables and their headers
    catalog_tables = [
        ('catalog_packs', 'Packs', [
            'id', 'pack_name', 'pack_code', 'description', 'total_power_watts',
            'is_active', 'created_at'
        ]),
        ('catalog_columns', 'Colunas', [
            'id', 'column_name', 'column_code', 'manufacturer', 'model',
            'height_meters', 'material', 'max_power_watts', 'connection_type',
            'mod1_luminaire', 'mod2_electrical_panel', 'mod3_fuse_box',
            'mod4_telemetry', 'mod5_ev_charger', 'mod6_mupi', 'mod7_lateral', 'mod8_antenna',
            'is_active', 'created_at'
        ]),
        ('catalog_luminaires', 'Luminárias', [
            'id', 'luminaire_name', 'luminaire_code', 'manufacturer', 'model',
            'power_watts', 'luminous_flux_lm', 'color_temperature_k', 'ip_rating',
            'is_active', 'created_at'
        ]),
        ('catalog_electrical_panels', 'Quadros Elétricos', [
            'id', 'panel_name', 'panel_code', 'manufacturer', 'model',
            'max_power_watts', 'num_circuits', 'ip_rating',
            'is_active', 'created_at'
        ]),
        ('catalog_fuse_boxes', 'Cofretes', [
            'id', 'fuse_box_name', 'fuse_box_code', 'manufacturer', 'model',
            'num_fuses', 'max_amps', 'ip_rating',
            'is_active', 'created_at'
        ]),
        ('catalog_telemetry_panels', 'Telemetria', [
            'id', 'panel_name', 'panel_code', 'manufacturer', 'model',
            'power_watts', 'communication_type', 'ip_rating',
            'is_active', 'created_at'
        ]),
        ('catalog_module_ev', 'Carregadores EV', [
            'id', 'module_name', 'module_code', 'manufacturer', 'model',
            'power_watts', 'connector_type', 'num_connectors',
            'is_active', 'created_at'
        ]),
        ('catalog_module_mupi', 'MUPI', [
            'id', 'module_name', 'module_code', 'manufacturer', 'model',
            'power_watts', 'screen_size_inches', 'resolution',
            'is_active', 'created_at'
        ]),
        ('catalog_module_lateral', 'Módulos Laterais', [
            'id', 'module_name', 'module_code', 'manufacturer', 'model',
            'power_watts', 'module_type', 'description',
            'is_active', 'created_at'
        ]),
        ('catalog_module_antenna', 'Antenas', [
            'id', 'module_name', 'module_code', 'manufacturer', 'model',
            'power_watts', 'frequency_band', 'antenna_type',
            'is_active', 'created_at'
        ]),
    ]

    first_sheet = True
    for table_name, sheet_name, headers in catalog_tables:
        # Check if table exists
        table_exists = bd.execute(
            f"SELECT name FROM sqlite_master WHERE type='table' AND name='{table_name}'"
        ).fetchone()

        if not table_exists:
            continue

        if first_sheet:
            ws = wb.active
            ws.title = sheet_name
            first_sheet = False
        else:
            ws = wb.create_sheet(sheet_name)

        style_header_row(ws, headers)

        # Get data
        rows = bd.execute(f'SELECT * FROM {table_name}').fetchall()

        for row_num, row in enumerate(rows, 2):
            row_dict = dict(row)
            for col_num, header in enumerate(headers, 1):
                value = row_dict.get(header, '')
                ws.cell(row=row_num, column=col_num, value=value)

        auto_adjust_columns(ws)

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'catalogo_export_{timestamp}.xlsx'

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@data_bp.route('/import/catalog', methods=['POST'])
@requer_admin
def import_catalog():
    """
    Import catalog data from Excel file.
    Each sheet corresponds to a catalog table.
    """
    try:
        import openpyxl
    except ImportError:
        return jsonify({'error': 'openpyxl não instalado'}), 500

    if 'file' not in request.files:
        return jsonify({'error': 'Ficheiro não fornecido'}), 400

    file = request.files['file']
    mode = request.form.get('mode', 'upsert')  # create_only, update_only, upsert

    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Formato de ficheiro inválido'}), 400

    try:
        wb = openpyxl.load_workbook(file)
    except Exception as e:
        return jsonify({'error': f'Erro ao ler ficheiro: {str(e)}'}), 400

    bd = obter_bd_catalogo()

    # Sheet name to table mapping
    sheet_to_table = {
        'Packs': ('catalog_packs', 'pack_code'),
        'Colunas': ('catalog_columns', 'column_code'),
        'Luminárias': ('catalog_luminaires', 'luminaire_code'),
        'Quadros Elétricos': ('catalog_electrical_panels', 'panel_code'),
        'Cofretes': ('catalog_fuse_boxes', 'fuse_box_code'),
        'Telemetria': ('catalog_telemetry_panels', 'panel_code'),
        'Carregadores EV': ('catalog_module_ev', 'module_code'),
        'MUPI': ('catalog_module_mupi', 'module_code'),
        'Módulos Laterais': ('catalog_module_lateral', 'module_code'),
        'Antenas': ('catalog_module_antenna', 'module_code'),
    }

    results = {}

    for sheet_name in wb.sheetnames:
        if sheet_name not in sheet_to_table:
            continue

        table_name, code_field = sheet_to_table[sheet_name]
        ws = wb[sheet_name]

        # Check if table exists
        table_exists = bd.execute(
            f"SELECT name FROM sqlite_master WHERE type='table' AND name='{table_name}'"
        ).fetchone()

        if not table_exists:
            results[sheet_name] = {'error': 'Tabela não existe'}
            continue

        # Get headers
        headers = [cell.value for cell in ws[1]]
        if 'id' in headers:
            headers.remove('id')  # Don't import ID

        imported = 0
        updated = 0
        skipped = 0
        errors = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row:
                continue

            row_data = dict(zip([cell.value for cell in ws[1]], row))

            # Get unique code field
            code_value = row_data.get(code_field)
            if not code_value:
                errors.append(f'Linha {row_num}: Código vazio')
                continue

            # Check if exists
            existing = bd.execute(
                f'SELECT id FROM {table_name} WHERE {code_field} = ?',
                (code_value,)
            ).fetchone()

            try:
                # Build field lists (excluding id and timestamps)
                fields = [h for h in headers if h not in ['id', 'created_at', 'updated_at']]
                values = [row_data.get(f) for f in fields]

                if existing:
                    if mode == 'create_only':
                        skipped += 1
                        continue

                    # Update
                    set_clause = ', '.join([f'{f} = ?' for f in fields])
                    bd.execute(
                        f'UPDATE {table_name} SET {set_clause} WHERE {code_field} = ?',
                        values + [code_value]
                    )
                    updated += 1
                else:
                    if mode == 'update_only':
                        skipped += 1
                        continue

                    # Insert
                    placeholders = ', '.join(['?' for _ in fields])
                    field_names = ', '.join(fields)
                    bd.execute(
                        f'INSERT INTO {table_name} ({field_names}) VALUES ({placeholders})',
                        values
                    )
                    imported += 1

            except Exception as e:
                errors.append(f'Linha {row_num}: {str(e)}')

        results[sheet_name] = {
            'imported': imported,
            'updated': updated,
            'skipped': skipped,
            'errors': errors[:10]
        }

    bd.commit()

    return jsonify({
        'message': 'Importação do catálogo concluída',
        'mode': mode,
        'results': results
    }), 200


@data_bp.route('/export/catalog/template', methods=['GET'])
@requer_admin
def get_catalog_template():
    """Get Excel template for catalog import."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return jsonify({'error': 'openpyxl não instalado'}), 500

    wb = openpyxl.Workbook()

    # Define template sheets
    templates = [
        ('Colunas', [
            'column_name', 'column_code', 'manufacturer', 'model',
            'height_meters', 'material', 'max_power_watts', 'connection_type',
            'mod1_luminaire', 'mod2_electrical_panel', 'mod3_fuse_box',
            'mod4_telemetry', 'mod5_ev_charger', 'mod6_mupi', 'mod7_lateral', 'mod8_antenna',
            'is_active'
        ]),
        ('Luminárias', [
            'luminaire_name', 'luminaire_code', 'manufacturer', 'model',
            'power_watts', 'luminous_flux_lm', 'color_temperature_k', 'ip_rating', 'is_active'
        ]),
        ('Quadros Elétricos', [
            'panel_name', 'panel_code', 'manufacturer', 'model',
            'max_power_watts', 'num_circuits', 'ip_rating', 'is_active'
        ]),
        ('Cofretes', [
            'fuse_box_name', 'fuse_box_code', 'manufacturer', 'model',
            'num_fuses', 'max_amps', 'ip_rating', 'is_active'
        ]),
        ('Telemetria', [
            'panel_name', 'panel_code', 'manufacturer', 'model',
            'power_watts', 'communication_type', 'ip_rating', 'is_active'
        ]),
        ('Carregadores EV', [
            'module_name', 'module_code', 'manufacturer', 'model',
            'power_watts', 'connector_type', 'num_connectors', 'is_active'
        ]),
        ('MUPI', [
            'module_name', 'module_code', 'manufacturer', 'model',
            'power_watts', 'screen_size_inches', 'resolution', 'is_active'
        ]),
        ('Módulos Laterais', [
            'module_name', 'module_code', 'manufacturer', 'model',
            'power_watts', 'module_type', 'description', 'is_active'
        ]),
        ('Antenas', [
            'module_name', 'module_code', 'manufacturer', 'model',
            'power_watts', 'frequency_band', 'antenna_type', 'is_active'
        ]),
    ]

    first_sheet = True
    for sheet_name, headers in templates:
        if first_sheet:
            ws = wb.active
            ws.title = sheet_name
            first_sheet = False
        else:
            ws = wb.create_sheet(sheet_name)

        style_header_row(ws, headers)
        auto_adjust_columns(ws)

    # Add instructions sheet
    ws_inst = wb.create_sheet('Instruções')
    instructions = [
        ['Instruções de Importação do Catálogo'],
        [''],
        ['1. Cada folha corresponde a um tipo de módulo'],
        ['2. Os campos *_code devem ser únicos'],
        ['3. is_active: 1 = ativo, 0 = inativo'],
        ['4. Campos mod1-mod8 em Colunas: 1 = compatível, 0 = não compatível'],
        ['5. Não modifique os cabeçalhos'],
    ]
    for row, inst in enumerate(instructions, 1):
        ws_inst.cell(row=row, column=1, value=inst[0] if inst else '')

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name='template_catalogo.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
